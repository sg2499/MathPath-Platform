import json
import os
import re
import secrets
import shutil
import socket
from datetime import datetime, timezone as datetime_timezone, timedelta, tzinfo
from zoneinfo import ZoneInfo
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.errors import api_error
from app.core.config import (
    TEMPORARY_ASSESSMENT_READINESS_BYPASS,
    ASSESSMENT_READINESS_GATE_MODE,
    ASSESSMENT_READINESS_GATE_LABEL,
    ASSESSMENT_TESTING_OVERRIDE_ENABLED,
    ASSESSMENT_TESTING_OVERRIDE_LABEL,
)
from app.core.security import hash_password
from app.core.rate_limit import limiter
from app.database import SessionLocal, get_db
from app.dependencies import require_roles
from app.models import User, Module, Level, Lesson, DPS, Assignment, Attempt, AttemptAnswer, GeneratedQuestionSet, GeneratedQuestion, QuestionOption, Student, Teacher, Batch, StudentBatch, Notification, AssignmentReattemptPermission, AssessmentBlueprint, AssessmentBlueprintLesson, AssessmentVersion, AssessmentAssignment, AssessmentAttempt, AssessmentResult, AssessmentReattemptApproval, AssessmentAttemptAnswer, StudentLevelPromotion, ParentReportEmailLog, AssessmentReadinessTestingOverride, AuditLog, CompetitionMockExam, CompetitionMockAssignment
from app.services.assignment_service import create_assignment
from app.services.attempt_service import result_payload
from app.services.reattempt_operational_service import CountNeedsReattemptConcepts, ClearedConceptAttempts, CurrentOperationalAttempts, NeedsReattemptAttempts
from app.services.curriculum_service import dps_config_payload, get_dps_or_404
from app.services.generation_service import build_preview_seed, generate_preview
from app.services.assessment_eligibility_service import assessment_eligibility_payload, eligibility_for_students
from app.services.auth_service import public_profile_photo_url, force_logout_user
from app.services.report_export_service import BuildWorkbookResponse, BuildParentProgressPdfResponse, BuildParentProgressPdfBytes, ReportGeneratedOn
from app.services.email_service import (
    DiagnoseSmtpConfiguration,
    EmailConfigurationError,
    EmailSendError,
    SendEmailWithAttachment,
)
from app.services.assessment_blueprint_service import (
    archive_blueprint,
    blueprint_payload,
    create_blueprint,
    delete_blueprint,
    is_section_wise_module,
    level_section_registry_config,
    list_blueprints,
    publish_blueprint,
    section_marks_metadata,
    update_blueprint,
)

from app.services.assessment_engine_service import (
    ApproveAssessmentReattempt,
    AssessmentAssignmentPayload,
    AssessmentResultPayload,
    NormalizeAssessmentScore,
    NormalizeAssessmentPercentage,
    AssessmentEngineFoundation,
    AssessmentReattemptApprovalPayload,
    BlueprintEngineState,
    GenerateAssessmentPreview,
    LatestGeneratedVersion,
    ListAssessmentReattemptApprovals,
    ListStudentLevelPromotions,
    PromoteAssessmentStudentToNextLevel,
    RejectAssessmentReattempt,
    SetAssessmentVersionAvailability,
)

from app.services.assessment_notification_service import NotifyAssessmentReattemptDecision, NotifyStudentPromoted
from app.services.practice_notification_service import NotifyPracticeFreshPracticeAssigned
from app.services.manual_intervention_service import BuildManualInterventionQueue, BuildManualRetryAssignment, MANUAL_INTERVENTION_STATUS
from app.services.assessment_feedback_service import upsert_assessment_remark, delete_assessment_remark, assessment_feedback_payload, active_assessment_remark
from app.services.parent_report_notification_service import (
    NotifyParentReportGenerated,
    NotifyParentReportDeliveryLogs,
    NotifyParentReportDeliveryDeleted,
)
from app.services.competition_mock_generation_service import (
    GenerateCompetitionMockDraft,
    ListCompetitionMockDrafts,
    CompetitionMockExamPayload,
    CompetitionMockSectionPlan,
    DeleteCompetitionMockExam,
    ArchiveCompetitionMockExam,
)

from app.services.competition_mock_assignment_service import (
    AssignCompetitionMockExams,
    ListCompetitionMockAssignments,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])
admin_dep = require_roles("SUPER_ADMIN", "ADMIN")



class CompetitionMockGenerateRequest(BaseModel):
    levelId: str
    title: str | None = None
    mockCode: str | None = None
    totalQuestions: int | None = None
    durationSeconds: int | None = None
    competitionScope: str | None = "GENERAL"
    difficultyBand: str | None = "COMPETITION"
    sectionCounts: dict[str, int] | None = None


class CompetitionMockAssignRequest(BaseModel):
    levelId: str
    mockExamIds: list[str]
    studentIds: list[str] | None = None
    assignToAllInLevel: bool = False
    maxAttempts: int | None = 1
    dueAt: str | None = None
    instructions: str | None = None

class AssessmentRemarkRequest(BaseModel):
    remarkText: str


def _admin_natural_sort_key(value: Any) -> list[Any]:
    text = str(value or "").strip().lower()
    parts = re.split(r"(\d+)", text)
    return [int(part) if part.isdigit() else part for part in parts]



def AssessmentReadinessGateAuditPayload(rows: list[dict[str, Any]]) -> dict[str, Any]:
    NotReadyRows = [Row for Row in rows if not Row.get("eligible")]
    return {
        "mode": ASSESSMENT_READINESS_GATE_MODE,
        "label": ASSESSMENT_READINESS_GATE_LABEL,
        "temporaryBypassEnabled": bool(TEMPORARY_ASSESSMENT_READINESS_BYPASS),
        "strictReadinessActive": not bool(TEMPORARY_ASSESSMENT_READINESS_BYPASS),
        "testingOverrideEnabled": bool(ASSESSMENT_TESTING_OVERRIDE_ENABLED),
        "testingOverrideLabel": ASSESSMENT_TESTING_OVERRIDE_LABEL,
        "notReadyStudentsImpacted": len(NotReadyRows),
        "assignmentImpactLabel": (
            "Testing bypass currently allows assessment assignment before readiness is complete."
            if TEMPORARY_ASSESSMENT_READINESS_BYPASS
            else "Strict readiness gate currently blocks assessment assignment until eligibility is complete."
        ),
        "nextPhaseNote": (
            "Global testing bypass is active. Use only for broad local testing; set TEMPORARY_ASSESSMENT_READINESS_BYPASS=false before live deployment."
            if TEMPORARY_ASSESSMENT_READINESS_BYPASS
            else (
                "Strict readiness gate is active with Admin Testing Override available for controlled QA/demo use."
                if ASSESSMENT_TESTING_OVERRIDE_ENABLED
                else "Strict readiness gate is active and Admin Testing Override is disabled. This is the recommended live-ready configuration."
            )
        ),
    }


def _admin_assessment_testing_override_payload(db: Session, OverrideValue: AssessmentReadinessTestingOverride) -> dict[str, Any]:
    EnabledBy = db.get(User, OverrideValue.enabled_by_user_id) if OverrideValue.enabled_by_user_id else None
    DisabledBy = db.get(User, OverrideValue.disabled_by_user_id) if OverrideValue.disabled_by_user_id else None
    return {
        "id": OverrideValue.id,
        "studentId": OverrideValue.student_id,
        "studentCode": OverrideValue.student_code,
        "moduleId": OverrideValue.module_id,
        "moduleCode": OverrideValue.module_code,
        "moduleName": OverrideValue.module_name,
        "levelId": OverrideValue.level_id,
        "levelCode": OverrideValue.level_code,
        "levelName": OverrideValue.level_name,
        "status": OverrideValue.status,
        "isActive": str(OverrideValue.status or "").upper() == "ACTIVE",
        "reason": OverrideValue.reason,
        "enabledByUserId": OverrideValue.enabled_by_user_id,
        "enabledBy": EnabledBy.full_name if EnabledBy else None,
        "enabledAt": OverrideValue.enabled_at.isoformat() if OverrideValue.enabled_at else None,
        "disabledByUserId": OverrideValue.disabled_by_user_id,
        "disabledBy": DisabledBy.full_name if DisabledBy else None,
        "disabledAt": OverrideValue.disabled_at.isoformat() if OverrideValue.disabled_at else None,
        "usedForAssessmentAssignmentId": OverrideValue.used_for_assessment_assignment_id,
        "usedAt": OverrideValue.used_at.isoformat() if OverrideValue.used_at else None,
        "createdAt": OverrideValue.created_at.isoformat() if OverrideValue.created_at else None,
        "updatedAt": OverrideValue.updated_at.isoformat() if OverrideValue.updated_at else None,
    }


def _admin_active_assessment_testing_override(db: Session, *, StudentId: str, LevelId: str, ModuleId: str | None = None) -> AssessmentReadinessTestingOverride | None:
    QueryValue = db.query(AssessmentReadinessTestingOverride).filter(
        AssessmentReadinessTestingOverride.student_id == StudentId,
        AssessmentReadinessTestingOverride.level_id == LevelId,
        AssessmentReadinessTestingOverride.status == "ACTIVE",
    )
    if ModuleId:
        QueryValue = QueryValue.filter(AssessmentReadinessTestingOverride.module_id == ModuleId)
    return QueryValue.order_by(AssessmentReadinessTestingOverride.enabled_at.desc()).first()


def _admin_testing_override_scope_payload(db: Session, *, StudentValue: Student, LevelValue: Level) -> tuple[Module | None, str | None, str | None]:
    ModuleValue = db.get(Module, LevelValue.module_id) if LevelValue.module_id else None
    ModuleId = ModuleValue.id if ModuleValue else getattr(StudentValue, "current_module_id", None)
    ModuleCode = ModuleValue.module_code if ModuleValue else None
    ModuleName = ModuleValue.module_name if ModuleValue else None
    return ModuleValue, ModuleCode, ModuleName

class AssessmentReattemptDecisionRequest(BaseModel):
    adminNote: str | None = None


class AssessmentPromotionRequest(BaseModel):
    targetLevelId: str | None = None
    targetLevelCode: str | None = None


class AssessmentTestingOverrideCreateRequest(BaseModel):
    studentId: str
    moduleId: str | None = None
    levelId: str
    reason: str | None = None


class AssessmentTestingOverrideDeactivateRequest(BaseModel):
    reason: str | None = None


class ParentReportEmailRequest(BaseModel):
    studentId: str
    moduleId: str | None = None
    levelId: str | None = None
    lessonId: str | None = None
    dpsId: str | None = None
    recipientMode: str
    customEmail: str | None = None
    timezone: str | None = None
    timezoneOffsetMinutes: int | None = None


class ParentReportResendRequest(BaseModel):
    recipientMode: str | None = "SAME"
    customEmail: str | None = None

UPLOAD_ROOT = Path("uploads/students")
PHOTO_DIR = UPLOAD_ROOT / "photos"
SIGNATURE_DIR = UPLOAD_ROOT / "signatures"
PHOTO_DIR.mkdir(parents=True, exist_ok=True)
SIGNATURE_DIR.mkdir(parents=True, exist_ok=True)

TEACHER_UPLOAD_ROOT = Path("uploads/teachers")
TEACHER_PHOTO_DIR = TEACHER_UPLOAD_ROOT / "photos"
TEACHER_SIGNATURE_DIR = TEACHER_UPLOAD_ROOT / "signatures"
TEACHER_PHOTO_DIR.mkdir(parents=True, exist_ok=True)
TEACHER_SIGNATURE_DIR.mkdir(parents=True, exist_ok=True)

MANDATORY_BULK_FIELDS = [
    "custom_id",
    "student_name",
    "dob",
    "gender",
    "school_name",
    "father_name",
    "father_mobile",
    "father_whatsapp",
    "mother_name",
    "mother_mobile",
    "mother_whatsapp",
    "student_code",
    "password",
    "module_code",
    "level_code",
    "status",
]

TEMPLATE_FIELDS = MANDATORY_BULK_FIELDS + [
    "teacher",
    "teacher_code",
    "admission_date",
    "blood_group",
    "class",
    "section",
    "father_email",
    "mother_email",
    "interest",
    "present_address",
    "permanent_address",
    "school_area",
    "father_occupation",
    "mother_occupation",
]


class PreviewRequest(BaseModel):
    questionCount: int | None = None
    seed: str | None = None


class AllowReattemptRequest(BaseModel):
    reason: str | None = None


class AssignmentRequest(BaseModel):
    assignmentType: str = "PRACTICE"
    dpsId: str
    assignedToType: str
    assignedToId: str
    title: str
    instructions: str | None = None
    allowReattempt: bool = False


class AssignmentStatusRequest(BaseModel):
    isActive: bool


class AssessmentLessonDistributionRequest(BaseModel):
    """Distribution row for an assessment blueprint. Doubles as both shapes:
    lessonId for YLM's original lesson-wise distribution, sectionKey for
    IM/MM's section-wise distribution (2026-07-22) -- exactly one of the two
    is populated depending on the target module, and
    assessment_blueprint_service.py's validators only ever read the one that
    applies. Both optional here so a single request schema (and a single
    admin UI payload shape) works for either mode without the frontend
    needing to know module-specific field names.
    """
    lessonId: str | None = None
    sectionKey: str | None = None
    questionCount: int
    conceptRules: dict[str, Any] | None = None


class AssessmentBlueprintCreateRequest(BaseModel):
    title: str
    moduleId: str
    levelId: str
    totalQuestions: int
    durationSeconds: int
    instructions: str | None = None
    status: str = "DRAFT"
    lessonDistribution: list[AssessmentLessonDistributionRequest]


class AssessmentBlueprintUpdateRequest(BaseModel):
    title: str | None = None
    totalQuestions: int | None = None
    durationSeconds: int | None = None
    instructions: str | None = None
    lessonDistribution: list[AssessmentLessonDistributionRequest] | None = None


class StudentCreateRequest(BaseModel):
    customId: str
    teacher: str
    teacherCode: str | None = None
    teacherId: str | None = None
    admissionDate: str
    studentName: str
    dob: str
    gender: str
    bloodGroup: str
    schoolName: str
    className: str
    section: str
    fatherName: str
    fatherMobile: str
    fatherEmail: str
    fatherWhatsapp: str
    motherName: str
    motherMobile: str
    motherEmail: str
    motherWhatsapp: str
    studentCode: str
    password: str
    moduleCode: str | None = None
    levelCode: str | None = None
    currentModuleId: str | None = None
    currentLevelId: str | None = None
    status: str = "ACTIVE"
    interest: str | None = None
    presentAddress: str | None = None
    permanentAddress: str | None = None
    schoolArea: str | None = None
    fatherOccupation: str | None = None
    motherOccupation: str | None = None


class StudentUpdateRequest(BaseModel):
    customId: str | None = None
    teacher: str | None = None
    teacherCode: str | None = None
    teacherId: str | None = None
    admissionDate: str | None = None
    studentName: str | None = None
    dob: str | None = None
    gender: str | None = None
    bloodGroup: str | None = None
    schoolName: str | None = None
    className: str | None = None
    section: str | None = None
    fatherName: str | None = None
    fatherMobile: str | None = None
    fatherEmail: str | None = None
    fatherWhatsapp: str | None = None
    motherName: str | None = None
    motherMobile: str | None = None
    motherEmail: str | None = None
    motherWhatsapp: str | None = None
    studentCode: str | None = None
    moduleCode: str | None = None
    levelCode: str | None = None
    currentModuleId: str | None = None
    currentLevelId: str | None = None
    status: str | None = None
    interest: str | None = None
    presentAddress: str | None = None
    permanentAddress: str | None = None
    schoolArea: str | None = None
    fatherOccupation: str | None = None
    motherOccupation: str | None = None


class StudentStatusRequest(BaseModel):
    isActive: bool


class ResetPasswordRequest(BaseModel):
    # No hardcoded default on purpose -- a caller (frontend or a direct API
    # call) must always supply a real password. A hardcoded fallback here
    # previously meant any reset-password call with an omitted/blank
    # password silently landed on a fixed, publicly-documented string.
    password: str



class TeacherCreateRequest(BaseModel):
    teacherName: str
    teacherCode: str
    email: str | None = None
    phone: str | None = None
    # No hardcoded default on purpose -- see ResetPasswordRequest above.
    password: str
    designation: str | None = None
    subjectSpecialization: str | None = None
    qualification: str | None = None
    joiningDate: str | None = None
    address: str | None = None
    notes: str | None = None
    status: str = "ACTIVE"


class TeacherUpdateRequest(BaseModel):
    teacherName: str | None = None
    teacherCode: str | None = None
    email: str | None = None
    phone: str | None = None
    designation: str | None = None
    subjectSpecialization: str | None = None
    qualification: str | None = None
    joiningDate: str | None = None
    address: str | None = None
    notes: str | None = None
    status: str | None = None


class TeacherStatusRequest(BaseModel):
    isActive: bool


class ResetTeacherPasswordRequest(BaseModel):
    # No hardcoded default on purpose -- see ResetPasswordRequest above.
    password: str


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    cleaned = str(value).strip()
    return cleaned or None


def clean_password(value: Any) -> str | None:
    """Preserve Excel passwords as entered as much as possible.

    Excel sometimes returns whole-number passwords as floats. For example,
    123456 may arrive as 123456.0. This normalises that case without
    replacing the password with a default value.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def required_text(value: Any, label: str) -> str:
    cleaned = clean_text(value)
    if not cleaned:
        api_error(400, "VALIDATION_ERROR", f"{label} is required.")
    return cleaned


def status_to_active(value: Any) -> bool:
    status = (clean_text(value) or "ACTIVE").upper().replace(" ", "_")
    if status in {"ACTIVE", "TRUE", "YES", "1", "ENABLED"}:
        return True
    if status in {"INACTIVE", "FALSE", "NO", "0", "DISABLED"}:
        return False
    api_error(400, "VALIDATION_ERROR", "Status must be ACTIVE or INACTIVE.")


def get_module_level(
    db: Session,
    *,
    module_id: str | None = None,
    level_id: str | None = None,
    module_code: str | None = None,
    level_code: str | None = None,
) -> tuple[Module, Level]:
    module = db.get(Module, module_id) if module_id else None
    if not module and module_code:
        module = db.query(Module).filter(Module.module_code == module_code).first()
    if not module:
        api_error(404, "NOT_FOUND", "Selected module was not found.")

    level = db.get(Level, level_id) if level_id else None
    if not level and level_code:
        level = (
            db.query(Level)
            .filter(Level.module_id == module.id, Level.level_code == level_code)
            .first()
        )
    if not level:
        api_error(404, "NOT_FOUND", "Selected level was not found.")
    if level.module_id != module.id:
        api_error(400, "VALIDATION_ERROR", "Selected level does not belong to the selected module.")

    return module, level


def ensure_unique_student(
    db: Session,
    *,
    custom_id: str | None,
    student_code: str | None,
    exclude_student_id: str | None = None,
) -> None:
    if custom_id:
        query = db.query(Student).filter(Student.custom_id == custom_id)
        if exclude_student_id:
            query = query.filter(Student.id != exclude_student_id)
        if query.first():
            api_error(409, "DUPLICATE_CUSTOM_ID", "A student with this Custom ID already exists.")

    if student_code:
        query = db.query(Student).filter(Student.student_code == student_code)
        if exclude_student_id:
            query = query.filter(Student.id != exclude_student_id)
        if query.first():
            api_error(409, "DUPLICATE_STUDENT_CODE", "A student with this Student Code already exists.")



def make_teacher_code(name: str) -> str:
    compact = re.sub(r"[^A-Z0-9]", "", name.upper())[:10] or "TEACHER"
    return f"T-{compact}"


def ensure_unique_teacher(
    db: Session,
    *,
    teacher_code: str | None,
    email: str | None = None,
    phone: str | None = None,
    exclude_teacher_id: str | None = None,
    exclude_user_id: str | None = None,
) -> None:
    if teacher_code:
        query = db.query(Teacher).filter(Teacher.teacher_code == teacher_code)
        if exclude_teacher_id:
            query = query.filter(Teacher.id != exclude_teacher_id)
        if query.first():
            api_error(409, "DUPLICATE_TEACHER_CODE", "Teacher code already exists.")

    if email:
        query = db.query(User).filter(User.email == email)
        if exclude_user_id:
            query = query.filter(User.id != exclude_user_id)
        if query.first():
            api_error(409, "DUPLICATE_EMAIL", "Email already exists.")

    if phone:
        query = db.query(User).filter(User.phone == phone)
        if exclude_user_id:
            query = query.filter(User.id != exclude_user_id)
        if query.first():
            api_error(409, "DUPLICATE_PHONE", "Phone already exists.")


def teacher_payload(db: Session, teacher: Teacher) -> dict:
    teacher_user = db.get(User, teacher.user_id)
    teacher_name = teacher_user.full_name if teacher_user else ""

    if hasattr(Student, "teacher_id"):
        student_query = db.query(Student).filter(
            (Student.teacher_id == teacher.id) | (Student.teacher == teacher_name)
        )
    else:
        student_query = db.query(Student).filter(Student.teacher == teacher_name)

    total_students = student_query.count()
    active_students = student_query.filter(Student.is_active == True).count()

    return {
        "teacherId": teacher.id,
        "userId": teacher.user_id,
        "teacherName": teacher_user.full_name if teacher_user else "",
        "teacherCode": teacher.teacher_code,
        "email": teacher_user.email if teacher_user else None,
        "phone": teacher_user.phone if teacher_user else None,
        "designation": teacher.designation,
        "subjectSpecialization": teacher.subject_specialization,
        "qualification": teacher.qualification,
        "joiningDate": teacher.joining_date,
        "address": teacher.address,
        "notes": teacher.notes,
        "photoUrl": public_profile_photo_url(teacher_user, teacher.photo_url or (teacher_user.photo_url if teacher_user else None)) if teacher_user else teacher.photo_url,
        "profilePhotoUrl": public_profile_photo_url(teacher_user, teacher.photo_url or (teacher_user.photo_url if teacher_user else None)) if teacher_user else teacher.photo_url,
        "signatureUrl": teacher.signature_url,
        "status": "ACTIVE" if teacher.is_active and (teacher_user.is_active if teacher_user else True) else "INACTIVE",
        "isActive": bool(teacher.is_active and teacher_user.is_active) if teacher_user else bool(teacher.is_active),
        "studentCount": total_students,
        "activeStudentCount": active_students,
        "inactiveStudentCount": max(total_students - active_students, 0),
        "createdAt": teacher_user.created_at.isoformat() if teacher_user and teacher_user.created_at else None,
    }


def find_or_create_teacher_for_student(db: Session, teacher_name: str | None, teacher_code: str | None = None) -> Teacher | None:
    name = clean_text(teacher_name)
    code = clean_text(teacher_code)
    if not name:
        return None

    teacher = None
    if code:
        teacher = db.query(Teacher).filter(Teacher.teacher_code == code).first()
    if not teacher:
        teacher = (
            db.query(Teacher)
            .join(User, Teacher.user_id == User.id)
            .filter(User.full_name == name)
            .first()
        )
    if teacher:
        return teacher

    base_code = code or make_teacher_code(name)
    final_code = base_code
    counter = 1
    while db.query(Teacher).filter(Teacher.teacher_code == final_code).first():
        counter += 1
        final_code = f"{base_code}-{counter}"

    # This path auto-creates a Teacher account on the fly when a student is
    # created/edited with a teacher name that doesn't exist yet -- there's no
    # password field in that flow for the admin to set. It previously landed
    # on a fixed, publicly-documented default password; now it's a random,
    # unguessable one instead. The admin still needs to open
    # Admin -> Users -> Teachers -> this teacher -> reset password to hand
    # this new teacher real, known login credentials.
    teacher_user = User(
        full_name=name,
        email=None,
        phone=None,
        password_hash=hash_password(secrets.token_urlsafe(18)),
        role="TEACHER",
        is_active=True,
    )
    db.add(teacher_user)
    db.flush()
    teacher = Teacher(
        user_id=teacher_user.id,
        teacher_code=final_code,
        is_active=True,
    )
    db.add(teacher)
    db.flush()
    return teacher


def student_payload(db: Session, student: Student) -> dict:
    student_user = db.get(User, student.user_id)
    module = db.get(Module, student.current_module_id) if student.current_module_id else None
    level = db.get(Level, student.current_level_id) if student.current_level_id else None
    teacher = db.get(Teacher, student.teacher_id) if hasattr(student, "teacher_id") and getattr(student, "teacher_id", None) else None
    teacher_user = db.get(User, teacher.user_id) if teacher else None

    return {
        "studentId": student.id,
        "userId": student.user_id,
        "customId": student.custom_id,
        "teacher": student.teacher,
        "teacherId": getattr(student, "teacher_id", None),
        "teacherCode": teacher.teacher_code if teacher else None,
        "teacherName": teacher_user.full_name if teacher_user else student.teacher,
        "admissionDate": student.admission_date,
        "studentName": student_user.full_name if student_user else "",
        "fullName": student_user.full_name if student_user else "",
        "dob": student.dob,
        "gender": student.gender,
        "bloodGroup": student.blood_group,
        "interest": student.interest,
        "photoUrl": public_profile_photo_url(student_user, student.photo_url or (student_user.photo_url if student_user else None)) if student_user else student.photo_url,
        "profilePhotoUrl": public_profile_photo_url(student_user, student.photo_url or (student_user.photo_url if student_user else None)) if student_user else student.photo_url,
        "signatureUrl": student.signature_url,
        "presentAddress": student.present_address,
        "permanentAddress": student.permanent_address,
        "schoolName": student.school_name,
        "schoolArea": student.school_area,
        "className": student.class_name,
        "section": student.section,
        "fatherName": student.father_name,
        "fatherOccupation": student.father_occupation,
        "fatherMobile": student.father_mobile,
        "fatherEmail": student.father_email,
        "fatherWhatsapp": student.father_whatsapp,
        "motherName": student.mother_name,
        "motherOccupation": student.mother_occupation,
        "motherMobile": student.mother_mobile,
        "motherEmail": student.mother_email,
        "motherWhatsapp": student.mother_whatsapp,
        "studentCode": student.student_code,
        "currentModuleId": student.current_module_id,
        "currentModuleCode": module.module_code if module else None,
        "currentModuleName": module.module_name if module else None,
        "currentLevelId": student.current_level_id,
        "currentLevelCode": level.level_code if level else None,
        "currentLevelName": level.level_name if level else None,
        "status": "ACTIVE" if student.is_active and (student_user.is_active if student_user else True) else "INACTIVE",
        "isActive": bool(student.is_active and student_user.is_active) if student_user else bool(student.is_active),
        "createdAt": student_user.created_at.isoformat() if student_user and student_user.created_at else None,
    }


def apply_student_fields(student: Student, payload: StudentCreateRequest | StudentUpdateRequest, module: Module, level: Level, db: Session | None = None) -> None:
    mapping = {
        "custom_id": "customId",
        "teacher": "teacher",
        "admission_date": "admissionDate",
        "dob": "dob",
        "gender": "gender",
        "blood_group": "bloodGroup",
        "interest": "interest",
        "present_address": "presentAddress",
        "permanent_address": "permanentAddress",
        "school_name": "schoolName",
        "school_area": "schoolArea",
        "class_name": "className",
        "section": "section",
        "father_name": "fatherName",
        "father_occupation": "fatherOccupation",
        "father_mobile": "fatherMobile",
        "father_email": "fatherEmail",
        "father_whatsapp": "fatherWhatsapp",
        "mother_name": "motherName",
        "mother_occupation": "motherOccupation",
        "mother_mobile": "motherMobile",
        "mother_email": "motherEmail",
        "mother_whatsapp": "motherWhatsapp",
        "student_code": "studentCode",
    }

    for db_field, payload_field in mapping.items():
        if hasattr(payload, payload_field):
            value = getattr(payload, payload_field)
            if value is not None:
                setattr(student, db_field, clean_text(value))

    if db is not None:
        teacher = find_or_create_teacher_for_student(
            db,
            getattr(payload, "teacher", None),
            getattr(payload, "teacherCode", None),
        )
        if teacher:
            teacher_user = db.get(User, teacher.user_id)
            if hasattr(student, "teacher_id"):
                student.teacher_id = teacher.id
            student.teacher = teacher_user.full_name if teacher_user else student.teacher

    student.current_module_id = module.id
    student.current_level_id = level.id
    if payload.status is not None:
        student.is_active = status_to_active(payload.status)


def safe_filename(filename: str, prefix: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
        api_error(400, "INVALID_FILE", "Only JPG, PNG, and WEBP images are allowed.")
    safe_prefix = re.sub(r"[^a-zA-Z0-9_-]", "-", prefix)[:80]
    return f"{safe_prefix}-{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}{suffix}"


def save_image(upload: UploadFile, folder: Path, prefix: str) -> str:
    filename = safe_filename(upload.filename or "image.png", prefix)
    path = folder / filename
    with path.open("wb") as buffer:
        shutil.copyfileobj(upload.file, buffer)
    return f"/uploads/students/{folder.name}/{filename}"



@router.get("/teachers")
def list_teachers(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    teachers = (
        db.query(Teacher)
        .join(User, Teacher.user_id == User.id)
        .order_by(User.full_name.asc())
        .all()
    )
    return {"teachers": [teacher_payload(db, teacher) for teacher in teachers]}


@router.post("/teachers")
def create_teacher_route(payload: TeacherCreateRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    teacher_name = required_text(payload.teacherName, "Teacher name")
    teacher_code = required_text(payload.teacherCode, "Teacher code")
    password = clean_password(payload.password)
    if not password:
        api_error(400, "VALIDATION_ERROR", "Password is required.")
    if len(password) < 6:
        api_error(400, "VALIDATION_ERROR", "Password must be at least 6 characters.")

    email = clean_text(payload.email)
    phone = clean_text(payload.phone)
    ensure_unique_teacher(db, teacher_code=teacher_code, email=email, phone=phone)

    teacher_user = User(
        full_name=teacher_name,
        email=email,
        phone=phone,
        password_hash=hash_password(password),
        role="TEACHER",
        is_active=status_to_active(payload.status),
    )
    db.add(teacher_user)
    db.flush()

    teacher = Teacher(
        user_id=teacher_user.id,
        teacher_code=teacher_code,
        designation=clean_text(payload.designation),
        subject_specialization=clean_text(payload.subjectSpecialization),
        qualification=clean_text(payload.qualification),
        joining_date=clean_text(payload.joiningDate),
        address=clean_text(payload.address),
        notes=clean_text(payload.notes),
        is_active=teacher_user.is_active,
    )
    db.add(teacher)
    db.commit()
    db.refresh(teacher)

    return {
        "created": True,
        "message": "Teacher created successfully.",
        "teacher": teacher_payload(db, teacher),
        "login": {"identifier": teacher_user.email or teacher_user.phone or teacher.teacher_code, "password": password},
    }


@router.patch("/teachers/{teacher_id}")
def update_teacher_route(teacher_id: str, payload: TeacherUpdateRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        api_error(404, "NOT_FOUND", "Teacher not found.")
    teacher_user = db.get(User, teacher.user_id)
    if not teacher_user:
        api_error(404, "NOT_FOUND", "Teacher user profile not found.")

    teacher_code = clean_text(payload.teacherCode) if payload.teacherCode is not None else teacher.teacher_code
    email = clean_text(payload.email) if payload.email is not None else teacher_user.email
    phone = clean_text(payload.phone) if payload.phone is not None else teacher_user.phone

    ensure_unique_teacher(
        db,
        teacher_code=teacher_code,
        email=email,
        phone=phone,
        exclude_teacher_id=teacher.id,
        exclude_user_id=teacher_user.id,
    )

    old_name = teacher_user.full_name
    if payload.teacherName is not None:
        teacher_user.full_name = required_text(payload.teacherName, "Teacher name")
    teacher_user.email = email
    teacher_user.phone = phone
    teacher.teacher_code = required_text(teacher_code, "Teacher code")

    if payload.designation is not None:
        teacher.designation = clean_text(payload.designation)
    if payload.subjectSpecialization is not None:
        teacher.subject_specialization = clean_text(payload.subjectSpecialization)
    if payload.qualification is not None:
        teacher.qualification = clean_text(payload.qualification)
    if payload.joiningDate is not None:
        teacher.joining_date = clean_text(payload.joiningDate)
    if payload.address is not None:
        teacher.address = clean_text(payload.address)
    if payload.notes is not None:
        teacher.notes = clean_text(payload.notes)
    if payload.status is not None:
        teacher.is_active = status_to_active(payload.status)
        teacher_user.is_active = teacher.is_active

    if hasattr(Student, "teacher_id"):
        db.query(Student).filter(Student.teacher_id == teacher.id).update({"teacher": teacher_user.full_name})
    if old_name != teacher_user.full_name:
        db.query(Student).filter(Student.teacher == old_name, Student.teacher_id.is_(None)).update({"teacher": teacher_user.full_name})

    db.commit()
    db.refresh(teacher)
    return {"updated": True, "message": "Teacher updated successfully.", "teacher": teacher_payload(db, teacher)}


@router.patch("/teachers/{teacher_id}/status")
def update_teacher_status_route(teacher_id: str, payload: TeacherStatusRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        api_error(404, "NOT_FOUND", "Teacher not found.")
    teacher_user = db.get(User, teacher.user_id)
    if not teacher_user:
        api_error(404, "NOT_FOUND", "Teacher user profile not found.")
    teacher.is_active = payload.isActive
    teacher_user.is_active = payload.isActive
    db.commit()
    return {"updated": True, "message": "Teacher status updated.", "teacher": teacher_payload(db, teacher)}


@router.post("/teachers/{teacher_id}/reset-password")
@limiter.limit("10/minute")
def reset_teacher_password_route(request: Request, teacher_id: str, payload: ResetTeacherPasswordRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        api_error(404, "NOT_FOUND", "Teacher not found.")
    teacher_user = db.get(User, teacher.user_id)
    if not teacher_user:
        api_error(404, "NOT_FOUND", "Teacher user profile not found.")
    password = clean_password(payload.password)
    if not password:
        api_error(400, "VALIDATION_ERROR", "Password is required.")
    if len(password) < 6:
        api_error(400, "VALIDATION_ERROR", "Password must be at least 6 characters.")
    teacher_user.password_hash = hash_password(password)
    db.commit()
    return {"updated": True, "message": "Teacher password reset successfully.", "login": {"identifier": teacher_user.email or teacher_user.phone or teacher.teacher_code, "password": password}}


@router.post("/teachers/{teacher_id}/force-logout")
@limiter.limit("10/minute")
def force_logout_teacher_route(request: Request, teacher_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    """Immediately invalidate every active session for this teacher, no password reset required."""
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        api_error(404, "NOT_FOUND", "Teacher not found.")
    teacher_user = db.get(User, teacher.user_id)
    if not teacher_user:
        api_error(404, "NOT_FOUND", "Teacher user profile not found.")
    force_logout_user(db, teacher_user)
    return {"updated": True, "message": "Teacher has been signed out of all active sessions."}



@router.post("/teachers/{teacher_id}/photo")
def upload_teacher_photo_route(
    teacher_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        api_error(404, "NOT_FOUND", "Teacher not found.")

    extension = Path(file.filename or "").suffix.lower()
    if extension not in [".png", ".jpg", ".jpeg", ".webp"]:
        api_error(400, "VALIDATION_ERROR", "Only PNG, JPG, JPEG, and WEBP images are allowed.")

    target = TEACHER_PHOTO_DIR / f"{teacher.id}{extension}"
    with target.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    teacher.photo_url = f"/uploads/teachers/photos/{target.name}"
    db.commit()
    db.refresh(teacher)

    return {"uploaded": True, "teacher": teacher_payload(db, teacher)}


@router.post("/teachers/{teacher_id}/signature")
def upload_teacher_signature_route(
    teacher_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        api_error(404, "NOT_FOUND", "Teacher not found.")

    extension = Path(file.filename or "").suffix.lower()
    if extension not in [".png", ".jpg", ".jpeg", ".webp"]:
        api_error(400, "VALIDATION_ERROR", "Only PNG, JPG, JPEG, and WEBP images are allowed.")

    target = TEACHER_SIGNATURE_DIR / f"{teacher.id}{extension}"
    with target.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    teacher.signature_url = f"/uploads/teachers/signatures/{target.name}"
    db.commit()
    db.refresh(teacher)

    return {"uploaded": True, "teacher": teacher_payload(db, teacher)}


@router.delete("/teachers/{teacher_id}")
def delete_teacher_route(teacher_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        api_error(404, "NOT_FOUND", "Teacher not found.")
    teacher_user = db.get(User, teacher.user_id)
    teacher_name = teacher_user.full_name if teacher_user else None

    # Permanent delete means removing the teacher account and safely clearing all optional references.
    # Student records are preserved and unassigned instead of being deleted.
    if hasattr(Student, "teacher_id"):
        db.query(Student).filter(Student.teacher_id == teacher.id).update({"teacher_id": None}, synchronize_session=False)
    if teacher_name:
        db.query(Student).filter(Student.teacher == teacher_name).update({"teacher": None}, synchronize_session=False)
    if hasattr(Batch, "teacher_id"):
        db.query(Batch).filter(Batch.teacher_id == teacher.id).update({"teacher_id": None}, synchronize_session=False)
    if hasattr(Assignment, "teacher_id"):
        db.query(Assignment).filter(Assignment.teacher_id == teacher.id).update({"teacher_id": None}, synchronize_session=False)
    if hasattr(AssessmentAssignment, "teacher_id"):
        db.query(AssessmentAssignment).filter(AssessmentAssignment.teacher_id == teacher.id).update({"teacher_id": None}, synchronize_session=False)
    if hasattr(Notification, "teacher_id"):
        db.query(Notification).filter(Notification.teacher_id == teacher.id).update({"teacher_id": None}, synchronize_session=False)

    photo_url = teacher.photo_url
    signature_url = teacher.signature_url
    db.delete(teacher)
    if teacher_user:
        db.delete(teacher_user)
    db.commit()

    for stored_url in [photo_url, signature_url]:
        if stored_url and stored_url.startswith("/uploads/teachers/"):
            file_path = Path(stored_url.lstrip("/"))
            try:
                if file_path.exists() and file_path.is_file():
                    file_path.unlink()
            except OSError:
                pass

    return {"deleted": True, "message": "Teacher deleted permanently.", "teacherId": teacher_id}

@router.get("/live-students")
def get_live_students(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    from datetime import datetime, timezone, timedelta
    five_mins_ago = datetime.now(timezone.utc) - timedelta(minutes=5)
    
    live_rows = (
        db.query(Student, User)
        .join(User, Student.user_id == User.id)
        .filter(User.last_active_at >= five_mins_ago)
        .order_by(User.last_active_at.desc())
        .all()
    )
    
    students = []
    for student, u in live_rows:
        students.append({
            "id": u.id,
            "full_name": u.full_name,
            "student_code": student.student_code,
            "last_active_at": u.last_active_at.isoformat() if u.last_active_at else None
        })
        
    return {"live_students": students, "count": len(students)}


@router.get("/students")
def list_students(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    rows = (
        db.query(Student)
        .join(User, Student.user_id == User.id)
        .order_by(User.full_name.asc())
        .all()
    )
    return {"students": [student_payload(db, student) for student in rows]}


@router.post("/students")
def create_student_route(payload: StudentCreateRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    custom_id = required_text(payload.customId, "Custom ID")
    student_name = required_text(payload.studentName, "Student name")
    student_code = required_text(payload.studentCode, "Student code")
    password = clean_password(payload.password)
    if not password:
        api_error(400, "VALIDATION_ERROR", "Password is required.")
    if len(password) < 6:
        api_error(400, "VALIDATION_ERROR", "Password must be at least 6 characters.")

    for value, label in [
        (payload.teacher, "Teacher"),
        (payload.admissionDate, "Admission date"),
        (payload.dob, "DOB"),
        (payload.gender, "Gender"),
        (payload.bloodGroup, "Blood group"),
        (payload.schoolName, "School name"),
        (payload.className, "Class"),
        (payload.section, "Section"),
        (payload.fatherName, "Father name"),
        (payload.fatherMobile, "Father mobile"),
        (payload.fatherEmail, "Father email"),
        (payload.fatherWhatsapp, "Father WhatsApp"),
        (payload.motherName, "Mother name"),
        (payload.motherMobile, "Mother mobile"),
        (payload.motherEmail, "Mother email"),
        (payload.motherWhatsapp, "Mother WhatsApp"),
    ]:
        required_text(value, label)

    module, level = get_module_level(
        db,
        module_id=payload.currentModuleId,
        level_id=payload.currentLevelId,
        module_code=payload.moduleCode,
        level_code=payload.levelCode,
    )
    ensure_unique_student(db, custom_id=custom_id, student_code=student_code)

    student_user = User(
        full_name=student_name,
        email=None,
        phone=None,
        password_hash=hash_password(password),
        role="STUDENT",
        is_active=status_to_active(payload.status),
    )
    db.add(student_user)
    db.flush()

    student = Student(user_id=student_user.id, student_code=student_code, is_active=student_user.is_active)
    apply_student_fields(student, payload, module, level, db)
    db.add(student)
    db.commit()
    db.refresh(student)

    return {
        "created": True,
        "message": "Student created successfully.",
        "student": student_payload(db, student),
        "login": {"identifier": student.student_code, "password": password},
    }


@router.patch("/students/{student_id}")
def update_student_route(student_id: str, payload: StudentUpdateRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")
    student_user = db.get(User, student.user_id)
    if not student_user:
        api_error(404, "NOT_FOUND", "Student user profile not found.")

    module, level = get_module_level(
        db,
        module_id=payload.currentModuleId or student.current_module_id,
        level_id=payload.currentLevelId or student.current_level_id,
        module_code=payload.moduleCode,
        level_code=payload.levelCode,
    )
    ensure_unique_student(
        db,
        custom_id=clean_text(payload.customId) if payload.customId is not None else student.custom_id,
        student_code=clean_text(payload.studentCode) if payload.studentCode is not None else student.student_code,
        exclude_student_id=student.id,
    )

    if payload.studentName is not None:
        student_user.full_name = required_text(payload.studentName, "Student name")
    if payload.status is not None:
        student_user.is_active = status_to_active(payload.status)
    apply_student_fields(student, payload, module, level, db)
    student.is_active = student_user.is_active if payload.status is not None else student.is_active

    db.commit()
    db.refresh(student)
    return {"updated": True, "message": "Student updated successfully.", "student": student_payload(db, student)}


@router.patch("/students/{student_id}/status")
def update_student_status_route(student_id: str, payload: StudentStatusRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")
    student_user = db.get(User, student.user_id)
    if not student_user:
        api_error(404, "NOT_FOUND", "Student user profile not found.")

    student.is_active = payload.isActive
    student_user.is_active = payload.isActive
    db.commit()
    return {"updated": True, "message": "Student status updated.", "student": student_payload(db, student)}


@router.post("/students/{student_id}/reset-password")
@limiter.limit("10/minute")
def reset_student_password_route(request: Request, student_id: str, payload: ResetPasswordRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")
    student_user = db.get(User, student.user_id)
    if not student_user:
        api_error(404, "NOT_FOUND", "Student user profile not found.")
    password = clean_password(payload.password)
    if not password:
        api_error(400, "VALIDATION_ERROR", "Password is required.")
    if len(password) < 6:
        api_error(400, "VALIDATION_ERROR", "Password must be at least 6 characters.")
    student_user.password_hash = hash_password(password)
    db.commit()
    return {"updated": True, "message": "Student password reset successfully.", "login": {"identifier": student.student_code, "password": password}}


@router.post("/students/{student_id}/force-logout")
@limiter.limit("10/minute")
def force_logout_student_route(request: Request, student_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    """Immediately invalidate every active session for this student, no password reset required."""
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")
    student_user = db.get(User, student.user_id)
    if not student_user:
        api_error(404, "NOT_FOUND", "Student user profile not found.")
    force_logout_user(db, student_user)
    return {"updated": True, "message": "Student has been signed out of all active sessions."}


@router.post("/students/{student_id}/photo")
def upload_student_photo(student_id: str, file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")
    student.photo_url = save_image(file, PHOTO_DIR, student.student_code)
    db.commit()
    return {"updated": True, "photoUrl": student.photo_url, "student": student_payload(db, student)}


@router.post("/students/{student_id}/signature")
def upload_student_signature(student_id: str, file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")
    student.signature_url = save_image(file, SIGNATURE_DIR, student.student_code)
    db.commit()
    return {"updated": True, "signatureUrl": student.signature_url, "student": student_payload(db, student)}



@router.delete("/students/{student_id}")
def delete_student_route(
    student_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")

    student_user = db.get(User, student.user_id)
    student_name = student_user.full_name if student_user else student.student_code
    student_code = student.student_code
    photo_url = student.photo_url
    signature_url = student.signature_url

    try:
        # Assessment-side cleanup. These records are shown in tracker/report views and
        # can block permanent student deletion on PostgreSQL when FK constraints are strict.
        assessment_assignment_ids = [
            row[0]
            for row in db.query(AssessmentAssignment.id)
            .filter(AssessmentAssignment.student_id == student.id)
            .all()
        ]
        assessment_attempt_ids = [
            row[0]
            for row in db.query(AssessmentAttempt.id)
            .filter(AssessmentAttempt.student_id == student.id)
            .all()
        ]
        assessment_result_ids = []
        if assessment_attempt_ids:
            assessment_result_ids = [
                row[0]
                for row in db.query(AssessmentResult.id)
                .filter(AssessmentResult.assessment_attempt_id.in_(assessment_attempt_ids))
                .all()
            ]
            db.query(AssessmentAttemptAnswer).filter(
                AssessmentAttemptAnswer.assessment_attempt_id.in_(assessment_attempt_ids)
            ).delete(synchronize_session=False)
            db.query(AssessmentReattemptApproval).filter(
                AssessmentReattemptApproval.assessment_attempt_id.in_(assessment_attempt_ids)
            ).delete(synchronize_session=False)
            db.query(StudentLevelPromotion).filter(
                StudentLevelPromotion.assessment_attempt_id.in_(assessment_attempt_ids)
            ).delete(synchronize_session=False)
            db.query(AssessmentResult).filter(
                AssessmentResult.assessment_attempt_id.in_(assessment_attempt_ids)
            ).delete(synchronize_session=False)
            db.query(AssessmentAttempt).filter(
                AssessmentAttempt.id.in_(assessment_attempt_ids)
            ).delete(synchronize_session=False)

        if assessment_result_ids:
            db.query(StudentLevelPromotion).filter(
                StudentLevelPromotion.assessment_result_id.in_(assessment_result_ids)
            ).delete(synchronize_session=False)

        if assessment_assignment_ids:
            db.query(StudentLevelPromotion).filter(
                StudentLevelPromotion.assessment_assignment_id.in_(assessment_assignment_ids)
            ).delete(synchronize_session=False)
            db.query(AssessmentReattemptApproval).filter(
                AssessmentReattemptApproval.assessment_assignment_id.in_(assessment_assignment_ids)
            ).delete(synchronize_session=False)
            db.query(AssessmentAssignment).filter(
                AssessmentAssignment.id.in_(assessment_assignment_ids)
            ).delete(synchronize_session=False)

        # Practice-side cleanup. Remove answers first, break generated-set references,
        # then remove generated questions/options/sets and assignment records.
        practice_attempts = db.query(Attempt).filter(Attempt.student_id == student.id).all()
        practice_attempt_ids = [attempt.id for attempt in practice_attempts]
        question_set_ids = [attempt.question_set_id for attempt in practice_attempts if attempt.question_set_id]

        if practice_attempt_ids:
            db.query(AuditLog).filter(AuditLog.attempt_id.in_(practice_attempt_ids)).update(
                {AuditLog.attempt_id: None},
                synchronize_session=False,
            )
            db.query(AttemptAnswer).filter(
                AttemptAnswer.attempt_id.in_(practice_attempt_ids)
            ).delete(synchronize_session=False)
            db.query(Attempt).filter(Attempt.id.in_(practice_attempt_ids)).update(
                {Attempt.question_set_id: None},
                synchronize_session=False,
            )
            db.query(Attempt).filter(
                Attempt.id.in_(practice_attempt_ids)
            ).delete(synchronize_session=False)

        if question_set_ids:
            question_ids = [
                row[0]
                for row in db.query(GeneratedQuestion.id)
                .filter(GeneratedQuestion.question_set_id.in_(question_set_ids))
                .all()
            ]
            if question_ids:
                db.query(QuestionOption).filter(
                    QuestionOption.question_id.in_(question_ids)
                ).delete(synchronize_session=False)
                db.query(GeneratedQuestion).filter(
                    GeneratedQuestion.id.in_(question_ids)
                ).delete(synchronize_session=False)
            db.query(GeneratedQuestionSet).filter(
                GeneratedQuestionSet.id.in_(question_set_ids)
            ).delete(synchronize_session=False)

        # Assignment and readiness governance records linked to the student.
        db.query(AssignmentReattemptPermission).filter(
            AssignmentReattemptPermission.student_id == student.id
        ).delete(synchronize_session=False)
        db.query(Assignment).filter(
            Assignment.assigned_to_type == "STUDENT",
            Assignment.assigned_to_id == student.id,
        ).delete(synchronize_session=False)
        db.query(Assignment).filter(
            Assignment.assigned_to_type == "STUDENT",
            Assignment.assigned_to_id == student.student_code,
        ).delete(synchronize_session=False)
        db.query(GeneratedQuestionSet).filter(
            GeneratedQuestionSet.student_id == student.id
        ).delete(synchronize_session=False)
        db.query(AssessmentReadinessTestingOverride).filter(
            AssessmentReadinessTestingOverride.student_id == student.id
        ).delete(synchronize_session=False)
        db.query(StudentBatch).filter(StudentBatch.student_id == student.id).delete(synchronize_session=False)

        # Preserve operational audit/delivery rows without letting nullable references
        # block permanent deletion.
        db.query(AuditLog).filter(AuditLog.student_id == student.id).update(
            {AuditLog.student_id: None},
            synchronize_session=False,
        )
        if student_user:
            db.query(AuditLog).filter(AuditLog.user_id == student_user.id).update(
                {AuditLog.user_id: None},
                synchronize_session=False,
            )
        db.query(Notification).filter(Notification.student_id == student.id).update(
            {Notification.student_id: None},
            synchronize_session=False,
        )
        db.query(ParentReportEmailLog).filter(ParentReportEmailLog.student_id == student.id).update(
            {ParentReportEmailLog.student_id: None},
            synchronize_session=False,
        )

        db.delete(student)
        if student_user:
            db.delete(student_user)

        db.commit()
    except Exception as exc:
        db.rollback()
        api_error(
            500,
            "STUDENT_DELETE_FAILED",
            "Student could not be deleted because linked production records could not be cleaned safely.",
            {"reason": str(exc)},
        )

    for stored_url in [photo_url, signature_url]:
        if stored_url and stored_url.startswith("/uploads/students/"):
            file_path = Path(stored_url.lstrip("/"))
            try:
                if file_path.exists() and file_path.is_file():
                    file_path.unlink()
            except OSError:
                pass

    return {
        "deleted": True,
        "message": "Student and linked operational records deleted successfully.",
        "studentId": student_id,
        "studentName": student_name,
        "studentCode": student_code,
    }


@router.get("/students/template")
def download_students_template(user: User = Depends(admin_dep)):
    wb = Workbook()
    ws = wb.active
    ws.title = "students"
    ws.append(TEMPLATE_FIELDS)
    sample = {
        "custom_id": "11030100225187",
        "teacher": "Teacher Name",
        "teacher_code": "T-TEACHER",
        "admission_date": "2026-05-01",
        "student_name": "Demo Student",
        "dob": "2017-01-15",
        "gender": "Female",
        "blood_group": "O+",
        "school_name": "Demo School",
        "class": "Class 3",
        "section": "A",
        "father_name": "Father Name",
        "father_mobile": "9999999999",
        "father_email": "father@example.com",
        "father_whatsapp": "9999999999",
        "mother_name": "Mother Name",
        "mother_mobile": "8888888888",
        "mother_email": "mother@example.com",
        "mother_whatsapp": "8888888888",
        "student_code": "MP-DEMO-1001",
        "password": "Choose-A-Strong-Password-1",
        "module_code": "YLM",
        "level_code": "YLM-L1",
        "status": "ACTIVE",
        "interest": "Maths",
        "present_address": "Present address",
        "permanent_address": "Permanent address",
        "school_area": "Kolkata",
        "father_occupation": "Business",
        "mother_occupation": "Teacher",
    }
    ws.append([sample.get(field, "") for field in TEMPLATE_FIELDS])
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 3, 14), 28)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    headers = {"Content-Disposition": "attachment; filename=mathpath_students_bulk_template.xlsx"}
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


MAX_BULK_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB


@router.post("/students/bulk-upload")
def bulk_upload_students(file: UploadFile = File(...), db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    FileName = file.filename or ""
    if not FileName.lower().endswith((".xlsx", ".xlsm")):
        api_error(400, "INVALID_FILE", "Please upload the official MathPath Excel template in .xlsx or .xlsm format.")

    # Guard against memory exhaustion before handing the raw stream to
    # openpyxl, which loads the entire workbook into memory to parse it. A
    # legitimate bulk-upload template is a few dozen KB even for hundreds of
    # students, so this ceiling is generous headroom, not a tight limit.
    file.file.seek(0, os.SEEK_END)
    FileSizeBytes = file.file.tell()
    file.file.seek(0)
    if FileSizeBytes > MAX_BULK_UPLOAD_BYTES:
        api_error(400, "FILE_TOO_LARGE", "Upload file must be under 5 MB.")

    try:
        WorkbookObject = load_workbook(file.file, data_only=True)
    except Exception:
        api_error(400, "INVALID_EXCEL", "The uploaded file could not be opened. Please download a fresh MathPath template and try again.")

    Sheet = WorkbookObject.active
    Rows = list(Sheet.iter_rows(values_only=True))
    if not Rows:
        api_error(400, "EMPTY_FILE", "Uploaded Excel file is empty.")

    def NormalizeHeader(Value) -> str:
        return re.sub(r"\s+", "_", str(Value or "").strip().lower())

    Headers = [NormalizeHeader(Value) for Value in Rows[0]]
    HeaderSet = {Header for Header in Headers if Header}
    MissingColumns = [Field for Field in MANDATORY_BULK_FIELDS if Field not in HeaderSet]
    UnknownColumns = [Header for Header in Headers if Header and Header not in set(TEMPLATE_FIELDS)]
    if MissingColumns:
        api_error(400, "MISSING_COLUMNS", f"Missing mandatory columns: {', '.join(MissingColumns)}")

    # First Pass: Collect all identifiers to limit DB memory usage
    RequestedCustomIds = set()
    RequestedStudentCodes = set()
    
    ParsedRows = []
    for RowNumber, RowValues in enumerate(Rows[1:], start=2):
        Row = {Headers[Index]: RowValues[Index] if Index < len(RowValues) else None for Index in range(len(Headers)) if Headers[Index]}
        if all(clean_text(Value) is None for Value in Row.values()):
            ParsedRows.append({"is_empty": True, "row_number": RowNumber})
            continue
            
        CId = clean_text(Row.get("custom_id"))
        if CId: RequestedCustomIds.add(CId)
            
        SCode = clean_text(Row.get("student_code"))
        if SCode: RequestedStudentCodes.add(SCode)
            
        ParsedRows.append({"is_empty": False, "row_number": RowNumber, "data": Row})

    Created = 0
    Failed = 0
    Skipped = 0
    Results = []
    SeenCustomIds = set()
    SeenStudentCodes = set()
    SeenIdentifiers = set()

    ModuleByCode = {ModuleValue.module_code: ModuleValue for ModuleValue in db.query(Module).all()}
    LevelsByModuleAndCode = {
        (LevelValue.module_id, LevelValue.level_code): LevelValue
        for LevelValue in db.query(Level).all()
    }
    TeachersByCode = {
        TeacherValue.teacher_code: TeacherValue
        for TeacherValue in db.query(Teacher).filter(Teacher.teacher_code.isnot(None)).all()
    }
    TeacherUsers = (
        db.query(Teacher, User)
        .join(User, Teacher.user_id == User.id)
        .filter(User.role == "TEACHER")
        .all()
    )
    TeachersByName: dict[str, list[Teacher]] = {}
    TeacherNamesById: dict[str, str] = {}
    for TeacherValue, TeacherUser in TeacherUsers:
        CleanName = clean_text(TeacherUser.full_name)
        if not CleanName:
            continue
        TeacherNamesById[TeacherValue.id] = CleanName
        TeachersByName.setdefault(CleanName.lower(), []).append(TeacherValue)

    # Targeted Queries for large tables (only fetch what is uploaded)
    ExistingCustomIds = {
        Value[0]
        for Value in db.query(Student.custom_id).filter(Student.custom_id.in_(RequestedCustomIds)).all()
        if Value[0]
    } if RequestedCustomIds else set()
    
    ExistingStudentCodes = {
        Value[0]
        for Value in db.query(Student.student_code).filter(Student.student_code.in_(RequestedStudentCodes)).all()
        if Value[0]
    } if RequestedStudentCodes else set()
    
    ExistingLoginIdentifiers = {
        Value[0]
        for Value in db.query(User.phone).filter(User.phone.in_(RequestedStudentCodes)).all()
        if Value[0]
    } if RequestedStudentCodes else set()

    def ResolveTeacher(Row: dict) -> tuple[Teacher | None, str | None]:
        TeacherName = clean_text(Row.get("teacher"))
        TeacherCode = clean_text(Row.get("teacher_code"))
        
        if not TeacherName and not TeacherCode:
            raise ValueError("A teacher or teacher_code must be provided.")
            
        if TeacherCode:
            TeacherRecord = TeachersByCode.get(TeacherCode)
            if TeacherRecord:
                if TeacherName:
                    ExistingName = TeacherNamesById.get(TeacherRecord.id)
                    if ExistingName and ExistingName.lower() != TeacherName.lower():
                        raise ValueError(f"Teacher code {TeacherCode} belongs to {ExistingName}, not {TeacherName}.")
                return TeacherRecord, TeacherNamesById.get(TeacherRecord.id) or TeacherName
            if TeacherName:
                NameMatches = TeachersByName.get(TeacherName.lower(), [])
                if len(NameMatches) == 1:
                    return NameMatches[0], TeacherName
                if len(NameMatches) > 1:
                    raise ValueError(f"Teacher code {TeacherCode} was not found and multiple teachers matched {TeacherName}.")
            raise ValueError(f"Teacher code not found: {TeacherCode}.")
            
        if TeacherName:
            NameMatches = TeachersByName.get(TeacherName.lower(), [])
            if len(NameMatches) == 1:
                return NameMatches[0], TeacherName
            if len(NameMatches) > 1:
                raise ValueError("Multiple teachers matched this teacher name. Use teacher_code in the template.")
            raise ValueError(f"Teacher not found: {TeacherName}.")
            
        return None, None

    ValidRowsToInsert = []

    for ParsedRow in ParsedRows:
        if ParsedRow["is_empty"]:
            Skipped += 1
            continue
            
        RowNumber = ParsedRow["row_number"]
        Row = ParsedRow["data"]

        try:
            MissingValues = [Field for Field in MANDATORY_BULK_FIELDS if not clean_text(Row.get(Field))]
            if MissingValues:
                raise ValueError(f"Missing mandatory values: {', '.join(MissingValues)}")

            CustomId = clean_text(Row.get("custom_id"))
            StudentCode = clean_text(Row.get("student_code"))
            ModuleCode = clean_text(Row.get("module_code"))
            LevelCode = clean_text(Row.get("level_code"))

            if CustomId in SeenCustomIds:
                raise ValueError("Duplicate custom_id inside uploaded file.")
            if StudentCode in SeenStudentCodes:
                raise ValueError("Duplicate student_code inside uploaded file.")
            if StudentCode in SeenIdentifiers:
                raise ValueError("Duplicate login identifier inside uploaded file.")
            if CustomId in ExistingCustomIds:
                raise ValueError("Custom ID already exists in system.")
            if StudentCode in ExistingStudentCodes:
                raise ValueError("Student code already exists in system.")
            if StudentCode in ExistingLoginIdentifiers:
                raise ValueError("Student login identifier already exists in system.")

            ModuleRecord = ModuleByCode.get(ModuleCode)
            if not ModuleRecord:
                raise ValueError(f"Module code not found: {ModuleCode}.")
            LevelRecord = LevelsByModuleAndCode.get((ModuleRecord.id, LevelCode))
            if not LevelRecord:
                raise ValueError(f"Level code {LevelCode} not found for module {ModuleCode}.")

            TeacherRecord, TeacherDisplayName = ResolveTeacher(Row)
            Active = status_to_active(Row.get("status"))
            Password = clean_password(Row.get("password"))
            if not Password or len(Password) < 6:
                raise ValueError("Password must be at least 6 characters.")

            SeenCustomIds.add(CustomId)
            SeenStudentCodes.add(StudentCode)
            SeenIdentifiers.add(StudentCode)
            
            ValidRowsToInsert.append({
                "row_number": RowNumber,
                "row_data": Row,
                "custom_id": CustomId,
                "student_code": StudentCode,
                "module_id": ModuleRecord.id,
                "level_id": LevelRecord.id,
                "teacher_record": TeacherRecord,
                "teacher_display_name": TeacherDisplayName,
                "active": Active,
                "password": Password,
                "level_code": LevelCode
            })
            
        except Exception as Exc:
            Failed += 1
            Results.append({"row": RowNumber, "status": "FAILED", "error": str(Exc)})

    # Batch Insert the valid rows
    if ValidRowsToInsert:
        try:
            UsersToInsert = []
            for ValidRow in ValidRowsToInsert:
                Row = ValidRow["row_data"]
                UserObj = User(
                    full_name=clean_text(Row.get("student_name")),
                    email=None,
                    phone=ValidRow["student_code"],
                    password_hash=hash_password(ValidRow["password"]),
                    role="STUDENT",
                    is_active=ValidRow["active"],
                )
                UsersToInsert.append(UserObj)
                
            db.add_all(UsersToInsert)
            db.flush()
            
            StudentsToInsert = []
            for Index, ValidRow in enumerate(ValidRowsToInsert):
                Row = ValidRow["row_data"]
                StudentUser = UsersToInsert[Index]
                TeacherRecord = ValidRow["teacher_record"]
                
                StudentObj = Student(
                    user_id=StudentUser.id,
                    custom_id=ValidRow["custom_id"],
                    teacher=ValidRow["teacher_display_name"],
                    teacher_id=TeacherRecord.id if TeacherRecord else None,
                    admission_date=clean_text(Row.get("admission_date")),
                    dob=clean_text(Row.get("dob")),
                    gender=clean_text(Row.get("gender")),
                    blood_group=clean_text(Row.get("blood_group")),
                    interest=clean_text(Row.get("interest")),
                    present_address=clean_text(Row.get("present_address")),
                    permanent_address=clean_text(Row.get("permanent_address")),
                    school_name=clean_text(Row.get("school_name")),
                    school_area=clean_text(Row.get("school_area")),
                    class_name=clean_text(Row.get("class")),
                    section=clean_text(Row.get("section")),
                    father_name=clean_text(Row.get("father_name")),
                    father_occupation=clean_text(Row.get("father_occupation")),
                    father_mobile=clean_text(Row.get("father_mobile")),
                    father_email=clean_text(Row.get("father_email")),
                    father_whatsapp=clean_text(Row.get("father_whatsapp")),
                    mother_name=clean_text(Row.get("mother_name")),
                    mother_occupation=clean_text(Row.get("mother_occupation")),
                    mother_mobile=clean_text(Row.get("mother_mobile")),
                    mother_email=clean_text(Row.get("mother_email")),
                    mother_whatsapp=clean_text(Row.get("mother_whatsapp")),
                    student_code=ValidRow["student_code"],
                    current_module_id=ValidRow["module_id"],
                    current_level_id=ValidRow["level_id"],
                    is_active=ValidRow["active"],
                )
                StudentsToInsert.append(StudentObj)
                
            db.add_all(StudentsToInsert)
            db.commit()
            
            for Index, ValidRow in enumerate(ValidRowsToInsert):
                Row = ValidRow["row_data"]
                StudentUser = UsersToInsert[Index]
                TeacherRecord = ValidRow["teacher_record"]
                Created += 1
                Results.append({
                    "row": ValidRow["row_number"],
                    "status": "CREATED",
                    "studentCode": ValidRow["student_code"],
                    "studentName": StudentUser.full_name,
                    "teacherCode": TeacherRecord.teacher_code if TeacherRecord else clean_text(Row.get("teacher_code")),
                    "teacherName": ValidRow["teacher_display_name"],
                    "levelCode": ValidRow["level_code"],
                })
        except Exception as Exc:
            db.rollback()
            for ValidRow in ValidRowsToInsert:
                Failed += 1
                Results.append({
                    "row": ValidRow["row_number"],
                    "status": "FAILED",
                    "error": f"Database batch insert error: {str(Exc)}"
                })

    return {
        "created": Created,
        "failed": Failed,
        "skipped": Skipped,
        "totalRows": Created + Failed,
        "unknownColumns": UnknownColumns,
        "results": Results,
    }



@router.get("/notifications")
def admin_notifications(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    notifications = []

    completed = (
        db.query(Attempt)
        .filter(Attempt.status.in_(["SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"]))
        .order_by(Attempt.submitted_at.desc().nullslast())
        .limit(30)
        .all()
    )

    for attempt in completed:
        student = db.get(Student, attempt.student_id)
        student_user = db.get(User, student.user_id) if student else None
        assignment = db.get(Assignment, attempt.assignment_id) if attempt.assignment_id else None
        accuracy = float(attempt.accuracy_percentage or 0)

        if accuracy < 70:
            tone = "danger"
            title = "Below benchmark attempt"
            message = f"{student_user.full_name if student_user else 'A student'} scored {accuracy:g}% and needs teacher follow-up."
        elif accuracy >= 90:
            tone = "success"
            title = "Excellent performance"
            message = f"{student_user.full_name if student_user else 'A student'} scored {accuracy:g}%."
        else:
            tone = "warning"
            title = "Good progress"
            message = f"{student_user.full_name if student_user else 'A student'} scored {accuracy:g}% and can improve further."

        notifications.append({
            "id": f"admin-attempt-{attempt.id}",
            "title": title,
            "message": message,
            "tone": tone,
            "targetUrl": f"/admin/results/{attempt.id}",
            "createdAt": attempt.submitted_at.isoformat() if attempt.submitted_at else None,
            "assignmentType": assignment.assignment_type if assignment else None,
        })

    assignments = db.query(Assignment).filter(Assignment.is_active == True).order_by(Assignment.created_at.desc()).limit(20).all()

    for assignment in assignments:
        creator = db.get(User, assignment.assigned_by_user_id) if assignment.assigned_by_user_id else None
        is_assessment = assignment.assignment_type == "ASSESSMENT"
        notifications.append({
            "id": f"admin-assignment-{assignment.id}",
            "title": "Assessment created" if is_assessment else "Practice assigned",
            "message": f"{assignment.title} was created by {creator.full_name if creator else 'Admin/Teacher'}.",
            "tone": "purple" if is_assessment else "info",
            "targetUrl": f"/admin/assessments/{assignment.id}" if is_assessment else f"/admin/assignments/{assignment.id}",
            "createdAt": assignment.created_at.isoformat() if assignment.created_at else None,
            "assignmentType": assignment.assignment_type,
        })

    permissions = db.query(AssignmentReattemptPermission).order_by(AssignmentReattemptPermission.allowed_at.desc()).limit(10).all()
    for permission in permissions:
        assignment = db.get(Assignment, permission.assignment_id)
        student = db.get(Student, permission.student_id)
        student_user = db.get(User, student.user_id) if student else None
        notifications.append({
            "id": f"admin-reattempt-{permission.id}",
            "title": "Reattempt unlocked",
            "message": f"{student_user.full_name if student_user else 'A student'} was allowed to reattempt {assignment.title if assignment else 'an assigned sheet'}.",
            "tone": "purple",
            "targetUrl": f"/admin/assessments/{assignment.id}" if assignment and assignment.assignment_type == "ASSESSMENT" else f"/admin/assignments/{assignment.id}" if assignment else "/admin/results",
            "createdAt": permission.allowed_at.isoformat() if permission.allowed_at else None,
        })

    notifications = sorted(notifications, key=lambda item: item.get("createdAt") or "", reverse=True)
    return {"notifications": notifications[:35]}


@router.get("/modules")
def modules(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    rows = db.query(Module).filter(Module.is_active == True).order_by(Module.display_order).all()
    return {"modules": [{"moduleId": m.id, "moduleCode": m.module_code, "moduleName": m.module_name, "displayOrder": m.display_order, "isActive": m.is_active} for m in rows]}


@router.get("/modules/{module_id}/levels")
def levels(module_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    module = db.get(Module, module_id)
    rows = db.query(Level).filter(Level.module_id == module_id, Level.is_active == True).order_by(Level.display_order).all()
    return {"moduleId": module_id, "moduleCode": module.module_code if module else None, "levels": [{"levelId": l.id, "levelCode": l.level_code, "levelName": l.level_name, "internalLevelNumber": l.internal_level_number, "displayOrder": l.display_order} for l in rows]}


@router.get("/levels/{level_id}/lessons")
def lessons(level_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    level = db.get(Level, level_id)
    rows = db.query(Lesson).filter(Lesson.level_id == level_id, Lesson.is_active == True).order_by(Lesson.lesson_number).all()
    return {"levelId": level_id, "levelCode": level.level_code if level else None, "lessons": [{"lessonId": l.id, "lessonNumber": l.lesson_number, "lessonTitle": l.lesson_title, "dpsCount": db.query(DPS).filter(DPS.lesson_id == l.id, DPS.is_active == True).count(), "isActive": l.is_active} for l in rows]}


@router.get("/levels/{level_id}/assessment-sections")
def assessment_sections_for_level(level_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    """Section-wise counterpart of /levels/{level_id}/lessons above -- lets
    the Assessment Blueprint Studio build a section-by-section distribution
    form for IM/MM levels. Returns the SAME section list/order that level's
    competition mock exam uses (read live from the registry, not a copy),
    so the studio can never silently drift out of sync with mocks. For YLM
    (or any module without section-wise assessments), isSectionWise is False
    and the frontend should fall back to the lesson-wise /lessons endpoint.
    """
    level = db.get(Level, level_id)
    if not level:
        api_error(404, "LEVEL_NOT_FOUND", "Level not found.")
    module = db.get(Module, level.module_id)
    if not module:
        api_error(404, "MODULE_NOT_FOUND", "Module not found.")

    section_wise = is_section_wise_module(module.module_code)
    if not section_wise:
        return {"levelId": level_id, "levelCode": level.level_code, "moduleCode": module.module_code, "isSectionWise": False, "sections": []}

    registry_config = level_section_registry_config(module.module_code, level)
    marks_meta = section_marks_metadata(module.module_code, registry_config)
    sections = [
        {
            "sectionKey": row["key"],
            "sectionNumber": row["number"],
            "sectionTitle": row["title"],
            "conceptCount": len(registry_config.get("sectionConceptPools", {}).get(row["key"], [])),
            # 2026-07-23: lets the Blueprint Studio auto-balance the paper to
            # always total 100 marks -- isWeighted sections are worth
            # marksPerQuestion (5) each, everything else is worth 1.
            "isWeighted": marks_meta.get(row["key"], {}).get("isWeighted", False),
            "marksPerQuestion": marks_meta.get(row["key"], {}).get("marksPerQuestion", 1.0),
        }
        for row in registry_config["sectionDefinitions"]
    ]
    return {"levelId": level_id, "levelCode": level.level_code, "moduleCode": module.module_code, "isSectionWise": True, "sections": sections}


@router.get("/lessons/{lesson_id}/dps")
def dps_list(lesson_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    lesson = db.get(Lesson, lesson_id)
    rows = db.query(DPS).filter(DPS.lesson_id == lesson_id, DPS.is_active == True).order_by(DPS.dps_number).all()
    return {
        "lessonId": lesson_id,
        "lessonNumber": lesson.lesson_number if lesson else None,
        "lessonTitle": lesson.lesson_title if lesson else None,
        "dps": [
            {
                "dpsId": d.id,
                "dpsNumber": d.dps_number,
                "dpsTitle": d.dps_title,
                "questionCount": d.default_question_count,
                "durationSeconds": d.default_duration_seconds,
                "layoutTemplate": d.layout_template,
                "answerType": d.answer_type,
                "optionsPerQuestion": d.options_per_question,
                "publicationStatus": getattr(d, "publication_status", "DRAFT") or "DRAFT",
                "publishedAt": d.published_at.isoformat() if getattr(d, "published_at", None) else None,
                "isActive": d.is_active,
            }
            for d in rows
        ],
    }


@router.get("/dps/{dps_id}")
def dps_config(dps_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return dps_config_payload(db, dps_id)


@router.post("/dps/{dps_id}/generate-preview")
def preview(dps_id: str, payload: PreviewRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    dps = get_dps_or_404(db, dps_id)
    preview_seed = payload.seed or build_preview_seed(dps)
    questions = generate_preview(db, dps, preview_seed)
    dps.last_preview_seed = preview_seed
    db.commit()
    return {"dpsId": dps.id, "title": dps.dps_title, "previewSeed": preview_seed, "questions": questions}

@router.post("/dps/{dps_id}/publish")
def publish_dps_route(dps_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    dps = get_dps_or_404(db, dps_id)
    if not getattr(dps, "last_preview_seed", None):
        api_error(400, "PREVIEW_REQUIRED", "Generate and review a fresh DPS preview before publishing.")

    dps.publication_status = "PUBLISHED"
    dps.published_seed = dps.last_preview_seed
    dps.published_at = datetime.now(datetime_timezone.utc)
    dps.published_by_user_id = user.id
    db.commit()
    db.refresh(dps)
    return {
        "published": True,
        "message": "DPS published successfully. Teachers can now assign this sheet.",
        "dps": {
            "dpsId": dps.id,
            "dpsNumber": dps.dps_number,
            "dpsTitle": dps.dps_title,
            "publicationStatus": dps.publication_status,
            "publishedAt": dps.published_at.isoformat() if dps.published_at else None,
        },
    }



def display_role(role: str | None) -> str:
    if role in ["SUPER_ADMIN", "ADMIN"]:
        return "ADMIN"
    if role == "TEACHER":
        return "TEACHER"
    if role == "STUDENT":
        return "STUDENT"
    return role or "SYSTEM"


def resolve_assignment_target_id(db: Session, assigned_to_type: str, assigned_to_id: str) -> str:
    target_type = clean_text(assigned_to_type)
    target_id = clean_text(assigned_to_id)

    if not target_type or not target_id:
        api_error(400, "VALIDATION_ERROR", "Assignment target type and target ID are required.")

    if target_type == "LEVEL":
        if not db.get(Level, target_id):
            api_error(404, "NOT_FOUND", "Selected level was not found.")
        return target_id

    if target_type == "STUDENT":
        student = db.get(Student, target_id)
        if not student:
            student = db.query(Student).filter(Student.student_code == target_id).first()
        if not student:
            student = db.query(Student).filter(Student.custom_id == target_id).first()
        if not student:
            api_error(
                404,
                "NOT_FOUND",
                "Selected student was not found. Please choose a valid student from the dropdown.",
            )
        return student.id

    # Batch assignment is retained for future expansion.
    return target_id


def assignment_target_payload(db: Session, assignment: Assignment) -> dict:
    target_type = assignment.assigned_to_type
    target_id = assignment.assigned_to_id
    label = target_id

    TargetStudent = None
    TargetStudentUser = None
    TargetTeacher = None
    TargetTeacherUser = None
    TargetTeacherName = None
    TargetTeacherCode = None
    TargetTeacherId = None

    if target_type == "LEVEL":
        level = db.get(Level, target_id)
        if level:
            label = f"{level.level_code} - {level.level_name}"
    elif target_type == "STUDENT":
        TargetStudent = db.get(Student, target_id)
        TargetStudentUser = db.get(User, TargetStudent.user_id) if TargetStudent else None
        if TargetStudent and TargetStudentUser:
            label = f"{TargetStudentUser.full_name} ({TargetStudent.student_code})"

        if TargetStudent:
            StoredTeacherName = clean_text(TargetStudent.teacher)
            if hasattr(TargetStudent, "teacher_id") and getattr(TargetStudent, "teacher_id", None):
                TargetTeacher = db.get(Teacher, getattr(TargetStudent, "teacher_id"))
            if not TargetTeacher and StoredTeacherName:
                TargetTeacher = (
                    db.query(Teacher)
                    .join(User, Teacher.user_id == User.id)
                    .filter(User.full_name == StoredTeacherName)
                    .first()
                )
            TargetTeacherUser = db.get(User, TargetTeacher.user_id) if TargetTeacher else None
            TargetTeacherName = TargetTeacherUser.full_name if TargetTeacherUser else StoredTeacherName
            TargetTeacherCode = TargetTeacher.teacher_code if TargetTeacher else None
            TargetTeacherId = TargetTeacher.id if TargetTeacher else None
    elif target_type == "BATCH":
        label = f"Batch: {target_id}"

    student_name = None
    student_code = None
    student_class = None
    student_section = None
    if target_type == "STUDENT":
        student = TargetStudent or db.get(Student, target_id)
        student_user = TargetStudentUser or (db.get(User, student.user_id) if student else None)
        if student:
            student_name = student_user.full_name if student_user else None
            student_code = student.student_code
            student_class = student.class_name
            student_section = student.section

    return {
        "assignedToType": target_type,
        "assignedToId": target_id,
        "assignedToLabel": label,
        "targetStudentName": student_name,
        "targetStudentCode": student_code,
        "targetClassName": student_class,
        "targetSection": student_section,
        "targetTeacherId": TargetTeacherId,
        "targetTeacherCode": TargetTeacherCode,
        "targetTeacherName": TargetTeacherName,
        "assignedTeacherId": TargetTeacherId,
        "assignedTeacherCode": TargetTeacherCode,
        "assignedTeacherName": TargetTeacherName,
        "teacherId": TargetTeacherId,
        "teacherCode": TargetTeacherCode,
        "teacherName": TargetTeacherName,
    }


def attempt_date_payload(attempt: Attempt | None) -> dict:
    if not attempt:
        return {
            "startedAt": None,
            "submittedAt": None,
            "attemptDate": None,
            "completedDate": None,
        }

    return {
        "startedAt": attempt.started_at.isoformat() if attempt.started_at else None,
        "submittedAt": attempt.submitted_at.isoformat() if attempt.submitted_at else None,
        "attemptDate": attempt.started_at.isoformat() if attempt.started_at else None,
        "completedDate": attempt.submitted_at.isoformat() if attempt.submitted_at else None,
    }


def latest_attempt_for_student_assignment(db: Session, assignment_id: str, student_id: str) -> Attempt | None:
    return (
        db.query(Attempt)
        .filter(Attempt.assignment_id == assignment_id, Attempt.student_id == student_id)
        .order_by(Attempt.started_at.desc())
        .first()
    )


def active_reattempt_permission(db: Session, assignment_id: str, student_id: str):
    return (
        db.query(AssignmentReattemptPermission)
        .filter(
            AssignmentReattemptPermission.assignment_id == assignment_id,
            AssignmentReattemptPermission.student_id == student_id,
            AssignmentReattemptPermission.status == "APPROVED",
            AssignmentReattemptPermission.used_at.is_(None),
        )
        .order_by(AssignmentReattemptPermission.allowed_at.desc())
        .first()
    )


def latest_reattempt_permission(db: Session, assignment_id: str, student_id: str):
    return (
        db.query(AssignmentReattemptPermission)
        .filter(
            AssignmentReattemptPermission.assignment_id == assignment_id,
            AssignmentReattemptPermission.student_id == student_id,
            AssignmentReattemptPermission.status == "APPROVED",
        )
        .order_by(AssignmentReattemptPermission.allowed_at.desc())
        .first()
    )


def reattempt_permission_payload(permission: AssignmentReattemptPermission | None) -> dict:
    if not permission:
        return {
            "reattemptPermissionId": None,
            "reattemptStatus": "NONE",
            "reattemptAllowedAt": None,
            "reattemptUsedAt": None,
            "reattemptReason": None,
            "usedAssignmentId": None,
        }

    status = "USED" if permission.used_at else permission.status
    return {
        "reattemptPermissionId": permission.id,
        "reattemptStatus": status,
        "reattemptAllowedAt": permission.allowed_at.isoformat() if permission.allowed_at else None,
        "reattemptUsedAt": permission.used_at.isoformat() if permission.used_at else None,
        "reattemptReason": permission.reason,
        "usedAssignmentId": permission.used_assignment_id,
    }


BENCHMARK_PERCENTAGE = 70.0


def benchmark_payload_for_attempt(attempt: Attempt | None) -> dict:
    if not attempt or attempt.accuracy_percentage is None:
        return {
            "benchmarkPercentage": BENCHMARK_PERCENTAGE,
            "benchmarkStatus": "PENDING",
            "requiresAttention": False,
            "benchmarkMessage": "Benchmark will be calculated after submission.",
        }
    below = float(attempt.accuracy_percentage or 0) < BENCHMARK_PERCENTAGE
    return {
        "benchmarkPercentage": BENCHMARK_PERCENTAGE,
        "benchmarkStatus": "BELOW_BENCHMARK" if below else "PASS",
        "requiresAttention": below,
        "benchmarkMessage": (
            "Caution: This student scored below the minimum benchmark of 70%. Teacher intervention is required."
            if below
            else "Meets the minimum benchmark of 70%."
        ),
    }




def _admin_attempt_sequence_number(db: Session, attempt: Attempt | None) -> int:
    """Return display sequence: 1=Original, 2=Re-Attempt 1, 3=Re-Attempt 2."""
    if not attempt:
        return 1
    StoredAttemptNumber = getattr(attempt, "attempt_number", None)
    if StoredAttemptNumber is not None:
        try:
            return int(StoredAttemptNumber or 0) + 1
        except (TypeError, ValueError):
            pass

    Query = db.query(Attempt).filter(
        Attempt.student_id == attempt.student_id,
        Attempt.dps_id == attempt.dps_id,
    )
    AttemptGroupId = getattr(attempt, "attempt_group_id", None)
    if AttemptGroupId:
        Query = Query.filter(Attempt.attempt_group_id == AttemptGroupId)
    elif attempt.assignment_id:
        Query = Query.filter(Attempt.assignment_id == attempt.assignment_id)
    else:
        Query = Query.filter(Attempt.assignment_id.is_(None))
    Attempts = Query.order_by(Attempt.attempt_number.asc().nullslast(), Attempt.started_at.asc().nullslast(), Attempt.submitted_at.asc().nullslast(), Attempt.id.asc()).all()
    for Index, AttemptValue in enumerate(Attempts, start=1):
        if AttemptValue.id == attempt.id:
            return Index
    return 1


def _admin_attempt_display_status(db: Session, attempt: Attempt | None) -> str:
    if not attempt:
        return "Pending"
    if bool(getattr(attempt, "requires_manual_intervention", False)) or str(getattr(attempt, "benchmark_status", "") or "").upper() == "MANUAL_INTERVENTION_REQUIRED":
        return "Manual Review Required"
    SequenceNumber = _admin_attempt_sequence_number(db, attempt)
    IsReattempt = SequenceNumber > 1
    StatusValue = (attempt.status or "").upper()
    IsCompleted = StatusValue in {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"}
    if not IsCompleted:
        return "Re-Attempt Pending" if IsReattempt else "Pending"
    BenchmarkMet = float(attempt.accuracy_percentage or 0) >= BENCHMARK_PERCENTAGE
    if IsReattempt:
        return "Re-Attempt Cleared" if BenchmarkMet else "Needs Re-Attempt"
    return "Cleared" if BenchmarkMet else "Needs Re-Attempt"


def _admin_attempt_metadata(db: Session, attempt: Attempt | None) -> dict:
    SequenceNumber = _admin_attempt_sequence_number(db, attempt) if attempt else 1
    return {
        "attemptNumber": SequenceNumber,
        "attemptLabel": "Re-Attempt %s" % (SequenceNumber - 1) if SequenceNumber > 1 else "Original",
        "isReattempt": SequenceNumber > 1,
        "displayStatus": _admin_attempt_display_status(db, attempt),
    }


def all_attempts_for_student_assignment(db: Session, assignment_id: str, student_id: str) -> list[Attempt]:
    return (
        db.query(Attempt)
        .filter(Attempt.assignment_id == assignment_id, Attempt.student_id == student_id)
        .order_by(Attempt.attempt_number.asc().nullslast(), Attempt.started_at.asc().nullslast(), Attempt.submitted_at.asc().nullslast(), Attempt.id.asc())
        .all()
    )


def admin_attempt_history_for_assignment_student(db: Session, assignment: Assignment, student: Student) -> list[dict]:
    dps = db.get(DPS, assignment.dps_id)
    lesson = db.get(Lesson, dps.lesson_id) if dps else None
    level = db.get(Level, lesson.level_id) if lesson else None
    module = db.get(Module, level.module_id) if level else None
    History = []
    for Index, AttemptRow in enumerate(all_attempts_for_student_assignment(db, assignment.id, student.id), start=1):
        SequenceNumber = _admin_attempt_sequence_number(db, AttemptRow)
        History.append({
            "assignmentId": assignment.id,
            "assignmentTitle": assignment.title,
            "assignmentType": assignment.assignment_type,
            "studentId": student.id,
            "studentCode": student.student_code,
            "className": student.class_name,
            "section": student.section,
            "attemptId": AttemptRow.id,
            "attemptStatus": AttemptRow.status,
            "status": AttemptRow.status,
            "attemptNumber": SequenceNumber,
            "attemptSequence": SequenceNumber,
            "attemptLabel": "Re-Attempt %s" % (SequenceNumber - 1) if SequenceNumber > 1 else "Original",
            "isReattempt": SequenceNumber > 1,
            "score": AttemptRow.total_score,
            "totalMarks": AttemptRow.max_score,
            "maxScore": AttemptRow.max_score,
            "accuracy": AttemptRow.accuracy_percentage,
            "accuracyPercentage": AttemptRow.accuracy_percentage,
            "correctCount": AttemptRow.correct_count,
            "wrongCount": AttemptRow.wrong_count,
            "unansweredCount": AttemptRow.unanswered_count,
            "timeTakenSeconds": AttemptRow.time_taken_seconds,
            **benchmark_payload_for_attempt(AttemptRow),
            **attempt_date_payload(AttemptRow),
            "dpsId": dps.id if dps else assignment.dps_id,
            "dpsNumber": dps.dps_number if dps else None,
            "dpsTitle": dps.dps_title if dps else None,
            "lessonId": lesson.id if lesson else None,
            "lessonNumber": lesson.lesson_number if lesson else None,
            "lessonTitle": lesson.lesson_title if lesson else None,
            "levelId": level.id if level else None,
            "levelCode": level.level_code if level else None,
            "levelName": level.level_name if level else None,
            "moduleId": module.id if module else None,
            "moduleCode": module.module_code if module else None,
            "moduleName": module.module_name if module else None,
        })
    return History

def admin_level_total_dps_count(db: Session, level_id: str | None) -> int:
    if not level_id:
        return 0
    return (
        db.query(DPS)
        .join(Lesson, DPS.lesson_id == Lesson.id)
        .filter(Lesson.level_id == level_id)
        .count()
    )


def assignment_payload(db: Session, assignment: Assignment) -> dict:
    dps = db.get(DPS, assignment.dps_id)
    lesson = db.get(Lesson, dps.lesson_id) if dps else None
    level = db.get(Level, lesson.level_id) if lesson else None
    module = db.get(Module, level.module_id) if level else None
    required_dps_count = admin_level_total_dps_count(db, level.id if level else None)
    assigned_by = db.get(User, assignment.assigned_by_user_id) if assignment.assigned_by_user_id else None

    attempts = db.query(Attempt).filter(Attempt.assignment_id == assignment.id).all()
    completed_statuses = {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"}
    completed_attempts = [attempt for attempt in attempts if attempt.status in completed_statuses]
    in_progress_attempts = [attempt for attempt in attempts if attempt.status == "IN_PROGRESS"]

    average_accuracy = (
        round(sum([attempt.accuracy_percentage or 0 for attempt in completed_attempts]) / len(completed_attempts))
        if completed_attempts
        else 0
    )

    return {
        "assignmentId": assignment.id,
        "assignmentType": assignment.assignment_type,
        "title": assignment.title,
        "instructions": assignment.instructions,
        "isActive": assignment.is_active,
        "status": "ACTIVE" if assignment.is_active else "INACTIVE",
        "allowReattempt": assignment.allow_reattempt,
        "attemptGroupId": assignment.attempt_group_id,
        "assignmentSource": assignment.assignment_source,
        "retryAttemptNumber": assignment.retry_attempt_number,
        "attemptNumber": int(assignment.retry_attempt_number or 0),
        "attemptLabel": "Re-Attempt %s" % assignment.retry_attempt_number if int(assignment.retry_attempt_number or 0) > 0 else "Original",
        "isReattempt": int(assignment.retry_attempt_number or 0) > 0,
        "createdAt": assignment.created_at.isoformat() if assignment.created_at else None,
        "assignedByUserId": assignment.assigned_by_user_id,
        "assignedByName": assigned_by.full_name if assigned_by else "System",
        "assignedByRole": display_role(assigned_by.role if assigned_by else "SYSTEM"),
        **assignment_target_payload(db, assignment),
        "dpsId": dps.id if dps else assignment.dps_id,
        "dpsNumber": dps.dps_number if dps else None,
        "dpsTitle": dps.dps_title if dps else None,
        "lessonId": lesson.id if lesson else None,
        "lessonNumber": lesson.lesson_number if lesson else None,
        "lessonTitle": lesson.lesson_title if lesson else None,
        "levelId": level.id if level else None,
        "levelCode": level.level_code if level else None,
        "levelName": level.level_name if level else None,
        "requiredDpsCount": required_dps_count,
        "requiredDPSCount": required_dps_count,
        "totalDpsCount": required_dps_count,
        "totalDPSCount": required_dps_count,
        "levelDpsCount": required_dps_count,
        "levelDPSCount": required_dps_count,
        "moduleId": module.id if module else None,
        "moduleCode": module.module_code if module else None,
        "moduleName": module.module_name if module else None,
        "attemptCount": len(attempts),
        "completedAttemptCount": len(completed_attempts),
        "inProgressAttemptCount": len(in_progress_attempts),
        "pendingAttemptCount": max(len(attempts) - len(completed_attempts) - len(in_progress_attempts), 0),
        "averageAccuracy": average_accuracy,
        "latestCompletedAt": max(
            [attempt.submitted_at for attempt in completed_attempts if attempt.submitted_at],
            default=None,
        ).isoformat()
        if completed_attempts and any(attempt.submitted_at for attempt in completed_attempts)
        else None,
        "attemptHistory": (
            admin_attempt_history_for_assignment_student(db, assignment, db.get(Student, assignment.assigned_to_id))
            if assignment.assigned_to_type == "STUDENT" and db.get(Student, assignment.assigned_to_id)
            else []
        ),
    }




@router.get("/assessment-attempts/{attempt_id}/result")
def admin_assessment_attempt_result_route(
    attempt_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    attempt = db.get(AssessmentAttempt, attempt_id)
    if not attempt:
        api_error(404, "ASSESSMENT_ATTEMPT_NOT_FOUND", "Assessment attempt not found.")
    return AssessmentResultPayload(db, attempt, IncludeReview=True)


@router.post("/assessment-attempts/{attempt_id}/remarks")
def admin_save_assessment_attempt_remark(
    attempt_id: str,
    payload: AssessmentRemarkRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    attempt = db.get(AssessmentAttempt, attempt_id)
    if not attempt:
        api_error(404, "ASSESSMENT_ATTEMPT_NOT_FOUND", "Assessment attempt not found.")
    remark = upsert_assessment_remark(db, attempt=attempt, actor=user, actor_role="ADMIN", remark_text=payload.remarkText)
    db.commit()
    db.refresh(remark)
    return {"teacherFeedback": assessment_feedback_payload(db, remark)}


@router.get("/assessment-attempts/{attempt_id}/remarks")
def admin_get_assessment_attempt_remark(
    attempt_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    attempt = db.get(AssessmentAttempt, attempt_id)
    if not attempt:
        api_error(404, "ASSESSMENT_ATTEMPT_NOT_FOUND", "Assessment attempt not found.")
    return {"teacherFeedback": assessment_feedback_payload(db, active_assessment_remark(db, attempt.id))}


@router.delete("/assessment-attempts/{attempt_id}/remarks")
def admin_delete_assessment_attempt_remark(
    attempt_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    attempt = db.get(AssessmentAttempt, attempt_id)
    if not attempt:
        api_error(404, "ASSESSMENT_ATTEMPT_NOT_FOUND", "Assessment attempt not found.")
    remark = delete_assessment_remark(db, attempt=attempt, actor=user)
    db.commit()
    return {"deleted": True, "remarkId": remark.id}


@router.get("/assessments")
def list_assessments_route(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    engine_assignments = (
        db.query(AssessmentAssignment)
        .filter(AssessmentAssignment.is_active == True)
        .order_by(AssessmentAssignment.assigned_at.desc())
        .all()
    )
    engine_rows = [AssessmentAssignmentPayload(db, assignment) for assignment in engine_assignments]

    legacy_rows = (
        db.query(Assignment)
        .filter(Assignment.assignment_type == "ASSESSMENT")
        .order_by(Assignment.created_at.desc())
        .all()
    )
    legacy_payload = [assignment_payload(db, assignment) for assignment in legacy_rows]

    rows = engine_rows + legacy_payload
    return {"assessments": rows}


@router.post("/assessments")
def create_assessment_route(payload: AssignmentRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    resolved_target_id = resolve_assignment_target_id(db, payload.assignedToType, payload.assignedToId)

    assessment = create_assignment(
        db,
        assignment_type="ASSESSMENT",
        dps_id=payload.dpsId,
        assigned_by_user_id=user.id,
        assigned_to_type=payload.assignedToType,
        assigned_to_id=resolved_target_id,
        title=payload.title,
        instructions=payload.instructions or "Complete this assessment carefully. Admin approval is required for any reattempt.",
        allow_reattempt=False,
    )
    db.commit()
    return {
        "assessmentId": assessment.id,
        "assignmentId": assessment.id,
        "created": True,
        "message": "Assessment created successfully.",
    }


@router.patch("/assessments/{assessment_assignment_id}/status")
def update_assessment_assignment_status_route(
    assessment_assignment_id: str,
    payload: AssignmentStatusRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    engine_assignment = db.get(AssessmentAssignment, assessment_assignment_id)
    if engine_assignment:
        engine_assignment.is_active = payload.isActive
        db.commit()
        db.refresh(engine_assignment)
        return {
            "updated": True,
            "message": "Assessment assignment status updated.",
            "assignment": AssessmentAssignmentPayload(db, engine_assignment),
        }

    legacy_assignment = db.get(Assignment, assessment_assignment_id)
    if legacy_assignment and legacy_assignment.assignment_type == "ASSESSMENT":
        legacy_assignment.is_active = payload.isActive
        db.commit()
        db.refresh(legacy_assignment)
        return {
            "updated": True,
            "message": "Assessment assignment status updated.",
            "assignment": assignment_payload(db, legacy_assignment),
        }

    api_error(404, "NOT_FOUND", "Assessment assignment not found.")


@router.delete("/assessments/{assessment_assignment_id}")
def delete_assessment_assignment_route(
    assessment_assignment_id: str,
    force: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    engine_assignment = db.get(AssessmentAssignment, assessment_assignment_id)
    if engine_assignment:
        attempts = (
            db.query(AssessmentAttempt)
            .filter(AssessmentAttempt.assessment_assignment_id == engine_assignment.id)
            .all()
        )
        if attempts and not force:
            api_error(
                400,
                "ASSESSMENT_HAS_ATTEMPTS",
                "This assessment has attempt history. Send force=true only after strong admin confirmation.",
            )

        attempt_ids = [attempt.id for attempt in attempts]
        if attempt_ids:
            db.query(AssessmentAttemptAnswer).filter(
                AssessmentAttemptAnswer.assessment_attempt_id.in_(attempt_ids)
            ).delete(synchronize_session=False)
            db.query(AssessmentResult).filter(
                AssessmentResult.assessment_attempt_id.in_(attempt_ids)
            ).delete(synchronize_session=False)
            db.query(AssessmentReattemptApproval).filter(
                AssessmentReattemptApproval.assessment_attempt_id.in_(attempt_ids)
            ).delete(synchronize_session=False)

        db.query(AssessmentReattemptApproval).filter(
            AssessmentReattemptApproval.assessment_assignment_id == engine_assignment.id
        ).delete(synchronize_session=False)
        db.query(StudentLevelPromotion).filter(
            StudentLevelPromotion.assessment_assignment_id == engine_assignment.id
        ).delete(synchronize_session=False)

        if attempt_ids:
            db.query(AssessmentAttempt).filter(AssessmentAttempt.id.in_(attempt_ids)).delete(synchronize_session=False)

        db.delete(engine_assignment)
        db.commit()
        return {
            "deleted": True,
            "message": "Assessment assignment deleted.",
            "assignmentId": assessment_assignment_id,
        }

    legacy_assignment = db.get(Assignment, assessment_assignment_id)
    if legacy_assignment and legacy_assignment.assignment_type == "ASSESSMENT":
        return delete_assignment_route(assessment_assignment_id, force=force, db=db, user=user)

    api_error(404, "NOT_FOUND", "Assessment assignment not found.")


@router.get("/student-level-promotions")
def list_student_level_promotions_route(
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    return ListStudentLevelPromotions(db)


@router.post("/assessments/{assessment_assignment_id}/promote")
def promote_assessment_assignment_route(
    assessment_assignment_id: str,
    payload: AssessmentPromotionRequest | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    Result = PromoteAssessmentStudentToNextLevel(
        db,
        assessment_assignment_id,
        PromotedByUserId=user.id,
        TargetLevelId=payload.targetLevelId if payload else None,
        TargetLevelCode=payload.targetLevelCode if payload else None,
    )
    PromotionPayload = Result.get("promotion") if isinstance(Result, dict) else None
    PromotionId = PromotionPayload.get("promotionId") if isinstance(PromotionPayload, dict) else None
    if PromotionId and Result.get("promoted"):
        NotifyStudentPromoted(db, promotion_id=PromotionId, actor_user_id=user.id)
        db.commit()
    return Result


@router.get("/assignments")
def list_assignments_route(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    rows = (
        db.query(Assignment)
        .filter(Assignment.assignment_type != "ASSESSMENT")
        .order_by(Assignment.created_at.desc())
        .all()
    )
    return {"assignments": [assignment_payload(db, assignment) for assignment in rows]}


@router.get("/assignments/manual-intervention-queue")
def admin_manual_intervention_queue(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    rows = BuildManualInterventionQueue(db, BenchmarkPercentage=BENCHMARK_PERCENTAGE)
    return {
        "summary": {
            "manualInterventionRows": len(rows),
            "uniqueStudents": len({row.get("studentId") for row in rows if row.get("studentId")}),
            "uniqueDps": len({row.get("dpsId") for row in rows if row.get("dpsId")}),
        },
        "rows": rows,
    }


@router.get("/assignments/{assignment_id}")
def get_assignment_route(assignment_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    assignment = db.get(Assignment, assignment_id)
    if not assignment:
        api_error(404, "NOT_FOUND", "Assignment not found.")

    attempts = (
        db.query(Attempt)
        .filter(Attempt.assignment_id == assignment.id)
        .order_by(Attempt.started_at.desc())
        .all()
    )

    detail = assignment_payload(db, assignment)
    attempt_rows = []
    latest_attempt_by_student: dict[str, Attempt] = {}

    for attempt in attempts:
        student = db.get(Student, attempt.student_id)
        student_user = db.get(User, student.user_id) if student else None

        if attempt.student_id and attempt.student_id not in latest_attempt_by_student:
            latest_attempt_by_student[attempt.student_id] = attempt

        attempt_rows.append({
            "attemptId": attempt.id,
            "studentId": student.id if student else None,
            "studentName": student_user.full_name if student_user else "-",
            "studentCode": student.student_code if student else "-",
            "className": student.class_name if student else None,
            "section": student.section if student else None,
            "status": _admin_attempt_display_status(db, attempt),
            "attemptStatus": attempt.status,
            **_admin_attempt_metadata(db, attempt),
            "score": attempt.total_score,
            "maxScore": attempt.max_score,
            "accuracyPercentage": attempt.accuracy_percentage,
            "correct": attempt.correct_count,
            "wrong": attempt.wrong_count,
            "unanswered": attempt.unanswered_count,
            "timeTakenSeconds": attempt.time_taken_seconds,
            **benchmark_payload_for_attempt(attempt),
            **attempt_date_payload(attempt),
        })

    target_students = []
    if assignment.assigned_to_type == "STUDENT":
        student = db.get(Student, assignment.assigned_to_id)
        if student:
            target_students = [student]
    elif assignment.assigned_to_type == "LEVEL":
        target_students = (
            db.query(Student)
            .filter(Student.current_level_id == assignment.assigned_to_id)
            .order_by(Student.student_code.asc())
            .all()
        )
    else:
        target_students = []

    student_rows = []
    completed_statuses = {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"}

    for student in target_students:
        student_user = db.get(User, student.user_id)
        attempt = latest_attempt_by_student.get(student.id)
        completed = bool(attempt and attempt.status in completed_statuses)

        reattempt_permission = latest_reattempt_permission(db, assignment.id, student.id)
        row_status = "COMPLETED" if completed else "PENDING"
        if reattempt_permission and completed and not reattempt_permission.used_at:
            row_status = "REATTEMPT_AVAILABLE"

        student_rows.append({
            "studentId": student.id,
            "studentName": student_user.full_name if student_user else "-",
            "studentCode": student.student_code,
            "className": student.class_name,
            "section": student.section,
            "status": row_status,
            "attemptId": attempt.id if attempt else None,
            "attemptStatus": attempt.status if attempt else None,
            "score": attempt.total_score if attempt else None,
            "maxScore": attempt.max_score if attempt else None,
            "accuracyPercentage": attempt.accuracy_percentage if attempt else None,
            "correct": attempt.correct_count if attempt else None,
            "wrong": attempt.wrong_count if attempt else None,
            "unanswered": attempt.unanswered_count if attempt else None,
            "timeTakenSeconds": attempt.time_taken_seconds if attempt else None,
            "attemptHistory": admin_attempt_history_for_assignment_student(db, assignment, student),
            "attemptNumber": getattr(attempt, "attempt_number", None) if attempt else None,
            "requiresManualIntervention": bool(getattr(attempt, "requires_manual_intervention", False)) if attempt else False,
            "nextAttemptNumber": (int(getattr(attempt, "attempt_number", 0) or 0) + 1) if attempt else None,
            "retryAttemptNumber": (int(getattr(attempt, "attempt_number", 0) or 0) + 1) if reattempt_permission and getattr(reattempt_permission, "used_assignment_id", None) else None,
            **benchmark_payload_for_attempt(attempt),
            **attempt_date_payload(attempt),
            **reattempt_permission_payload(reattempt_permission),
        })

    return {
        "assignment": detail,
        "attempts": attempt_rows,
        "students": student_rows,
        "summary": {
            "assignedStudentCount": len(student_rows),
            "completedStudentCount": len([row for row in student_rows if row["status"] == "COMPLETED"]),
            "pendingStudentCount": len([row for row in student_rows if row["status"] == "PENDING"]),
            "attemptCount": len(attempt_rows),
        },
    }





@router.post("/assignments/{assignment_id}/students/{student_id}/allow-reattempt")
def allow_assignment_reattempt_route(
    assignment_id: str,
    student_id: str,
    payload: AllowReattemptRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    assignment = db.get(Assignment, assignment_id)
    if not assignment:
        api_error(404, "NOT_FOUND", "Assignment not found.")

    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")

    if assignment.assigned_to_type == "STUDENT" and assignment.assigned_to_id != student.id:
        api_error(403, "FORBIDDEN", "This student is not assigned to this assignment.")
    if assignment.assigned_to_type == "LEVEL" and assignment.assigned_to_id != student.current_level_id:
        api_error(403, "FORBIDDEN", "This student is not part of this assignment level.")

    latest_attempt = latest_attempt_for_student_assignment(db, assignment.id, student.id)
    if not latest_attempt or latest_attempt.status not in ["SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"]:
        api_error(400, "VALIDATION_ERROR", "Reattempt can be allowed only after a submitted attempt.")

    if not (bool(getattr(latest_attempt, "requires_manual_intervention", False)) or str(getattr(latest_attempt, "benchmark_status", "") or "").upper() == "MANUAL_INTERVENTION_REQUIRED"):
        api_error(400, "MANUAL_REATTEMPT_NOT_READY", "Manual re-attempt approval is available only after the automatic retry cycle requires teacher review.")

    existing = latest_reattempt_permission(db, assignment.id, student.id)
    if existing and existing.used_assignment_id:
        used_assignment = db.get(Assignment, existing.used_assignment_id)
        return {
            "created": False,
            "message": f"Re-Attempt {getattr(used_assignment, 'retry_attempt_number', None) or 'next'} has already been assigned for this exhausted attempt cycle.",
            "permission": reattempt_permission_payload(existing),
            "freshAssignmentId": used_assignment.id if used_assignment else existing.used_assignment_id,
            "freshAssignmentTitle": used_assignment.title if used_assignment else None,
        }

    if existing and not existing.used_at:
        permission = existing
    else:
        permission = AssignmentReattemptPermission(
            assignment_id=assignment.id,
            dps_id=assignment.dps_id,
            student_id=student.id,
            allowed_by_user_id=user.id,
            reason=payload.reason or "Additional re-attempt approved after all available attempts were used.",
            status="APPROVED",
        )
        db.add(permission)
        db.flush()

    fresh_assignment = BuildManualRetryAssignment(
        db,
        StudentItem=student,
        SourceAttempt=latest_attempt,
        AssignedByUserId=user.id,
        Instructions="A new re-attempt has been assigned for the same concept with a different question set.",
    )
    permission.status = "APPROVED"
    permission.used_at = datetime.now(datetime_timezone.utc)
    permission.used_assignment_id = fresh_assignment.id
    NotifyPracticeFreshPracticeAssigned(db, assignment_id=fresh_assignment.id, actor_user_id=user.id, source_attempt_id=latest_attempt.id)
    db.commit()
    db.refresh(permission)
    db.refresh(fresh_assignment)

    return {
        "created": True,
        "message": f"Re-Attempt {fresh_assignment.retry_attempt_number} assigned. The student will find the new sheet in the Practice tab.",
        "permission": reattempt_permission_payload(permission),
        "freshAssignmentId": fresh_assignment.id,
        "freshAssignmentTitle": fresh_assignment.title,
    }


@router.patch("/assignments/{assignment_id}/status")
def update_assignment_status_route(
    assignment_id: str,
    payload: AssignmentStatusRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    assignment = db.get(Assignment, assignment_id)
    if not assignment:
        api_error(404, "NOT_FOUND", "Assignment not found.")

    assignment.is_active = payload.isActive
    db.commit()
    db.refresh(assignment)

    return {
        "updated": True,
        "message": "Assignment status updated.",
        "assignment": assignment_payload(db, assignment),
    }


@router.delete("/assignments/{assignment_id}")
def delete_assignment_route(assignment_id: str, force: bool = False, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    assignment = db.get(Assignment, assignment_id)
    if not assignment:
        api_error(404, "NOT_FOUND", "Assignment not found.")

    attempts = db.query(Attempt).filter(Attempt.assignment_id == assignment.id).all()
    if attempts and not force:
        api_error(
            400,
            "ASSIGNMENT_HAS_ATTEMPTS",
            "This assignment has submitted/in-progress attempts. Send force=true only after strong admin confirmation.",
        )

    question_set_ids = [attempt.question_set_id for attempt in attempts if attempt.question_set_id]
    attempt_ids = [attempt.id for attempt in attempts]

    if attempt_ids:
        # Audit rows preserve the operational trail but must not block permanent removal
        # of a practice assignment/attempt during controlled admin cleanup.
        db.query(AuditLog).filter(AuditLog.attempt_id.in_(attempt_ids)).update(
            {AuditLog.attempt_id: None},
            synchronize_session=False,
        )
        db.query(AttemptAnswer).filter(AttemptAnswer.attempt_id.in_(attempt_ids)).delete(synchronize_session=False)
        # Break the Attempt -> GeneratedQuestionSet reference before deleting generated sets.
        db.query(Attempt).filter(Attempt.id.in_(attempt_ids)).update(
            {Attempt.question_set_id: None},
            synchronize_session=False,
        )
        db.query(Attempt).filter(Attempt.id.in_(attempt_ids)).delete(synchronize_session=False)

    if question_set_ids:
        db.query(GeneratedQuestionSet).filter(GeneratedQuestionSet.id.in_(question_set_ids)).update(
            {GeneratedQuestionSet.assignment_id: None},
            synchronize_session=False,
        )
        question_ids = [q.id for q in db.query(GeneratedQuestion).filter(GeneratedQuestion.question_set_id.in_(question_set_ids)).all()]
        if question_ids:
            db.query(QuestionOption).filter(QuestionOption.question_id.in_(question_ids)).delete(synchronize_session=False)
            db.query(GeneratedQuestion).filter(GeneratedQuestion.id.in_(question_ids)).delete(synchronize_session=False)
        db.query(GeneratedQuestionSet).filter(GeneratedQuestionSet.id.in_(question_set_ids)).delete(synchronize_session=False)

    db.query(AssignmentReattemptPermission).filter(AssignmentReattemptPermission.assignment_id == assignment.id).delete(synchronize_session=False)
    db.delete(assignment)
    db.commit()
    return {
        "deleted": True,
        "message": "Assignment and related attempt history deleted permanently." if attempts else "Assignment deleted successfully.",
        "assignmentId": assignment_id,
        "deletedAttempts": len(attempts),
    }


@router.post("/assignments")
def create_assignment_route(payload: AssignmentRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    resolved_target_id = resolve_assignment_target_id(db, payload.assignedToType, payload.assignedToId)
    dps = db.get(DPS, payload.dpsId)
    if not dps:
        api_error(404, "NOT_FOUND", "DPS not found.")
    if (payload.assignmentType or "PRACTICE").upper() == "PRACTICE" and (getattr(dps, "publication_status", "DRAFT") or "DRAFT") != "PUBLISHED":
        api_error(403, "DPS_NOT_PUBLISHED", "This DPS has not been published from Learning Path Studio yet.")

    assignment = create_assignment(
        db,
        assignment_type=payload.assignmentType,
        dps_id=payload.dpsId,
        assigned_by_user_id=user.id,
        assigned_to_type=payload.assignedToType,
        assigned_to_id=resolved_target_id,
        title=payload.title,
        instructions=payload.instructions,
        allow_reattempt=payload.allowReattempt,
    )
    db.commit()
    return {"assignmentId": assignment.id, "created": True, "message": "Assignment created successfully."}


def _admin_teacher_context_for_student(db: Session, student: Student | None) -> dict:
    if not student:
        return {"teacherId": None, "teacherName": None, "teacherCode": None}
    teacher_name = clean_text(getattr(student, "teacher", None))
    if not teacher_name:
        return {"teacherId": None, "teacherName": None, "teacherCode": None}
    teacher = (
        db.query(Teacher)
        .join(User, Teacher.user_id == User.id)
        .filter((User.full_name == teacher_name) | (Teacher.teacher_code == teacher_name))
        .first()
    )
    teacher_user = db.get(User, teacher.user_id) if teacher else None
    return {
        "teacherId": teacher.id if teacher else None,
        "teacherName": teacher_user.full_name if teacher_user else teacher_name,
        "teacherCode": teacher.teacher_code if teacher else None,
    }


def _admin_student_matches_teacher(db: Session, student: Student | None, teacher_id: str | None) -> bool:
    if not teacher_id:
        return True
    teacher = db.get(Teacher, teacher_id)
    if not teacher:
        return False
    teacher_user = db.get(User, teacher.user_id)
    teacher_name = teacher_user.full_name if teacher_user else None
    student_teacher = clean_text(getattr(student, "teacher", None)) if student else None
    return bool(student_teacher and student_teacher in {teacher_name, teacher.teacher_code})


@router.get("/dps/{dps_id}/results")
def dps_results(dps_id: str, teacherId: str | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    attempts = db.query(Attempt).filter(Attempt.dps_id == dps_id).all()
    results = []
    for a in attempts:
        student = db.get(Student, a.student_id)
        if not _admin_student_matches_teacher(db, student, teacherId):
            continue
        student_user = db.get(User, student.user_id) if student else None
        teacher_context = _admin_teacher_context_for_student(db, student)
        results.append({
            "attemptId": a.id,
            "studentId": a.student_id,
            "studentName": student_user.full_name if student_user else None,
            "studentCode": student.student_code if student else None,
            **teacher_context,
            "score": a.total_score,
            "maxScore": a.max_score,
            "accuracyPercentage": a.accuracy_percentage,
            "correct": a.correct_count,
            "wrong": a.wrong_count,
            "unanswered": a.unanswered_count,
            "timeTakenSeconds": a.time_taken_seconds,
            "status": _admin_attempt_display_status(db, a),
            "attemptStatus": a.status,
            **_admin_attempt_metadata(db, a),
            **benchmark_payload_for_attempt(a),
            **attempt_date_payload(a),
        })
    avg_acc = sum(r["accuracyPercentage"] for r in results) / len(results) if results else 0
    avg_score = sum(r["score"] for r in results) / len(results) if results else 0
    return {"dpsId": dps_id, "summary": {"totalAttempts": len(results), "averageScore": round(avg_score), "averageAccuracy": round(avg_acc)}, "results": results}


@router.get("/results/level")
def level_results(levelId: str, moduleId: str | None = None, teacherId: str | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    level = db.get(Level, levelId)
    if not level:
        api_error(404, "NOT_FOUND", "Level not found.")
    if moduleId and level.module_id != moduleId:
        api_error(400, "INVALID_SCOPE", "Selected level does not belong to the selected module.")

    module = db.get(Module, level.module_id) if level else None
    lessons = db.query(Lesson).filter(Lesson.level_id == levelId).order_by(Lesson.display_order.asc(), Lesson.lesson_number.asc()).all()
    lesson_ids = [lesson.id for lesson in lessons]
    dps_items = db.query(DPS).filter(DPS.lesson_id.in_(lesson_ids)).order_by(DPS.dps_number.asc()).all() if lesson_ids else []
    dps_ids = [dps.id for dps in dps_items]

    attempt_rows = db.query(Attempt).filter(Attempt.dps_id.in_(dps_ids)).all() if dps_ids else []
    students_by_id: dict[str, Student] = {}
    if dps_ids:
        for attempt in attempt_rows:
            student = db.get(Student, attempt.student_id)
            if student and _admin_student_matches_teacher(db, student, teacherId):
                students_by_id[student.id] = student

    enrolled_query = db.query(Student).filter(Student.current_level_id == levelId)
    for student in enrolled_query.all():
        if _admin_student_matches_teacher(db, student, teacherId):
            students_by_id[student.id] = student

    latest_attempt_by_student_dps: dict[tuple[str, str], Attempt] = {}
    for attempt in attempt_rows:
        student = students_by_id.get(attempt.student_id)
        if not student:
            continue
        key = (attempt.student_id, attempt.dps_id)
        current = latest_attempt_by_student_dps.get(key)
        attempt_date = attempt.submitted_at or attempt.started_at
        current_date = (current.submitted_at or current.started_at) if current else None
        if not current or (attempt_date and current_date and attempt_date > current_date) or (attempt_date and not current_date):
            latest_attempt_by_student_dps[key] = attempt

    required_dps = len(dps_items)
    rows = []
    for student in sorted(students_by_id.values(), key=lambda item: item.student_code or ""):
        student_user = db.get(User, student.user_id)
        teacher_context = _admin_teacher_context_for_student(db, student)
        student_attempts = [attempt for (student_id, _), attempt in latest_attempt_by_student_dps.items() if student_id == student.id]
        completed_attempts = [attempt for attempt in student_attempts if (attempt.status or "").upper() in {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"}]
        passed_attempts = [attempt for attempt in completed_attempts if float(attempt.accuracy_percentage or 0) >= 70]
        needs_reattempt = NeedsReattemptAttempts(completed_attempts, BENCHMARK_PERCENTAGE)
        assigned_dps = len({attempt.dps_id for attempt in student_attempts})
        completed_dps = len({attempt.dps_id for attempt in completed_attempts})
        pending_dps = max(required_dps - completed_dps, 0)
        avg_score = round(sum(float(attempt.total_score or 0) for attempt in completed_attempts) / len(completed_attempts)) if completed_attempts else 0
        avg_accuracy = round(sum(float(attempt.accuracy_percentage or 0) for attempt in completed_attempts) / len(completed_attempts)) if completed_attempts else 0
        if avg_accuracy >= 90:
            performance_zone = "Excellence Zone"
        elif avg_accuracy >= 70:
            performance_zone = "Growth Zone"
        else:
            performance_zone = "Needs Improvement" if completed_attempts else "Not Started"
        readiness = "Ready" if required_dps > 0 and completed_dps >= required_dps and len(needs_reattempt) == 0 else "Not Ready"
        last_activity = max([(attempt.submitted_at or attempt.started_at) for attempt in completed_attempts if (attempt.submitted_at or attempt.started_at)], default=None)
        rows.append({
            "studentId": student.id,
            "studentName": student_user.full_name if student_user else None,
            "studentCode": student.student_code,
            **teacher_context,
            "moduleId": module.id if module else None,
            "moduleCode": module.module_code if module else None,
            "moduleName": module.module_name if module else None,
            "levelId": level.id,
            "levelCode": level.level_code,
            "levelName": level.level_name,
            "requiredDps": required_dps,
            "assignedDps": assigned_dps,
            "completedDps": completed_dps,
            "passedDps": len(passed_attempts),
            "pendingDps": pending_dps,
            "needsReattempt": CountNeedsReattemptConcepts(completed_attempts, BENCHMARK_PERCENTAGE),
            "averageScore": avg_score,
            "averageAccuracy": avg_accuracy,
            "performanceZone": performance_zone,
            "assessmentReadiness": readiness,
            "lastActivity": last_activity.isoformat() if last_activity else None,
        })

    avg_accuracy = round(sum(float(row["averageAccuracy"] or 0) for row in rows) / len(rows)) if rows else 0
    ready_count = len([row for row in rows if row["assessmentReadiness"] == "Ready"])
    return {
        "moduleId": module.id if module else None,
        "levelId": level.id,
        "summary": {
            "students": len(rows),
            "ready": ready_count,
            "notReady": max(len(rows) - ready_count, 0),
            "averageAccuracy": avg_accuracy,
            "requiredDps": required_dps,
        },
        "results": rows,
    }



def _admin_scope_for_dps(db: Session, dps_id: str) -> dict:
    dps = db.get(DPS, dps_id)
    lesson = db.get(Lesson, dps.lesson_id) if dps else None
    level = db.get(Level, lesson.level_id) if lesson else None
    module = db.get(Module, level.module_id) if level else None
    return {
        "dps": dps,
        "lesson": lesson,
        "level": level,
        "module": module,
        "dpsLabel": f"DPS {dps.dps_number} - {dps.dps_title}" if dps else "-",
        "lessonLabel": f"Lesson {lesson.lesson_number} - {lesson.lesson_title}" if lesson else "-",
        "levelLabel": f"{level.level_code} - {level.level_name}" if level else "-",
        "moduleLabel": f"{module.module_code} - {module.module_name}" if module else "-",
    }


def _admin_scope_for_level(db: Session, level_id: str) -> dict:
    level = db.get(Level, level_id)
    module = db.get(Module, level.module_id) if level else None
    return {
        "level": level,
        "module": module,
        "levelLabel": f"{level.level_code} - {level.level_name}" if level else "-",
        "moduleLabel": f"{module.module_code} - {module.module_name}" if module else "-",
    }


INDIA_REPORT_TIMEZONE = datetime_timezone(timedelta(hours=5, minutes=30), name="IST")
INDIA_REPORT_TIMEZONE_NAME = "Asia/Kolkata"


def _admin_report_timezone(TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> tzinfo:
    return INDIA_REPORT_TIMEZONE


def _admin_report_timezone_name(TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> str:
    return INDIA_REPORT_TIMEZONE_NAME


def _admin_parse_report_datetime(Value: Any, TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> datetime | None:
    if not Value or Value == "-":
        return None
    if isinstance(Value, datetime):
        ParsedValue = Value
    else:
        TextValue = str(Value).strip()
        if not TextValue:
            return None
        try:
            ParsedValue = datetime.fromisoformat(TextValue.replace("Z", "+00:00"))
        except ValueError:
            return None
    if ParsedValue.tzinfo is None:
        ParsedValue = ParsedValue.replace(tzinfo=datetime_timezone.utc)
    return ParsedValue.astimezone(_admin_report_timezone(TimezoneName, TimezoneOffsetMinutes))


def _admin_report_date(Value: Any, TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> str:
    ParsedValue = _admin_parse_report_datetime(Value, TimezoneName, TimezoneOffsetMinutes)
    return ParsedValue.strftime("%d-%b-%Y") if ParsedValue else "-"


def _admin_report_time(Value: Any, TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> str:
    ParsedValue = _admin_parse_report_datetime(Value, TimezoneName, TimezoneOffsetMinutes)
    return ParsedValue.strftime("%I:%M %p") if ParsedValue else "-"


def _admin_report_datetime(Value: Any, TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> str:
    ParsedValue = _admin_parse_report_datetime(Value, TimezoneName, TimezoneOffsetMinutes)
    return ParsedValue.strftime("%d-%b-%Y, %I:%M %p") if ParsedValue else "-"


def _admin_report_duration(Value: Any) -> str:
    try:
        TotalSeconds = int(round(float(Value or 0)))
    except (TypeError, ValueError):
        return "-"
    if TotalSeconds < 60:
        return f"{TotalSeconds} sec"
    Minutes, Seconds = divmod(TotalSeconds, 60)
    return f"{Minutes} min {Seconds:02d} sec" if Seconds else f"{Minutes} min"

def _admin_promotion_records_for_scope(
    db: Session,
    student_id: str | None = None,
    teacher_id: str | None = None,
    module_id: str | None = None,
    level_id: str | None = None,
) -> list[StudentLevelPromotion]:
    Query = db.query(StudentLevelPromotion).filter(StudentLevelPromotion.status == "PROMOTED")
    if student_id:
        Query = Query.filter(StudentLevelPromotion.student_id == student_id)
    if module_id:
        Query = Query.filter(
            (StudentLevelPromotion.from_module_id == module_id)
            | (StudentLevelPromotion.to_module_id == module_id)
        )
    if level_id:
        Query = Query.filter(
            (StudentLevelPromotion.from_level_id == level_id)
            | (StudentLevelPromotion.to_level_id == level_id)
        )
    Promotions = Query.order_by(
        StudentLevelPromotion.promoted_at.desc().nullslast(),
        StudentLevelPromotion.created_at.desc(),
    ).all()
    if not teacher_id:
        return Promotions
    Filtered: list[StudentLevelPromotion] = []
    for Promotion in Promotions:
        StudentValue = db.get(Student, Promotion.student_id)
        if _admin_student_matches_teacher(db, StudentValue, teacher_id):
            Filtered.append(Promotion)
    return Filtered


def _admin_promotion_payload_for_report(db: Session, Promotion: StudentLevelPromotion) -> dict[str, Any]:
    StudentValue = db.get(Student, Promotion.student_id) if Promotion else None
    StudentUser = db.get(User, StudentValue.user_id) if StudentValue else None
    PromotedBy = db.get(User, Promotion.promoted_by_user_id) if Promotion and Promotion.promoted_by_user_id else None
    AssignmentValue = db.get(AssessmentAssignment, Promotion.assessment_assignment_id) if Promotion and Promotion.assessment_assignment_id else None
    BlueprintValue = db.get(AssessmentBlueprint, AssignmentValue.blueprint_id) if AssignmentValue else None
    return {
        "promotionId": Promotion.id,
        "studentId": Promotion.student_id,
        "studentName": StudentUser.full_name if StudentUser else "-",
        "studentCode": Promotion.student_code or (StudentValue.student_code if StudentValue else "-"),
        "fromModuleId": Promotion.from_module_id,
        "fromModuleCode": Promotion.from_module_code,
        "fromLevelId": Promotion.from_level_id,
        "fromLevelCode": Promotion.from_level_code,
        "toModuleId": Promotion.to_module_id,
        "toModuleCode": Promotion.to_module_code,
        "toLevelId": Promotion.to_level_id,
        "toLevelCode": Promotion.to_level_code,
        "assessmentAssignmentId": Promotion.assessment_assignment_id,
        "assessmentAttemptId": Promotion.assessment_attempt_id,
        "assessmentResultId": Promotion.assessment_result_id,
        "assessmentTitle": BlueprintValue.title if BlueprintValue else "Assessment",
        "score": Promotion.score,
        "maxScore": Promotion.max_score,
        "percentage": Promotion.percentage,
        "promotionStatus": "Promoted",
        "promotedAt": Promotion.promoted_at.isoformat() if Promotion.promoted_at else None,
        "promotedByName": PromotedBy.full_name if PromotedBy else "Admin",
        "promotedByUserId": Promotion.promoted_by_user_id,
    }


def _admin_latest_promotion_for_level(
    Promotions: list[StudentLevelPromotion],
    LevelId: str | None,
) -> StudentLevelPromotion | None:
    if not LevelId:
        return None
    Matching = [
        Promotion
        for Promotion in Promotions
        if Promotion.from_level_id == LevelId
    ]
    return Matching[0] if Matching else None


def _admin_dps_export_rows(db: Session, dps_id: str, teacher_id: str | None, TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> tuple[list[dict], dict]:
    Payload = dps_results(dps_id, teacher_id, db)
    Scope = _admin_scope_for_dps(db, dps_id)
    Rows = []
    for AttemptRow in Payload.get("results", []):
        CompletedValue = AttemptRow.get("completedDate") or AttemptRow.get("submittedAt")
        Rows.append({
            "Student Name": AttemptRow.get("studentName") or "-",
            "Student Code": AttemptRow.get("studentCode") or "-",
            "Teacher Name": AttemptRow.get("teacherName") or "Not Assigned",
            "Teacher Code": AttemptRow.get("teacherCode") or "-",
            "Status": AttemptRow.get("displayStatus") or AttemptRow.get("status") or "-",
            "Score": AttemptRow.get("score") or 0,
            "Total Marks": AttemptRow.get("maxScore") or 0,
            "Accuracy %": AttemptRow.get("accuracyPercentage") or 0,
            "Benchmark Status": "Needs Improvement" if AttemptRow.get("requiresAttention") else "Benchmark Met",
            "Correct Answers": AttemptRow.get("correct") or 0,
            "Completed Date": _admin_report_date(CompletedValue, TimezoneName, TimezoneOffsetMinutes),
            "Completion Time": _admin_report_time(CompletedValue, TimezoneName, TimezoneOffsetMinutes),
            "Time Taken": _admin_report_duration(AttemptRow.get("timeTakenSeconds")),
        })
    return Rows, {**Payload.get("summary", {}), **Scope}


def _admin_level_export_rows(db: Session, module_id: str | None, level_id: str, teacher_id: str | None, TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> tuple[list[dict], dict]:
    Payload = level_results(level_id, module_id, teacher_id, db)
    Scope = _admin_scope_for_level(db, level_id)
    Rows = []
    for Row in Payload.get("results", []):
        Rows.append({
            "Student Name": Row.get("studentName") or "-",
            "Student Code": Row.get("studentCode") or "-",
            "Teacher Name": Row.get("teacherName") or "Not Assigned",
            "Teacher Code": Row.get("teacherCode") or "-",
            "Required DPS": Row.get("requiredDps") or 0,
            "Completed DPS": Row.get("completedDps") or 0,
            "Passed DPS": Row.get("passedDps") or 0,
            "Pending DPS": Row.get("pendingDps") or 0,
            "Needs Re-Attempt": Row.get("needsReattempt") or 0,
            "Average Score": Row.get("averageScore") or 0,
            "Average Accuracy %": Row.get("averageAccuracy") or 0,
            "Performance Zone": Row.get("performanceZone") or "-",
            "Assessment Readiness": Row.get("assessmentReadiness") or "-",
            "Last Activity Date": _admin_report_date(Row.get("lastActivity"), timezone, timezoneOffsetMinutes),
        })
    return Rows, {**Payload.get("summary", {}), **Scope}


def _admin_attempt_scope(db: Session, attempt: Attempt) -> dict:
    dps = db.get(DPS, attempt.dps_id)
    lesson = db.get(Lesson, dps.lesson_id) if dps else None
    level = db.get(Level, lesson.level_id) if lesson else None
    module = db.get(Module, level.module_id) if level else None
    assignment = db.get(Assignment, attempt.assignment_id) if attempt.assignment_id else None
    return {"dps": dps, "lesson": lesson, "level": level, "module": module, "assignment": assignment}


def _admin_assessment_attempt_status(AttemptValue: AssessmentAttempt | None, ResultValue: AssessmentResult | None = None) -> str:
    if not AttemptValue:
        return "Pending"
    StatusValue = (AttemptValue.status or "").upper()
    IsCompleted = StatusValue in {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED", "CLEARED", "NEEDS_RE_ATTEMPT"}
    IsReattempt = (AttemptValue.attempt_type or "").upper() in {"RE_ATTEMPT", "REATTEMPT", "RE-ATTEMPT"} or int(AttemptValue.attempt_number or 1) > 1
    if not IsCompleted:
        return "Re-Attempt Pending" if IsReattempt else "Pending"
    PercentageValue = float(
        ResultValue.percentage
        if ResultValue and ResultValue.percentage is not None
        else AttemptValue.percentage or 0
    )
    Cleared = bool(ResultValue.cleared) if ResultValue else PercentageValue >= BENCHMARK_PERCENTAGE
    if IsReattempt:
        return "Re-Attempt Cleared" if Cleared else "Needs Re-Attempt"
    return "Cleared" if Cleared else "Needs Re-Attempt"


def _admin_assessment_attempt_label(AttemptValue: AssessmentAttempt | None) -> str:
    if not AttemptValue:
        return "Original"
    AttemptNumber = int(AttemptValue.attempt_number or 1)
    AttemptType = (AttemptValue.attempt_type or "ORIGINAL").upper()
    if AttemptType in {"RE_ATTEMPT", "REATTEMPT", "RE-ATTEMPT"} or AttemptNumber > 1:
        return f"Re-Attempt {max(AttemptNumber - 1, 1)}"
    return "Original"


def _admin_assessment_attempt_date_payload(AttemptValue: AssessmentAttempt | None, ResultValue: AssessmentResult | None = None) -> dict:
    if not AttemptValue:
        return {
            "startedAt": None,
            "submittedAt": None,
            "attemptDate": None,
            "completedDate": None,
        }
    CompletedAt = ResultValue.completion_date if ResultValue and ResultValue.completion_date else AttemptValue.submitted_at
    return {
        "startedAt": AttemptValue.started_at.isoformat() if AttemptValue.started_at else None,
        "submittedAt": AttemptValue.submitted_at.isoformat() if AttemptValue.submitted_at else None,
        "attemptDate": AttemptValue.started_at.isoformat() if AttemptValue.started_at else None,
        "completedDate": CompletedAt.isoformat() if CompletedAt else None,
    }


def _admin_assessment_attempt_report_rows(
    db: Session,
    student: Student,
    teacher_context: dict,
    module_id: str | None = None,
    level_id: str | None = None,
    lesson_id: str | None = None,
    dps_id: str | None = None,
) -> list[dict]:
    if dps_id:
        return []

    LessonScopeLevelId = None
    if lesson_id:
        LessonScope = db.get(Lesson, lesson_id)
        LessonScopeLevelId = LessonScope.level_id if LessonScope else None

    Attempts = (
        db.query(AssessmentAttempt)
        .filter(AssessmentAttempt.student_id == student.id)
        .order_by(AssessmentAttempt.started_at.desc().nullslast(), AssessmentAttempt.submitted_at.desc().nullslast(), AssessmentAttempt.id.desc())
        .all()
    )
    Rows: list[dict] = []
    for AttemptValue in Attempts:
        AssignmentValue = db.get(AssessmentAssignment, AttemptValue.assessment_assignment_id)
        VersionValue = db.get(AssessmentVersion, AttemptValue.assessment_version_id)
        BlueprintValue = db.get(AssessmentBlueprint, AssignmentValue.blueprint_id) if AssignmentValue else None
        if not BlueprintValue and VersionValue:
            BlueprintValue = db.get(AssessmentBlueprint, VersionValue.blueprint_id)
        LevelValue = db.get(Level, BlueprintValue.level_id) if BlueprintValue else None
        ModuleValue = db.get(Module, BlueprintValue.module_id) if BlueprintValue else None
        if module_id and (not ModuleValue or ModuleValue.id != module_id):
            continue
        if level_id and (not LevelValue or LevelValue.id != level_id):
            continue
        if LessonScopeLevelId and (not LevelValue or LevelValue.id != LessonScopeLevelId):
            continue

        ResultValue = (
            db.query(AssessmentResult)
            .filter(AssessmentResult.assessment_attempt_id == AttemptValue.id)
            .first()
        )
        TeacherValue = db.get(Teacher, AssignmentValue.teacher_id) if AssignmentValue and AssignmentValue.teacher_id else None
        TeacherUser = db.get(User, TeacherValue.user_id) if TeacherValue else None
        MaxScore = float(ResultValue.max_score) if ResultValue else float(AttemptValue.max_score or (VersionValue.total_marks if VersionValue else 100))
        RawScore = float(ResultValue.score) if ResultValue else float(AttemptValue.total_score or 0)
        Score = NormalizeAssessmentScore(RawScore, MaxScore)
        Percentage = NormalizeAssessmentPercentage(Score, MaxScore) if MaxScore else NormalizeAssessmentPercentage(ResultValue.percentage if ResultValue else AttemptValue.percentage or 0, 100)
        Cleared = bool(ResultValue.cleared) if ResultValue else Percentage >= BENCHMARK_PERCENTAGE
        DisplayStatus = _admin_assessment_attempt_status(AttemptValue, ResultValue)
        Rows.append({
            "attemptId": AttemptValue.id,
            "assessmentAttemptId": AttemptValue.id,
            "assignmentId": AttemptValue.assessment_assignment_id,
            "assessmentAssignmentId": AttemptValue.assessment_assignment_id,
            "assignmentType": "ASSESSMENT",
            "assessmentId": BlueprintValue.id if BlueprintValue else None,
            "blueprintId": BlueprintValue.id if BlueprintValue else None,
            "assessmentVersionId": AttemptValue.assessment_version_id,
            "assessmentTitle": BlueprintValue.title if BlueprintValue else "Assessment",
            "assignmentTitle": BlueprintValue.title if BlueprintValue else "Assessment",
            "versionNumber": VersionValue.version_number if VersionValue else None,
            "moduleId": ModuleValue.id if ModuleValue else None,
            "moduleCode": ModuleValue.module_code if ModuleValue else None,
            "moduleName": ModuleValue.module_name if ModuleValue else None,
            "levelId": LevelValue.id if LevelValue else None,
            "levelCode": LevelValue.level_code if LevelValue else None,
            "levelName": LevelValue.level_name if LevelValue else None,
            "lessonId": None,
            "lessonNumber": None,
            "lessonTitle": None,
            "dpsId": None,
            "dpsNumber": None,
            "dpsTitle": None,
            "teacherId": TeacherValue.id if TeacherValue else teacher_context.get("teacherId"),
            "teacherName": TeacherUser.full_name if TeacherUser else teacher_context.get("teacherName") or "Not Assigned",
            "teacherCode": TeacherValue.teacher_code if TeacherValue else teacher_context.get("teacherCode"),
            "status": AttemptValue.status,
            "attemptStatus": AttemptValue.status,
            "displayStatus": DisplayStatus,
            "attemptNumber": AttemptValue.attempt_number,
            "attemptType": AttemptValue.attempt_type,
            "attemptLabel": _admin_assessment_attempt_label(AttemptValue),
            "isReattempt": DisplayStatus.startswith("Re-Attempt") or int(AttemptValue.attempt_number or 1) > 1,
            "score": Score,
            "maxScore": NormalizeAssessmentScore(MaxScore, MaxScore),
            "totalMarks": NormalizeAssessmentScore(MaxScore, MaxScore),
            "accuracyPercentage": Percentage,
            "percentage": Percentage,
            "correct": AttemptValue.correct_count,
            "wrong": AttemptValue.wrong_count,
            "unanswered": AttemptValue.unanswered_count,
            "timeTakenSeconds": AttemptValue.time_taken_seconds,
            "benchmarkPercentage": BENCHMARK_PERCENTAGE,
            "benchmarkStatus": "PASS" if Cleared else "BELOW_BENCHMARK",
            "requiresAttention": not Cleared,
            "benchmarkMessage": "Assessment meets the 70% benchmark." if Cleared else "Assessment is below benchmark and needs re-attempt support.",
            "createdAt": AssignmentValue.assigned_at.isoformat() if AssignmentValue and AssignmentValue.assigned_at else None,
            "assignedAt": AssignmentValue.assigned_at.isoformat() if AssignmentValue and AssignmentValue.assigned_at else None,
            **_admin_assessment_attempt_date_payload(AttemptValue, ResultValue),
        })
    return Rows


def _admin_build_student_report(db: Session, student_id: str, module_id: str | None = None, level_id: str | None = None, lesson_id: str | None = None, dps_id: str | None = None) -> dict:
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")
    student_user = db.get(User, student.user_id)
    teacher_context = _admin_teacher_context_for_student(db, student)
    promotion_records = _admin_promotion_records_for_scope(
        db,
        student_id=student.id,
        module_id=module_id,
        level_id=level_id,
    )
    promotion_rows = [_admin_promotion_payload_for_report(db, Promotion) for Promotion in promotion_records]
    all_attempts = db.query(Attempt).filter(Attempt.student_id == student.id).order_by(Attempt.started_at.desc()).all()

    filtered_attempts = []
    for attempt in all_attempts:
        scope = _admin_attempt_scope(db, attempt)
        level = scope.get("level")
        module = scope.get("module")
        lesson = scope.get("lesson")
        dps = scope.get("dps")
        if module_id and (not module or module.id != module_id):
            continue
        if level_id and (not level or level.id != level_id):
            continue
        if lesson_id and (not lesson or lesson.id != lesson_id):
            continue
        if dps_id and (not dps or dps.id != dps_id):
            continue
        filtered_attempts.append((attempt, scope))

    dps_rows = []
    assessment_rows = []
    completed_attempts = []
    level_ids: set[str] = set()
    for attempt, scope in filtered_attempts:
        dps = scope.get("dps")
        lesson = scope.get("lesson")
        level = scope.get("level")
        module = scope.get("module")
        assignment = scope.get("assignment")
        if level:
            level_ids.add(level.id)
        is_completed = (attempt.status or "").upper() in {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"}
        if is_completed:
            completed_attempts.append(attempt)
        common = {
            "attemptId": attempt.id,
            "assignmentType": (assignment.assignment_type if assignment else attempt.mode) or "PRACTICE",
            **teacher_context,
            "moduleId": module.id if module else None,
            "moduleCode": module.module_code if module else None,
            "moduleName": module.module_name if module else None,
            "levelId": level.id if level else None,
            "levelCode": level.level_code if level else None,
            "levelName": level.level_name if level else None,
            "lessonId": lesson.id if lesson else None,
            "lessonNumber": lesson.lesson_number if lesson else None,
            "lessonTitle": lesson.lesson_title if lesson else None,
            "dpsId": dps.id if dps else None,
            "dpsNumber": dps.dps_number if dps else None,
            "dpsTitle": dps.dps_title if dps else None,
            "status": _admin_attempt_display_status(db, attempt),
            "attemptStatus": attempt.status,
            **_admin_attempt_metadata(db, attempt),
            "score": attempt.total_score,
            "maxScore": attempt.max_score,
            "accuracyPercentage": attempt.accuracy_percentage,
            "correct": attempt.correct_count,
            "wrong": attempt.wrong_count,
            "unanswered": attempt.unanswered_count,
            "timeTakenSeconds": attempt.time_taken_seconds,
            **benchmark_payload_for_attempt(attempt),
            **attempt_date_payload(attempt),
        }
        if ((assignment.assignment_type if assignment else "PRACTICE") or "PRACTICE").upper() == "ASSESSMENT":
            assessment_rows.append(common)
        else:
            dps_rows.append(common)

    assessment_attempt_rows = _admin_assessment_attempt_report_rows(
        db,
        student,
        teacher_context,
        module_id=module_id,
        level_id=level_id,
        lesson_id=lesson_id,
        dps_id=dps_id,
    )
    ExistingAssessmentAttemptIds = {Row.get("attemptId") for Row in assessment_rows if Row.get("attemptId")}
    assessment_rows.extend([
        Row for Row in assessment_attempt_rows
        if Row.get("attemptId") not in ExistingAssessmentAttemptIds
    ])
    completed_assessment_rows = [
        Row for Row in assessment_rows
        if (Row.get("completedDate") or Row.get("submittedAt"))
    ]
    for Row in assessment_rows:
        if Row.get("levelId"):
            level_ids.add(Row.get("levelId"))

    if student.current_level_id and (not level_id or student.current_level_id == level_id):
        current_level = db.get(Level, student.current_level_id)
        if current_level and (not module_id or current_level.module_id == module_id):
            level_ids.add(current_level.id)

    level_progress = []
    for Promotion in promotion_records:
        if Promotion.from_level_id:
            level_ids.add(Promotion.from_level_id)
        if Promotion.to_level_id and (not level_id or Promotion.to_level_id == level_id):
            level_ids.add(Promotion.to_level_id)

    for level_key in sorted(level_ids):
        level = db.get(Level, level_key)
        if not level:
            continue
        module = db.get(Module, level.module_id)
        lessons = db.query(Lesson).filter(Lesson.level_id == level.id).all()
        lesson_ids = [lesson.id for lesson in lessons]
        dps_items = db.query(DPS).filter(DPS.lesson_id.in_(lesson_ids)).all() if lesson_ids else []
        dps_ids = {dps.id for dps in dps_items}
        level_attempts = [attempt for attempt, scope in filtered_attempts if scope.get("level") and scope["level"].id == level.id and attempt.dps_id in dps_ids]
        latest_by_dps: dict[str, Attempt] = {}
        for attempt in level_attempts:
            current = latest_by_dps.get(attempt.dps_id)
            attempt_date = attempt.submitted_at or attempt.started_at
            current_date = (current.submitted_at or current.started_at) if current else None
            if not current or (attempt_date and (not current_date or attempt_date > current_date)):
                latest_by_dps[attempt.dps_id] = attempt
        completed = [attempt for attempt in latest_by_dps.values() if (attempt.status or "").upper() in {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"}]
        passed = ClearedConceptAttempts(completed, BENCHMARK_PERCENTAGE)
        needs_reattempt = NeedsReattemptAttempts(completed, BENCHMARK_PERCENTAGE)
        avg_accuracy = round(sum(float(attempt.accuracy_percentage or 0) for attempt in completed) / len(completed)) if completed else 0
        avg_score = round(sum(float(attempt.total_score or 0) for attempt in completed) / len(completed)) if completed else 0
        performance_zone = "Excellence Zone" if avg_accuracy >= 90 else "Growth Zone" if avg_accuracy >= 70 else "Needs Improvement" if completed else "Not Started"
        required_dps = len(dps_items)
        completed_dps = len({attempt.dps_id for attempt in completed})
        readiness = "Ready" if required_dps > 0 and completed_dps >= required_dps and not needs_reattempt else "Not Ready"
        last_activity = max([(attempt.submitted_at or attempt.started_at) for attempt in completed if (attempt.submitted_at or attempt.started_at)], default=None)
        PromotionValue = _admin_latest_promotion_for_level(promotion_records, level.id)
        PromotionPayload = _admin_promotion_payload_for_report(db, PromotionValue) if PromotionValue else None
        PromotionStatus = "Promoted" if PromotionValue else "Not Promoted"
        level_progress.append({
            "moduleId": module.id if module else None,
            "moduleCode": module.module_code if module else None,
            "moduleName": module.module_name if module else None,
            "levelId": level.id,
            "levelCode": level.level_code,
            "levelName": level.level_name,
            "requiredDps": required_dps,
            "attemptedDps": len(latest_by_dps),
            "completedDps": completed_dps,
            "passedDps": len(passed),
            "pendingDps": max(required_dps - completed_dps, 0),
            "needsReattempt": len(needs_reattempt),
            "averageScore": avg_score,
            "averageAccuracy": avg_accuracy,
            "performanceZone": performance_zone,
            "assessmentReadiness": readiness,
            "promotionStatus": PromotionStatus,
            "fromLevelCode": PromotionPayload.get("fromLevelCode") if PromotionPayload else None,
            "toLevelCode": PromotionPayload.get("toLevelCode") if PromotionPayload else None,
            "promotedAt": PromotionPayload.get("promotedAt") if PromotionPayload else None,
            "promotedByName": PromotionPayload.get("promotedByName") if PromotionPayload else None,
            "lastActivity": last_activity.isoformat() if last_activity else (PromotionPayload.get("promotedAt") if PromotionPayload else None),
        })

    dps_cleared_rows = [
        Row for Row in dps_rows
        if ((Row.get("attemptStatus") or Row.get("status") or "").upper() in {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED", "CLEARED"}
            or (Row.get("completedDate") or Row.get("submittedAt")))
        and not Row.get("requiresAttention")
        and float(Row.get("accuracyPercentage") or 0) >= BENCHMARK_PERCENTAGE
    ]
    assessment_cleared_rows = [
        Row for Row in assessment_rows
        if ((Row.get("attemptStatus") or Row.get("status") or "").upper() in {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED", "CLEARED"}
            or (Row.get("completedDate") or Row.get("submittedAt")))
        and not Row.get("requiresAttention")
        and float(Row.get("accuracyPercentage") or 0) >= BENCHMARK_PERCENTAGE
    ]
    cleared_level_ids = {Row.get("levelId") for Row in assessment_cleared_rows if Row.get("levelId")}
    levels_cleared = len(cleared_level_ids)

    candidate_module_ids = {Row.get("moduleId") for Row in level_progress if Row.get("moduleId")}
    candidate_module_ids.update(Row.get("moduleId") for Row in assessment_cleared_rows if Row.get("moduleId"))
    modules_cleared = 0
    for module_key in candidate_module_ids:
        module_levels = db.query(Level).filter(Level.module_id == module_key).all()
        module_level_ids = {Item.id for Item in module_levels}
        if module_level_ids and module_level_ids.issubset(cleared_level_ids):
            modules_cleared += 1

    completed_dps_percentages = [
        float(Row.get("accuracyPercentage") or 0)
        for Row in dps_rows
        if Row.get("completedDate") or Row.get("submittedAt")
    ]
    completed_assessment_rows_for_accuracy = [
        Row for Row in assessment_rows
        if Row.get("completedDate") or Row.get("submittedAt")
    ]

    def AssessmentGroupKey(Row: dict) -> str:
        ModuleKey = str(Row.get("moduleId") or Row.get("moduleCode") or "module")
        LevelKey = str(Row.get("levelId") or Row.get("levelCode") or Row.get("assessmentId") or "level")
        return f"{ModuleKey}|{LevelKey}"

    def AssessmentRowSortKey(Row: dict) -> tuple[int, str]:
        AttemptNumber = Row.get("attemptNumber") or 1
        try:
            AttemptNumber = int(AttemptNumber)
        except (TypeError, ValueError):
            AttemptNumber = 1
        ActivityValue = (
            Row.get("completedDate")
            or Row.get("submittedAt")
            or Row.get("startedAt")
            or Row.get("assignedAt")
            or Row.get("createdAt")
            or ""
        )
        return AttemptNumber, str(ActivityValue)

    def AssessmentRowCleared(Row: dict) -> bool:
        DisplayStatus = str(Row.get("displayStatus") or Row.get("status") or "").strip().lower()
        BenchmarkStatus = str(Row.get("benchmarkStatus") or "").strip().upper()
        try:
            AccuracyValue = float(Row.get("accuracyPercentage") or Row.get("percentage") or 0)
        except (TypeError, ValueError):
            AccuracyValue = 0
        return (
            AccuracyValue >= BENCHMARK_PERCENTAGE
            or DisplayStatus in {"cleared", "re-attempt cleared"}
            or BenchmarkStatus == "PASS"
        )

    current_assessment_rows_by_level: dict[str, dict] = {}
    for Row in completed_assessment_rows_for_accuracy:
        GroupKey = AssessmentGroupKey(Row)
        CurrentRow = current_assessment_rows_by_level.get(GroupKey)
        RowCleared = AssessmentRowCleared(Row)
        CurrentCleared = AssessmentRowCleared(CurrentRow) if CurrentRow else False
        ShouldReplace = False
        if not CurrentRow:
            ShouldReplace = True
        elif RowCleared and not CurrentCleared:
            ShouldReplace = True
        elif RowCleared == CurrentCleared and AssessmentRowSortKey(Row) >= AssessmentRowSortKey(CurrentRow):
            ShouldReplace = True
        if ShouldReplace:
            current_assessment_rows_by_level[GroupKey] = Row

    current_assessment_percentages = [
        float(Row.get("accuracyPercentage") or Row.get("percentage") or 0)
        for Row in current_assessment_rows_by_level.values()
    ]
    completed_row_percentages = completed_dps_percentages + current_assessment_percentages
    practice_average_accuracy = int((sum(completed_dps_percentages) / len(completed_dps_percentages)) + 0.5) if completed_dps_percentages else 0
    assessment_average_accuracy = int((sum(current_assessment_percentages) / len(current_assessment_percentages)) + 0.5) if current_assessment_percentages else 0
    overall_accuracy = int((sum(completed_row_percentages) / len(completed_row_percentages)) + 0.5) if completed_row_percentages else 0
    activity_dates = []
    for ActivityValue in [(attempt.submitted_at or attempt.started_at) for attempt in completed_attempts if (attempt.submitted_at or attempt.started_at)]:
        if getattr(ActivityValue, "tzinfo", None):
            ActivityValue = ActivityValue.astimezone(datetime_timezone.utc).replace(tzinfo=None)
        activity_dates.append(ActivityValue)
    for Row in completed_assessment_rows:
        ActivityValue = Row.get("completedDate") or Row.get("submittedAt") or Row.get("attemptDate")
        if not ActivityValue:
            continue
        try:
            ParsedActivity = datetime.fromisoformat(str(ActivityValue))
            if ParsedActivity.tzinfo:
                ParsedActivity = ParsedActivity.astimezone(datetime_timezone.utc).replace(tzinfo=None)
            activity_dates.append(ParsedActivity)
        except ValueError:
            pass
    last_activity = max(activity_dates, default=None)
    student_payload = {
        "studentId": student.id,
        "studentName": student_user.full_name if student_user else "-",
        "studentCode": student.student_code,
        "className": student.class_name,
        "section": student.section,
        "status": "Active" if student.is_active else "Inactive",
        "currentModuleId": student.current_module_id,
        "currentLevelId": student.current_level_id,
        **teacher_context,
    }
    return {
        "student": student_payload,
        "summary": {
            "levels": len(level_progress),
            "levelsCleared": levels_cleared,
            "levelsCompleted": levels_cleared,
            "modulesCleared": modules_cleared,
            "modulesCompleted": modules_cleared,
            "dpsAttempts": len(dps_rows),
            "dpsCleared": len(dps_cleared_rows),
            "dpsCompleted": len(dps_cleared_rows),
            "needsReattempt": CountNeedsReattemptConcepts(completed_attempts, BENCHMARK_PERCENTAGE),
            "assessmentAttempts": len(assessment_rows),
            "assessmentsCleared": len(assessment_cleared_rows),
            "promotedLevels": len(promotion_records),
            "promotionHistoryCount": len(promotion_records),
            "completedAttempts": len(completed_attempts) + len(completed_assessment_rows),
            "practiceAverageAccuracy": practice_average_accuracy,
            "dpsAverageAccuracy": practice_average_accuracy,
            "assessmentAverageAccuracy": assessment_average_accuracy,
            "overallAverageAccuracy": overall_accuracy,
            "averageAccuracy": overall_accuracy,
            "lastActivity": last_activity.isoformat() if last_activity else None,
        },
        "levelProgress": level_progress,
        "dpsAttempts": dps_rows,
        "assessmentHistory": assessment_rows,
        "promotionHistory": promotion_rows,
    }




def _admin_filter_attempts_for_learning_performance(db: Session, teacher_id: str | None = None, module_id: str | None = None, level_id: str | None = None, lesson_id: str | None = None, dps_id: str | None = None) -> list[tuple[Attempt, dict]]:
    Attempts = db.query(Attempt).all()
    FilteredAttempts: list[tuple[Attempt, dict]] = []
    for AttemptValue in Attempts:
        Scope = _admin_attempt_scope(db, AttemptValue)
        AssignmentValue = Scope.get("assignment")
        if ((AssignmentValue.assignment_type if AssignmentValue else AttemptValue.mode) or "PRACTICE").upper() == "ASSESSMENT":
            continue
        StudentValue = db.get(Student, AttemptValue.student_id)
        ModuleValue = Scope.get("module")
        LevelValue = Scope.get("level")
        LessonValue = Scope.get("lesson")
        DpsValue = Scope.get("dps")
        if not _admin_student_matches_teacher(db, StudentValue, teacher_id):
            continue
        if module_id and (not ModuleValue or ModuleValue.id != module_id):
            continue
        if level_id and (not LevelValue or LevelValue.id != level_id):
            continue
        if lesson_id and (not LessonValue or LessonValue.id != lesson_id):
            continue
        if dps_id and (not DpsValue or DpsValue.id != dps_id):
            continue
        FilteredAttempts.append((AttemptValue, Scope))
    return FilteredAttempts


def _admin_learning_scope_type(module_id: str | None, level_id: str | None, lesson_id: str | None, dps_id: str | None) -> str:
    if dps_id:
        return "DPS"
    if lesson_id:
        return "Lesson"
    if level_id:
        return "Level"
    if module_id:
        return "Module"
    return "Platform"


def _admin_scope_label(db: Session, ModelValue: Any, IdValue: str | None, AllLabel: str, Formatter) -> str:
    if not IdValue:
        return AllLabel
    Item = db.get(ModelValue, IdValue)
    return Formatter(Item) if Item else IdValue


def _admin_teacher_label(db: Session, teacher_id: str | None) -> str:
    if not teacher_id:
        return "All Teachers"
    TeacherValue = db.get(Teacher, teacher_id)
    TeacherUser = db.get(User, TeacherValue.user_id) if TeacherValue else None
    if not TeacherValue:
        return teacher_id
    return f"{TeacherUser.full_name if TeacherUser else '-'} ({TeacherValue.teacher_code})"


def _admin_module_label(db: Session, module_id: str | None) -> str:
    return _admin_scope_label(db, Module, module_id, "All Modules", lambda Item: f"{Item.module_code} - {Item.module_name}")


def _admin_level_label(db: Session, level_id: str | None) -> str:
    return _admin_scope_label(db, Level, level_id, "All Levels", lambda Item: f"{Item.level_code} - {Item.level_name}")


def _admin_lesson_label(db: Session, lesson_id: str | None) -> str:
    return _admin_scope_label(db, Lesson, lesson_id, "All Lessons", lambda Item: f"Lesson {Item.lesson_number}: {Item.lesson_title}")


def _admin_dps_label(db: Session, dps_id: str | None) -> str:
    return _admin_scope_label(db, DPS, dps_id, "All DPS", lambda Item: f"DPS {Item.dps_number}: {Item.dps_title}")


def _admin_safe_filename_part(Value: Any) -> str:
    TextValue = re.sub(r"[^A-Za-z0-9]+", "_", str(Value or "Report")).strip("_")
    return TextValue[:90] or "Report"


def _admin_safe_level_filename_part(Value: Any) -> str:
    TextValue = re.sub(r"[^A-Za-z0-9-]+", "_", str(Value or "Level")).strip("_")
    return TextValue[:90] or "Level"


def _admin_report_file_date() -> str:
    return datetime.now(INDIA_REPORT_TIMEZONE).strftime("%d%b%Y")


def _admin_learning_export_filename(db: Session, teacher_id: str | None, module_id: str | None, level_id: str | None, lesson_id: str | None, dps_id: str | None) -> str:
    ScopeCode = "Scope"
    if dps_id:
        DpsValue = db.get(DPS, dps_id)
        ScopeCode = f"DPS-{DpsValue.dps_number}" if DpsValue else "DPS"
    elif lesson_id:
        LessonValue = db.get(Lesson, lesson_id)
        ScopeCode = f"Lesson-{LessonValue.lesson_number}" if LessonValue else "Lesson"
    elif level_id:
        LevelValue = db.get(Level, level_id)
        ScopeCode = LevelValue.level_code if LevelValue else "Level"
    elif module_id:
        ModuleValue = db.get(Module, module_id)
        ScopeCode = ModuleValue.module_code if ModuleValue else "Module"
    Parts = ["MP", "Learning", "Performance", ScopeCode, _admin_report_file_date()]
    return _admin_safe_filename_part("_".join(Parts)) + ".xlsx"


def _admin_student_export_filename(StudentValue: dict, db: Session, module_id: str | None, level_id: str | None, lesson_id: str | None, dps_id: str | None) -> str:
    Parts = ["MP", "Student", "History", StudentValue.get("studentCode") or "Student", _admin_report_file_date()]
    return _admin_safe_filename_part("_".join(Parts)) + ".xlsx"


def _admin_learning_performance_payload(db: Session, teacher_id: str | None = None, module_id: str | None = None, level_id: str | None = None, lesson_id: str | None = None, dps_id: str | None = None) -> dict:
    FilteredAttempts = _admin_filter_attempts_for_learning_performance(db, teacher_id, module_id, level_id, lesson_id, dps_id)
    Rows = []
    CompletedRows = []
    StudentIds: set[str] = set()
    BenchmarkMet = 0
    NeedsImprovement = 0
    for AttemptValue, Scope in FilteredAttempts:
        StudentValue = db.get(Student, AttemptValue.student_id)
        StudentUser = db.get(User, StudentValue.user_id) if StudentValue else None
        TeacherContext = _admin_teacher_context_for_student(db, StudentValue)
        ModuleValue = Scope.get("module")
        LevelValue = Scope.get("level")
        LessonValue = Scope.get("lesson")
        DpsValue = Scope.get("dps")
        BenchmarkPayload = benchmark_payload_for_attempt(AttemptValue)
        RequiresAttention = bool(BenchmarkPayload.get("requiresAttention"))
        IsCompleted = (AttemptValue.status or "").upper() in {"SUBMITTED", "AUTO_SUBMITTED", "COMPLETED"}
        if StudentValue:
            StudentIds.add(StudentValue.id)
        if IsCompleted:
            CompletedRows.append(AttemptValue)
            if RequiresAttention:
                NeedsImprovement += 1
            else:
                BenchmarkMet += 1
        Rows.append({
            "attemptId": AttemptValue.id,
            "studentId": AttemptValue.student_id,
            "studentName": StudentUser.full_name if StudentUser else None,
            "studentCode": StudentValue.student_code if StudentValue else None,
            **TeacherContext,
            "moduleId": ModuleValue.id if ModuleValue else None,
            "moduleCode": ModuleValue.module_code if ModuleValue else None,
            "moduleName": ModuleValue.module_name if ModuleValue else None,
            "levelId": LevelValue.id if LevelValue else None,
            "levelCode": LevelValue.level_code if LevelValue else None,
            "levelName": LevelValue.level_name if LevelValue else None,
            "lessonId": LessonValue.id if LessonValue else None,
            "lessonNumber": LessonValue.lesson_number if LessonValue else None,
            "lessonTitle": LessonValue.lesson_title if LessonValue else None,
            "dpsId": DpsValue.id if DpsValue else None,
            "dpsNumber": DpsValue.dps_number if DpsValue else None,
            "dpsTitle": DpsValue.dps_title if DpsValue else None,
            "score": AttemptValue.total_score,
            "maxScore": AttemptValue.max_score,
            "accuracyPercentage": AttemptValue.accuracy_percentage,
            "correct": AttemptValue.correct_count,
            "wrong": AttemptValue.wrong_count,
            "unanswered": AttemptValue.unanswered_count,
            "timeTakenSeconds": AttemptValue.time_taken_seconds,
            "status": _admin_attempt_display_status(db, AttemptValue),
            "attemptStatus": AttemptValue.status,
            **_admin_attempt_metadata(db, AttemptValue),
            **BenchmarkPayload,
            **attempt_date_payload(AttemptValue),
        })
    AverageAccuracy = round(sum(float(AttemptValue.accuracy_percentage or 0) for AttemptValue in CompletedRows) / len(CompletedRows)) if CompletedRows else 0
    AverageScore = round(sum(float(AttemptValue.total_score or 0) for AttemptValue in CompletedRows) / len(CompletedRows)) if CompletedRows else 0
    PromotionRecords = _admin_promotion_records_for_scope(
        db,
        teacher_id=teacher_id,
        module_id=module_id if module_id and module_id != "__ALL__" else None,
        level_id=level_id if level_id and level_id != "__ALL__" else None,
    )
    PromotedStudentIds = {Promotion.student_id for Promotion in PromotionRecords if Promotion.student_id}
    return {
        "summary": {
            "scopeType": _admin_learning_scope_type(module_id, level_id, lesson_id, dps_id),
            "studentsCovered": len(StudentIds),
            "attemptsReviewed": len(Rows),
            "completedAttempts": len(CompletedRows),
            "dpsCompleted": len(CompletedRows),
            "benchmarkMet": BenchmarkMet,
            "needsImprovement": NeedsImprovement,
            "promotedStudents": len(PromotedStudentIds),
            "promotionRecords": len(PromotionRecords),
            "averageScore": AverageScore,
            "averageAccuracy": AverageAccuracy,
        },
        "scope": {
            "teacherScope": _admin_teacher_label(db, teacher_id),
            "moduleScope": _admin_module_label(db, module_id),
            "levelScope": _admin_level_label(db, level_id),
            "lessonScope": _admin_lesson_label(db, lesson_id),
            "dpsScope": _admin_dps_label(db, dps_id),
        },
        "results": Rows,
    }


def _admin_combined_lesson_label(LessonNumber: Any, LessonTitle: Any) -> str:
    if not LessonNumber and not LessonTitle:
        return "-"
    NumberText = LessonNumber if LessonNumber else "-"
    TitleText = str(LessonTitle or "-").strip() or "-"
    return f"Lesson {NumberText}: {TitleText}"


def _admin_combined_dps_label(DpsNumber: Any, DpsTitle: Any) -> str:
    if not DpsNumber and not DpsTitle:
        return "-"
    NumberText = DpsNumber if DpsNumber else "-"
    TitleText = str(DpsTitle or "-").strip() or "-"
    return f"DPS {NumberText}: {TitleText}"


def _admin_learning_export_rows(Payload: dict, TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None) -> tuple[list[dict], list[dict]]:
    DetailRows = []
    ImprovementRows = []
    for Row in Payload.get("results", []):
        CompletedValue = Row.get("completedDate") or Row.get("submittedAt")
        ExportRow = {
            "Student Name": Row.get("studentName") or "-",
            "Student Code": Row.get("studentCode") or "-",
            "Teacher Name": Row.get("teacherName") or "Not Assigned",
            "Teacher Code": Row.get("teacherCode") or "-",
            "Module Code": Row.get("moduleCode") or "-",
            "Module": Row.get("moduleName") or "-",
            "Level Code": Row.get("levelCode") or "-",
            "Level": Row.get("levelName") or "-",
            "Lesson": _admin_combined_lesson_label(Row.get("lessonNumber"), Row.get("lessonTitle")),
            "DPS": _admin_combined_dps_label(Row.get("dpsNumber"), Row.get("dpsTitle")),
            "Status": Row.get("displayStatus") or Row.get("status") or "-",
            "Score": Row.get("score") or 0,
            "Total Marks": Row.get("maxScore") or 0,
            "Accuracy %": Row.get("accuracyPercentage") or 0,
            "Benchmark Status": "Needs Improvement" if Row.get("requiresAttention") else "Benchmark Met",
            "Correct Answers": Row.get("correct") or 0,
            "Completed Date": _admin_report_date(CompletedValue, TimezoneName, TimezoneOffsetMinutes),
            "Completion Time": _admin_report_time(CompletedValue, TimezoneName, TimezoneOffsetMinutes),
            "Time Taken": _admin_report_duration(Row.get("timeTakenSeconds")),
        }
        DetailRows.append(ExportRow)
        if Row.get("requiresAttention"):
            ImprovementRows.append(ExportRow)
    return DetailRows, ImprovementRows


@router.get("/results/learning-performance")
def learning_performance_results(teacherId: str | None = None, moduleId: str | None = None, levelId: str | None = None, lessonId: str | None = None, dpsId: str | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return _admin_learning_performance_payload(db, teacherId, moduleId, levelId, lessonId, dpsId)


@router.get("/results/export/learning-performance")
def export_learning_performance_results(teacherId: str | None = None, moduleId: str | None = None, levelId: str | None = None, lessonId: str | None = None, dpsId: str | None = None, timezone: str | None = None, timezoneOffsetMinutes: int | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    Payload = _admin_learning_performance_payload(db, teacherId, moduleId, levelId, lessonId, dpsId)
    Summary = Payload.get("summary", {})
    Scope = Payload.get("scope", {})
    DetailRows, ImprovementRows = _admin_learning_export_rows(Payload, timezone, timezoneOffsetMinutes)
    return BuildWorkbookResponse(
        _admin_learning_export_filename(db, teacherId, moduleId, levelId, lessonId, dpsId),
        [
            ("Report Type", "Learning Performance"),
            ("Generated Date", _admin_report_date(datetime.now(datetime_timezone.utc), timezone, timezoneOffsetMinutes)),
            ("Generated Time", _admin_report_time(datetime.now(datetime_timezone.utc), timezone, timezoneOffsetMinutes)),
            ("Teacher Scope", Scope.get("teacherScope", "All Teachers")),
            ("Module Scope", Scope.get("moduleScope", "All Modules")),
            ("Level Scope", Scope.get("levelScope", "All Levels")),
            ("Lesson Scope", Scope.get("lessonScope", "All Lessons")),
            ("DPS Scope", Scope.get("dpsScope", "All DPS")),
            ("Students Covered", Summary.get("studentsCovered", 0)),
            ("Attempts Reviewed", Summary.get("attemptsReviewed", 0)),
            ("Completed Attempts", Summary.get("completedAttempts", 0)),
            ("Benchmark Met", Summary.get("benchmarkMet", 0)),
            ("Needs Improvement", Summary.get("needsImprovement", 0)),
            ("Promoted Students", Summary.get("promotedStudents", 0)),
            ("Promotion Records", Summary.get("promotionRecords", 0)),
            ("Average Score", Summary.get("averageScore", 0)),
            ("Average Accuracy %", Summary.get("averageAccuracy", 0)),
            ("Benchmark Rule", "70% minimum accuracy"),
        ],
        [
            ("Student Performance", DetailRows),
            ("Improvement Focus", ImprovementRows),
        ],
    )


@router.get("/results/student")
def student_results(studentId: str, moduleId: str | None = None, levelId: str | None = None, lessonId: str | None = None, dpsId: str | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return _admin_build_student_report(db, studentId, moduleId, levelId, lessonId, dpsId)


@router.get("/results/export/dps")
def export_dps_results(dpsId: str, teacherId: str | None = None, timezone: str | None = None, timezoneOffsetMinutes: int | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    Rows, Summary = _admin_dps_export_rows(db, dpsId, teacherId, timezone, timezoneOffsetMinutes)
    TeacherLabel = "All Teachers"
    if teacherId:
        teacher = db.get(Teacher, teacherId)
        teacher_user = db.get(User, teacher.user_id) if teacher else None
        TeacherLabel = f"{teacher_user.full_name if teacher_user else '-'} ({teacher.teacher_code if teacher else '-'})"
    return BuildWorkbookResponse(
        _admin_learning_export_filename(db, teacherId, Summary.get("moduleId"), Summary.get("levelId"), Summary.get("lessonId"), dpsId),
        [
            ("Report Type", "DPS Report"),
            ("Generated Date", _admin_report_date(datetime.now(datetime_timezone.utc), timezone, timezoneOffsetMinutes)),
            ("Generated Time", _admin_report_time(datetime.now(datetime_timezone.utc), timezone, timezoneOffsetMinutes)),
            ("Teacher Filter", TeacherLabel),
            ("Module", Summary.get("moduleLabel")),
            ("Level", Summary.get("levelLabel")),
            ("Lesson", Summary.get("lessonLabel")),
            ("DPS", Summary.get("dpsLabel")),
            ("Rows Reviewed", len(Rows)),
            ("Benchmark Rule", "70% minimum accuracy"),
        ],
        [("DPS Results", Rows)],
    )


@router.get("/results/export/level")
def export_level_results(levelId: str, moduleId: str | None = None, teacherId: str | None = None, timezone: str | None = None, timezoneOffsetMinutes: int | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    Rows, Summary = _admin_level_export_rows(db, moduleId, levelId, teacherId, timezone, timezoneOffsetMinutes)
    TeacherLabel = "All Teachers"
    if teacherId:
        teacher = db.get(Teacher, teacherId)
        teacher_user = db.get(User, teacher.user_id) if teacher else None
        TeacherLabel = f"{teacher_user.full_name if teacher_user else '-'} ({teacher.teacher_code if teacher else '-'})"
    return BuildWorkbookResponse(
        _admin_learning_export_filename(db, teacherId, moduleId, levelId, None, None),
        [
            ("Report Type", "Level Report"),
            ("Generated Date", _admin_report_date(datetime.now(datetime_timezone.utc), timezone, timezoneOffsetMinutes)),
            ("Generated Time", _admin_report_time(datetime.now(datetime_timezone.utc), timezone, timezoneOffsetMinutes)),
            ("Teacher Filter", TeacherLabel),
            ("Module", Summary.get("moduleLabel")),
            ("Level", Summary.get("levelLabel")),
            ("Rows Reviewed", len(Rows)),
            ("Ready", Summary.get("ready", 0)),
            ("Not Ready", Summary.get("notReady", 0)),
            ("Benchmark Rule", "70% minimum accuracy"),
        ],
        [("Level Results", Rows)],
    )


@router.get("/results/export/student")
def export_student_results(studentId: str, moduleId: str | None = None, levelId: str | None = None, lessonId: str | None = None, dpsId: str | None = None, timezone: str | None = None, timezoneOffsetMinutes: int | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    Payload = _admin_build_student_report(db, studentId, moduleId, levelId, lessonId, dpsId)
    Student = Payload.get("student", {})
    Summary = Payload.get("summary", {})
    ProfileRows = [{
        "Student Name": Student.get("studentName"),
        "Student Code": Student.get("studentCode"),
        "Class": Student.get("className") or "-",
        "Section": Student.get("section") or "-",
        "Account Status": Student.get("status"),
        "Current / Assigned Teacher": Student.get("teacherName") or "Not Assigned",
        "Teacher Code": Student.get("teacherCode") or "-",
        "Levels In Report": Summary.get("levels", 0),
        "Modules Cleared": Summary.get("modulesCleared", Summary.get("modulesCompleted", 0)),
        "Levels Cleared": Summary.get("levelsCleared", Summary.get("levelsCompleted", 0)),
        "DPS Cleared": Summary.get("dpsCleared", Summary.get("dpsCompleted", 0)),
        "DPS Attempts": Summary.get("dpsAttempts", 0),
        "Assessments Cleared": Summary.get("assessmentsCleared", 0),
        "Promoted Levels": Summary.get("promotedLevels", 0),
        "Promotion History Records": Summary.get("promotionHistoryCount", 0),
        "Average Accuracy %": Summary.get("averageAccuracy", 0),
        "Last Activity Date": _admin_report_date(Summary.get("lastActivity"), timezone, timezoneOffsetMinutes),
    }]
    LevelRows = []
    for Row in Payload.get("levelProgress", []):
        LevelRows.append({
            "Module Code": Row.get("moduleCode") or "-",
            "Module": Row.get("moduleName") or "-",
            "Level Code": Row.get("levelCode") or "-",
            "Level": Row.get("levelName") or "-",
            "Teacher Name": Row.get("teacherName") or Student.get("teacherName") or "Not Assigned",
            "Teacher Code": Row.get("teacherCode") or Student.get("teacherCode") or "-",
            "Required DPS": Row.get("requiredDps") or 0,
            "Completed DPS": Row.get("completedDps") or 0,
            "Passed DPS": Row.get("passedDps") or 0,
            "Pending DPS": Row.get("pendingDps") or 0,
            "Needs Re-Attempt": Row.get("needsReattempt") or 0,
            "Average Score": Row.get("averageScore") or 0,
            "Average Accuracy %": Row.get("averageAccuracy") or 0,
            "Performance Zone": Row.get("performanceZone") or "-",
            "Assessment Readiness": Row.get("assessmentReadiness") or "-",
            "Promotion Status": Row.get("promotionStatus") or "Not Promoted",
            "From Level": Row.get("fromLevelCode") or "-",
            "To Level": Row.get("toLevelCode") or "-",
            "Promoted Date": _admin_report_date(Row.get("promotedAt"), timezone, timezoneOffsetMinutes),
            "Promoted Time": _admin_report_time(Row.get("promotedAt"), timezone, timezoneOffsetMinutes),
            "Promoted By": Row.get("promotedByName") or "-",
            "Last Activity Date": _admin_report_date(Row.get("lastActivity"), timezone, timezoneOffsetMinutes),
        })
    LevelRows = sorted(
        LevelRows,
        key=lambda Row: (
            _admin_natural_sort_key(Row.get("Module Code")),
            _admin_natural_sort_key(Row.get("Level Code")),
        ),
    )

    PromotionRows = []
    for Row in Payload.get("promotionHistory", []):
        PromotionRows.append({
            "Student Name": Row.get("studentName") or Student.get("studentName") or "-",
            "Student Code": Row.get("studentCode") or Student.get("studentCode") or "-",
            "Promotion Status": Row.get("promotionStatus") or "Promoted",
            "From Module": Row.get("fromModuleCode") or "-",
            "From Level": Row.get("fromLevelCode") or "-",
            "To Module": Row.get("toModuleCode") or "-",
            "To Level": Row.get("toLevelCode") or "-",
            "Promotion Assessment": Row.get("assessmentTitle") or "Assessment",
            "Promotion Score": NormalizeAssessmentScore(Row.get("score") or 0, Row.get("maxScore") or Row.get("totalMarks") or 100),
            "Promotion Percentage": NormalizeAssessmentPercentage(Row.get("percentage") or 0, 100),
            "Promoted Date": _admin_report_date(Row.get("promotedAt"), timezone, timezoneOffsetMinutes),
            "Promoted Time": _admin_report_time(Row.get("promotedAt"), timezone, timezoneOffsetMinutes),
            "Promoted By": Row.get("promotedByName") or "Admin",
        })
    DpsRows = []
    for Row in Payload.get("dpsAttempts", []):
        CompletedValue = Row.get("completedDate") or Row.get("submittedAt")
        DpsRows.append({
            "Module Code": Row.get("moduleCode") or "-",
            "Module": Row.get("moduleName") or "-",
            "Level Code": Row.get("levelCode") or "-",
            "Level": Row.get("levelName") or "-",
            "Lesson": _admin_combined_lesson_label(Row.get("lessonNumber"), Row.get("lessonTitle")),
            "DPS": _admin_combined_dps_label(Row.get("dpsNumber"), Row.get("dpsTitle")),
            "Teacher Name": Row.get("teacherName") or Student.get("teacherName") or "Not Assigned",
            "Teacher Code": Row.get("teacherCode") or Student.get("teacherCode") or "-",
            "Status": Row.get("displayStatus") or Row.get("status") or "-",
            "Score": Row.get("score") or 0,
            "Total Marks": Row.get("maxScore") or 0,
            "Accuracy %": Row.get("accuracyPercentage") or 0,
            "Benchmark Status": "Needs Improvement" if Row.get("requiresAttention") else "Benchmark Met",
            "Correct Answers": Row.get("correct") or 0,
            "Completed Date": _admin_report_date(CompletedValue, timezone, timezoneOffsetMinutes),
            "Completion Time": _admin_report_time(CompletedValue, timezone, timezoneOffsetMinutes),
            "Time Taken": _admin_report_duration(Row.get("timeTakenSeconds")),
        })
    AssessmentRows = []
    for Row in Payload.get("assessmentHistory", []):
        CompletedValue = Row.get("completedDate") or Row.get("submittedAt")
        AssessmentRows.append({
            "Module Code": Row.get("moduleCode") or "-",
            "Module": Row.get("moduleName") or "-",
            "Level Code": Row.get("levelCode") or "-",
            "Level": Row.get("levelName") or "-",
            "Assessment": Row.get("assessmentTitle") or Row.get("assignmentTitle") or f"Level Assessment - {Row.get('levelCode') or '-'}",
            "Attempt": Row.get("attemptLabel") or Row.get("attemptType") or "Original",
            "Teacher Name": Row.get("teacherName") or Student.get("teacherName") or "Not Assigned",
            "Teacher Code": Row.get("teacherCode") or Student.get("teacherCode") or "-",
            "Status": Row.get("displayStatus") or Row.get("status") or "-",
            "Score": NormalizeAssessmentScore(Row.get("score") or 0, Row.get("maxScore") or Row.get("totalMarks") or 100),
            "Total Marks": NormalizeAssessmentScore(Row.get("maxScore") or Row.get("totalMarks") or 0, Row.get("maxScore") or Row.get("totalMarks") or 0),
            "Accuracy %": NormalizeAssessmentPercentage(Row.get("accuracyPercentage") or Row.get("percentage") or 0, 100),
            "Result": "Needs Re-Attempt" if Row.get("requiresAttention") else "Cleared",
            "Correct Answers": Row.get("correct") or 0,
            "Wrong Answers": Row.get("wrong") or 0,
            "Unanswered": Row.get("unanswered") or 0,
            "Assigned Date": _admin_report_date(Row.get("createdAt"), timezone, timezoneOffsetMinutes),
            "Completed Date": _admin_report_date(CompletedValue, timezone, timezoneOffsetMinutes),
            "Completion Time": _admin_report_time(CompletedValue, timezone, timezoneOffsetMinutes),
            "Time Taken": _admin_report_duration(Row.get("timeTakenSeconds")),
        })
    SelectedModuleLabel = "All Modules"
    if moduleId:
        ModuleValue = db.get(Module, moduleId)
        SelectedModuleLabel = f"{ModuleValue.module_code} - {ModuleValue.module_name}" if ModuleValue else moduleId
    SelectedLevelLabel = "All Levels"
    if levelId:
        LevelValue = db.get(Level, levelId)
        SelectedLevelLabel = f"{LevelValue.level_code} - {LevelValue.level_name}" if LevelValue else levelId
    SelectedLessonLabel = _admin_lesson_label(db, lessonId)
    SelectedDpsLabel = _admin_dps_label(db, dpsId)

    return BuildWorkbookResponse(
        _admin_student_export_filename(Student, db, moduleId, levelId, lessonId, dpsId),
        [
            ("Report Type", "Student History"),
            ("Generated Date", _admin_report_date(datetime.now(datetime_timezone.utc), timezone, timezoneOffsetMinutes)),
            ("Generated Time", _admin_report_time(datetime.now(datetime_timezone.utc), timezone, timezoneOffsetMinutes)),
            ("Student", f"{Student.get('studentName')} ({Student.get('studentCode')})"),
            ("Class", Student.get("className") or "-"),
            ("Section", Student.get("section") or "-"),
            ("Account Status", Student.get("status") or "-"),
            ("Current / Assigned Teacher", Student.get("teacherName") or "Not Assigned"),
            ("Teacher Code", Student.get("teacherCode") or "-"),
            ("Module Scope", SelectedModuleLabel),
            ("Level Scope", SelectedLevelLabel),
            ("Lesson Scope", SelectedLessonLabel),
            ("DPS Scope", SelectedDpsLabel),
            ("Modules Cleared", Summary.get("modulesCleared", Summary.get("modulesCompleted", 0))),
            ("Levels Cleared", Summary.get("levelsCleared", Summary.get("levelsCompleted", 0))),
            ("Levels In Scope", Summary.get("levels", 0)),
            ("DPS Attempts Reviewed", Summary.get("dpsAttempts", 0)),
            ("DPS Cleared", Summary.get("dpsCleared", Summary.get("dpsCompleted", 0))),
            ("Assessments Cleared", Summary.get("assessmentsCleared", 0)),
            ("Promoted Levels", Summary.get("promotedLevels", 0)),
            ("Promotion History Records", Summary.get("promotionHistoryCount", 0)),
            ("DPS Avg Accuracy %", Summary.get("practiceAverageAccuracy", Summary.get("dpsAverageAccuracy", 0))),
            ("Assessment Avg Accuracy %", Summary.get("assessmentAverageAccuracy", 0)),
            ("Benchmark Rule", "70% minimum accuracy"),
        ],
        [
            ("Level Progress", LevelRows),
            ("Promotion History", PromotionRows),
            ("DPS Learning History", DpsRows),
            ("Assessment History", AssessmentRows),
        ],
    )


def _admin_parent_report_latest_by_date(Rows: list[dict]) -> dict | None:
    def SortValue(Row: dict) -> str:
        return str(
            Row.get("completedDate")
            or Row.get("submittedAt")
            or Row.get("startedAt")
            or Row.get("assignedAt")
            or Row.get("createdAt")
            or ""
        )
    return sorted(Rows, key=SortValue, reverse=True)[0] if Rows else None


def _admin_parent_assessment_status(Row: dict | None) -> str:
    if not Row:
        return "Assessment Not Started"
    DisplayStatus = str(Row.get("displayStatus") or Row.get("status") or "").strip()
    UpperStatus = DisplayStatus.upper()
    if "NEEDS" in UpperStatus:
        return "Needs More Practice"
    if "PENDING" in UpperStatus:
        return "Assessment Pending"
    try:
        AccuracyValue = float(Row.get("accuracyPercentage") or Row.get("percentage") or 0)
    except (TypeError, ValueError):
        AccuracyValue = 0
    BenchmarkStatus = str(Row.get("benchmarkStatus") or "").upper()
    if AccuracyValue >= BENCHMARK_PERCENTAGE or BenchmarkStatus == "PASS" or "CLEARED" in UpperStatus:
        return "Assessment Cleared"
    return DisplayStatus or "Assessment In Progress"



def _admin_find_module_by_code_or_name(db: Session, ModuleCodeOrName: str | None) -> Module | None:
    LookupValue = str(ModuleCodeOrName or "").strip()
    if not LookupValue or LookupValue in {"-", "Learning Journey"}:
        return None
    return (
        db.query(Module)
        .filter((Module.module_code == LookupValue) | (Module.module_name == LookupValue))
        .first()
    )


def _admin_ordered_active_modules(db: Session) -> list[Module]:
    Modules = db.query(Module).filter(Module.is_active == True).all()
    return sorted(
        Modules,
        key=lambda ModuleValue: (
            int(ModuleValue.display_order or 0),
            _admin_natural_sort_key(ModuleValue.module_code),
            _admin_natural_sort_key(ModuleValue.module_name),
        ),
    )


def _admin_ordered_active_levels(db: Session, ModuleId: str) -> list[Level]:
    Levels = db.query(Level).filter(Level.module_id == ModuleId, Level.is_active == True).all()
    return sorted(
        Levels,
        key=lambda LevelValue: (
            int(LevelValue.display_order or 0),
            int(LevelValue.internal_level_number or 0),
            _admin_natural_sort_key(LevelValue.level_code),
            _admin_natural_sort_key(LevelValue.level_name),
        ),
    )


def _admin_parent_report_next_learning_destination(db: Session | None, ModuleCodeOrName: str | None, CompletedLevelCode: str | None) -> dict[str, Any]:
    """Return the next valid learning destination after a completed report level.

    This intentionally does not depend on whether Admin has already clicked Promote.
    Completed-level parent reports must show the next curriculum destination, not the
    student's still-current active level while promotion is pending.
    """
    EmptyResult = {
        "moduleCode": None,
        "moduleName": None,
        "levelCode": None,
        "levelName": None,
        "destinationType": "NEXT_LEVEL_SETUP_REQUIRED",
    }
    if db is None:
        return EmptyResult

    ModuleValue = _admin_find_module_by_code_or_name(db, ModuleCodeOrName)
    CompletedCode = str(CompletedLevelCode or "").strip()
    if not ModuleValue or not CompletedCode or CompletedCode in {"-", "Completed Level", "Not Started"}:
        return EmptyResult

    CurrentModuleLevels = _admin_ordered_active_levels(db, ModuleValue.id)
    CurrentIndex = next((Index for Index, LevelValue in enumerate(CurrentModuleLevels) if LevelValue.level_code == CompletedCode), -1)
    if CurrentIndex >= 0 and CurrentIndex + 1 < len(CurrentModuleLevels):
        NextLevelValue = CurrentModuleLevels[CurrentIndex + 1]
        return {
            "moduleCode": ModuleValue.module_code,
            "moduleName": ModuleValue.module_name,
            "levelCode": NextLevelValue.level_code,
            "levelName": NextLevelValue.level_name,
            "destinationType": "SAME_MODULE_NEXT_LEVEL",
        }

    OrderedModules = _admin_ordered_active_modules(db)
    ModuleIndex = next((Index for Index, ModuleItem in enumerate(OrderedModules) if ModuleItem.id == ModuleValue.id), -1)
    if ModuleIndex >= 0:
        for NextModuleValue in OrderedModules[ModuleIndex + 1:]:
            NextLevels = _admin_ordered_active_levels(db, NextModuleValue.id)
            if NextLevels:
                NextLevelValue = NextLevels[0]
                return {
                    "moduleCode": NextModuleValue.module_code,
                    "moduleName": NextModuleValue.module_name,
                    "levelCode": NextLevelValue.level_code,
                    "levelName": NextLevelValue.level_name,
                    "destinationType": "NEXT_MODULE_FIRST_LEVEL",
                }

    return EmptyResult

def _admin_build_parent_progress_pdf_data(Payload: dict, TimezoneName: str | None = None, TimezoneOffsetMinutes: int | None = None, db: Session | None = None) -> dict:
    StudentValue = Payload.get("student", {}) or {}
    SummaryValue = Payload.get("summary", {}) or {}
    LevelRows = Payload.get("levelProgress", []) or []
    PromotionRows = Payload.get("promotionHistory", []) or []
    AssessmentRows = Payload.get("assessmentHistory", []) or []

    CurrentLevelId = StudentValue.get("currentLevelId")
    CurrentLevelRow = next((Row for Row in LevelRows if Row.get("levelId") == CurrentLevelId), None) or (LevelRows[0] if LevelRows else {})
    LatestPromotion = _admin_parent_report_latest_by_date(PromotionRows)
    LatestAssessment = _admin_parent_report_latest_by_date(AssessmentRows)

    def _row_for_level(LevelCode: str | None) -> dict:
        if not LevelCode:
            return {}
        return next((Row for Row in LevelRows if str(Row.get("levelCode") or "") == str(LevelCode)), None) or {}

    def _assessment_for_level(LevelCode: str | None) -> dict | None:
        MatchingRows = [Row for Row in AssessmentRows if str(Row.get("levelCode") or "") == str(LevelCode or "")]
        ClearedRows = [Row for Row in MatchingRows if _admin_parent_assessment_status(Row) == "Assessment Cleared"]
        return _admin_parent_report_latest_by_date(ClearedRows or MatchingRows)

    ReportSource = "current"
    ReportLevelRow: dict = {}
    ReportAssessment: dict | None = None
    ReportModuleCode = "Learning Journey"
    ReportModuleName = "Learning Journey"
    ReportLevelCode = "Not Started"
    ReportLevelName = "Level Progress"
    NextLevelCode = CurrentLevelRow.get("levelCode") or "Next Level"
    NextLevelName = CurrentLevelRow.get("levelName") or "Next Level"
    AssessmentTitle = "Level Assessment"
    AssessmentScore = 0
    AssessmentTotal = 100
    AssessmentPercentage = 0
    AssessmentDate = "-"
    ResultLabel = "Assessment Milestone Cleared"

    if LatestPromotion:
        ReportSource = "promotion"
        ReportModuleCode = LatestPromotion.get("fromModuleCode") or LatestPromotion.get("fromModuleName") or "Learning Journey"
        ReportModuleName = LatestPromotion.get("fromModuleName") or LatestPromotion.get("fromModuleCode") or ReportModuleCode
        ReportLevelCode = LatestPromotion.get("fromLevelCode") or "Completed Level"
        ReportLevelName = LatestPromotion.get("fromLevelName") or LatestPromotion.get("fromLevelCode") or ReportLevelCode
        NextLevelCode = LatestPromotion.get("toLevelCode") or CurrentLevelRow.get("levelCode") or "Next Level"
        NextLevelName = LatestPromotion.get("toLevelName") or LatestPromotion.get("toLevelCode") or NextLevelCode
        AssessmentTitle = LatestPromotion.get("assessmentTitle") or "Level Assessment"
        AssessmentScore = NormalizeAssessmentScore(LatestPromotion.get("score") or 0, LatestPromotion.get("maxScore") or LatestPromotion.get("totalMarks") or 100)
        AssessmentTotal = NormalizeAssessmentScore(LatestPromotion.get("maxScore") or LatestPromotion.get("totalMarks") or 100, LatestPromotion.get("maxScore") or LatestPromotion.get("totalMarks") or 100)
        AssessmentPercentage = NormalizeAssessmentPercentage(LatestPromotion.get("percentage") or 0, 100)
        AssessmentDate = _admin_report_datetime(LatestPromotion.get("promotedAt"), TimezoneName, TimezoneOffsetMinutes)
        ReportLevelRow = _row_for_level(ReportLevelCode)
        ReportAssessment = _assessment_for_level(ReportLevelCode)
    elif LatestAssessment:
        ReportSource = "assessment"
        ReportModuleCode = LatestAssessment.get("moduleCode") or LatestAssessment.get("moduleName") or "Learning Journey"
        ReportModuleName = LatestAssessment.get("moduleName") or LatestAssessment.get("moduleCode") or ReportModuleCode
        ReportLevelCode = LatestAssessment.get("levelCode") or "Completed Level"
        ReportLevelName = LatestAssessment.get("levelName") or LatestAssessment.get("levelCode") or ReportLevelCode
        NextLevelCode = CurrentLevelRow.get("levelCode") or "Next Level"
        NextLevelName = CurrentLevelRow.get("levelName") or NextLevelCode
        AssessmentTitle = LatestAssessment.get("assessmentTitle") or LatestAssessment.get("assignmentTitle") or "Level Assessment"
        AssessmentScore = NormalizeAssessmentScore(LatestAssessment.get("score") or 0, LatestAssessment.get("maxScore") or LatestAssessment.get("totalMarks") or 100)
        AssessmentTotal = NormalizeAssessmentScore(LatestAssessment.get("maxScore") or LatestAssessment.get("totalMarks") or 100, LatestAssessment.get("maxScore") or LatestAssessment.get("totalMarks") or 100)
        AssessmentPercentage = NormalizeAssessmentPercentage(LatestAssessment.get("accuracyPercentage") or LatestAssessment.get("percentage") or 0, 100)
        AssessmentDate = _admin_report_datetime(LatestAssessment.get("completedDate") or LatestAssessment.get("submittedAt"), TimezoneName, TimezoneOffsetMinutes)
        ResultLabel = _admin_parent_assessment_status(LatestAssessment)
        ReportLevelRow = _row_for_level(ReportLevelCode)
        ReportAssessment = LatestAssessment
    else:
        ReportLevelRow = CurrentLevelRow
        ReportModuleCode = CurrentLevelRow.get("moduleCode") or CurrentLevelRow.get("moduleName") or "Learning Journey"
        ReportModuleName = CurrentLevelRow.get("moduleName") or CurrentLevelRow.get("moduleCode") or ReportModuleCode
        ReportLevelCode = CurrentLevelRow.get("levelCode") or "Not Started"
        ReportLevelName = CurrentLevelRow.get("levelName") or CurrentLevelRow.get("levelCode") or ReportLevelCode
        NextLevelCode = "Next Level"
        NextLevelName = "Next Level"
        ResultLabel = "Progress In Review"

    NextDestination = _admin_parent_report_next_learning_destination(db, ReportModuleCode, ReportLevelCode)
    if NextDestination.get("levelCode"):
        if not LatestPromotion or not NextLevelCode or NextLevelCode in {"-", "Next Level", ReportLevelCode}:
            NextLevelCode = NextDestination.get("levelCode") or NextLevelCode
            NextLevelName = NextDestination.get("levelName") or NextLevelCode
    elif not LatestPromotion and ReportSource in {"assessment", "promotion"}:
        NextLevelCode = "Next Level Pending Setup"
        NextLevelName = "Next Level Pending Setup"

    RequiredDps = int(ReportLevelRow.get("requiredDps") or 0)
    CompletedDps = int(ReportLevelRow.get("completedDps") or ReportLevelRow.get("passedDps") or 0)
    PracticeAccuracy = int(round(float(ReportLevelRow.get("averageAccuracy") or 0)))
    if not RequiredDps and SummaryValue.get("dpsCleared"):
        CompletedDps = int(SummaryValue.get("dpsCleared") or 0)
    PracticeProgress = f"{CompletedDps} / {RequiredDps} Practice Sheets" if RequiredDps else f"{CompletedDps} Practice Sheets"
    ScoreText = f"{AssessmentScore} / {AssessmentTotal}"
    PercentageText = f"{AssessmentPercentage}%"
    ModuleDisplayNames = {
        "YLM": "Young Learners Module",
    }
    if ReportModuleName == ReportModuleCode and ReportModuleCode in ModuleDisplayNames:
        ReportModuleName = ModuleDisplayNames[ReportModuleCode]
    if not ReportModuleName or ReportModuleName == "-":
        ReportModuleName = ModuleDisplayNames.get(ReportModuleCode, ReportModuleCode or "Learning Journey")

    StudentName = StudentValue.get("studentName") or "This learner"
    FirstName = str(StudentName).split()[0] if StudentName else "The learner"

    if ReportSource in {"promotion", "assessment"} and ResultLabel in {"Assessment Milestone Cleared", "Assessment Cleared"}:
        if NextLevelCode == "Next Level Pending Setup":
            Message = f"{FirstName} has successfully completed {ReportLevelCode} by clearing {AssessmentTitle} with {ScoreText}. The next structured MathPath level is pending setup."
        else:
            Message = f"{FirstName} has successfully completed {ReportLevelCode} by clearing {AssessmentTitle} with {ScoreText}. {FirstName} is now ready to begin {NextLevelCode} with confidence."
    elif "Cleared" in ResultLabel:
        Message = f"{FirstName} has completed the assessment milestone for {ReportLevelCode}. The next step is to continue the planned MathPath journey with teacher guidance."
    elif "Practice" in ResultLabel or "Needs" in ResultLabel:
        Message = f"{FirstName} is strengthening the concepts from {ReportLevelCode}. Focused revision and calm, regular practice will help build confidence for the next milestone."
    else:
        Message = f"{FirstName}'s progress for {ReportLevelCode} is ready for review. This report summarizes the latest learning, practice, and assessment records."

    ParentGuidance = f"Celebrate the completion of {ReportLevelCode}. The next learning focus is {NextLevelCode}, where regular practice will help convert this milestone into stronger speed, accuracy, and confidence."
    if NextLevelCode == "Next Level Pending Setup":
        ParentGuidance = f"Celebrate the completion of {ReportLevelCode}. The next structured MathPath level should be set up before the next learning journey begins."
    elif ResultLabel not in {"Assessment Milestone Cleared", "Assessment Cleared"}:
        ParentGuidance = f"The immediate focus should be steady guided practice for {ReportLevelCode}. Accuracy should remain the priority before speed."

    MovementRows = []
    if LatestPromotion:
        MovementRows.append({
            "fromLevel": LatestPromotion.get("fromLevelCode") or ReportLevelCode,
            "toLevel": LatestPromotion.get("toLevelCode") or NextLevelCode,
            "assessment": LatestPromotion.get("assessmentTitle") or AssessmentTitle,
            "score": ScoreText,
            "percentage": PercentageText,
            "date": AssessmentDate,
        })

    return {
        "student": {
            "name": StudentName,
            "code": StudentValue.get("studentCode") or "-",
            "classSection": f"{StudentValue.get('className') or '-'} / {StudentValue.get('section') or '-'}",
        },
        "report": {
            "title": "Student Progress Report",
            "reportLevelCode": ReportLevelCode,
            "reportLevelName": ReportLevelName,
            "reportModuleCode": ReportModuleCode,
            "reportModuleName": ReportModuleName,
            "nextLevelCode": NextLevelCode,
            "nextLevelName": NextLevelName,
            "generatedOn": _admin_report_datetime(datetime.now(datetime_timezone.utc), TimezoneName, TimezoneOffsetMinutes),
        },
        "performance": {
            "assessmentName": AssessmentTitle,
            "assessmentScore": ScoreText,
            "assessmentPercentage": PercentageText,
            "assessmentResult": ResultLabel if ResultLabel != "Assessment Cleared" else "Assessment Milestone Cleared",
            "assessmentDate": AssessmentDate,
            "practiceProgress": PracticeProgress,
            "practiceAccuracy": f"{PracticeAccuracy}%",
            "practiceCompleted": CompletedDps,
            "practiceTotal": RequiredDps,
        },
        "summary": {
            "message": Message,
            "parentGuidance": ParentGuidance,
            "nextStep": f"Begin {NextLevelCode} Practice" if NextLevelCode not in {"-", "Next Level", "Next Level Pending Setup"} else "Continue Teacher-Guided Practice",
        },
        "movements": MovementRows,
        "generatedOn": _admin_report_datetime(datetime.now(datetime_timezone.utc), TimezoneName, TimezoneOffsetMinutes),
    }


@router.get("/results/export/parent-summary")
def export_parent_progress_summary(studentId: str, moduleId: str | None = None, levelId: str | None = None, lessonId: str | None = None, dpsId: str | None = None, timezone: str | None = None, timezoneOffsetMinutes: int | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    StudentRecord, ModuleRecord, LevelRecord = _admin_validate_parent_report_scope(db, StudentId=studentId, ModuleId=moduleId, LevelId=levelId)
    Payload = _admin_build_student_report(db, StudentRecord.id, ModuleRecord.id, LevelRecord.id, lessonId, dpsId)
    StudentValue = Payload.get("student", {}) or {}
    ReportData = _admin_build_parent_progress_pdf_data(Payload, timezone, timezoneOffsetMinutes, db)
    ReportMeta, _ = _admin_validate_parent_report_data(ReportData, ExpectedModule=ModuleRecord, ExpectedLevel=LevelRecord)
    StudentName = StudentValue.get("studentName") or StudentValue.get("studentCode") or "Student"
    ReportLevel = ReportMeta.get("reportLevelCode") or "Level"
    FileName = f"{_admin_safe_filename_part(StudentName)}-Progress_Report-{_admin_safe_level_filename_part(ReportLevel)}"
    NotifyParentReportGenerated(
        db,
        actor_user_id=user.id,
        student_id=StudentRecord.id,
        student_code=StudentRecord.student_code,
        module_code=ReportMeta.get("reportModuleCode"),
        level_code=ReportMeta.get("reportLevelCode"),
        file_name=f"{FileName}.pdf",
    )
    db.commit()
    return BuildParentProgressPdfResponse(FileName, ReportData)


def _admin_parent_report_file_name(StudentName: str | None, ReportLevel: str | None) -> str:
    return f"{_admin_safe_filename_part(StudentName or 'Student')}-Progress_Report-{_admin_safe_level_filename_part(ReportLevel or 'Level')}.pdf"


def _admin_clean_email(Value: str | None) -> str:
    return str(Value or "").strip()


def _admin_valid_email(Value: str) -> bool:
    CleanValue = Value.strip()
    if not re.match(r"^[A-Za-z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$", CleanValue):
        return False
    LocalPart, DomainPart = CleanValue.rsplit("@", 1)
    if not LocalPart or not DomainPart:
        return False
    if ".." in LocalPart or ".." in DomainPart:
        return False
    if DomainPart.startswith(".") or DomainPart.endswith(".") or DomainPart.startswith("-") or DomainPart.endswith("-"):
        return False
    return True


def _admin_email_domain_has_mail_server(Value: str) -> bool:
    CleanValue = Value.strip().lower()
    if not _admin_valid_email(CleanValue):
        return False
    DomainPart = CleanValue.rsplit("@", 1)[1]
    try:
        import dns.resolver  # type: ignore

        Resolver = dns.resolver.Resolver()
        Resolver.lifetime = 3.0
        Resolver.timeout = 2.0
        try:
            MxAnswers = Resolver.resolve(DomainPart, "MX")
            if list(MxAnswers):
                return True
        except Exception:
            # Some domains accept mail on A/AAAA records even without MX. Use this as a
            # conservative fallback instead of accepting obviously unresolved domains.
            pass
    except Exception:
        pass

    try:
        socket.getaddrinfo(DomainPart, 25)
        return True
    except Exception:
        return False


def _admin_validate_parent_report_recipient_email(Value: str):
    if not _admin_valid_email(Value) or not _admin_email_domain_has_mail_server(Value):
        api_error(400, "INVALID_PARENT_EMAIL", "Invalid recipient email. Please enter a valid email address with an active mail domain.")


def _admin_validate_parent_report_scope(db: Session, *, StudentId: str, ModuleId: str | None = None, LevelId: str | None = None) -> tuple[Student, Module, Level]:
    StudentValue = db.get(Student, StudentId)
    if not StudentValue:
        api_error(404, "STUDENT_NOT_FOUND", "Student not found.")

    if not LevelId:
        api_error(400, "REPORT_LEVEL_REQUIRED", "Please choose the completed level report before generating or sending a parent report.")

    LevelValue = db.get(Level, LevelId)
    if not LevelValue:
        api_error(404, "REPORT_LEVEL_NOT_FOUND", "The selected report level could not be found.")

    ModuleValue = db.get(Module, ModuleId) if ModuleId else db.get(Module, LevelValue.module_id)
    if not ModuleValue:
        api_error(404, "REPORT_MODULE_NOT_FOUND", "The selected report module could not be found.")

    if LevelValue.module_id != ModuleValue.id:
        api_error(400, "REPORT_LEVEL_MODULE_MISMATCH", "The selected report level does not belong to the selected module.")

    if StudentValue.current_module_id and ModuleId and StudentValue.current_module_id != ModuleId:
        # Do not block historical completed-level reports from earlier modules. This check is kept
        # intentionally non-fatal because parent reports may be regenerated after promotion.
        pass

    return StudentValue, ModuleValue, LevelValue


def _admin_parse_percentage(Value: Any) -> float:
    try:
        return float(str(Value or "0").replace("%", "").strip())
    except (TypeError, ValueError):
        return 0.0


def _admin_validate_parent_report_data(ReportData: dict, *, ExpectedModule: Module | None = None, ExpectedLevel: Level | None = None) -> tuple[dict, dict]:
    ReportMeta = ReportData.get("report", {}) or {}
    PerformanceMeta = ReportData.get("performance", {}) or {}

    ReportLevelCode = str(ReportMeta.get("reportLevelCode") or "").strip()
    ReportModuleCode = str(ReportMeta.get("reportModuleCode") or "").strip()
    AssessmentName = str(PerformanceMeta.get("assessmentName") or "").strip()
    AssessmentResult = str(PerformanceMeta.get("assessmentResult") or "").strip().upper()
    PercentageValue = _admin_parse_percentage(PerformanceMeta.get("assessmentPercentage"))

    if ExpectedLevel and ReportLevelCode != ExpectedLevel.level_code:
        api_error(400, "PARENT_REPORT_SCOPE_MISMATCH", "The generated parent report does not match the selected completed level.")

    if ExpectedModule and ReportModuleCode not in {ExpectedModule.module_code, str(ExpectedModule.module_name or "")}:
        api_error(400, "PARENT_REPORT_MODULE_MISMATCH", "The generated parent report does not match the selected module.")

    if not ReportLevelCode or ReportLevelCode in {"-", "Not Started", "Completed Level"}:
        api_error(400, "PARENT_REPORT_LEVEL_INVALID", "A completed level is required before generating a parent progress report.")

    if not AssessmentName or AssessmentName in {"-", "Level Assessment"}:
        api_error(400, "PARENT_REPORT_ASSESSMENT_MISSING", "No cleared assessment record was found for this completed level.")

    if PercentageValue < BENCHMARK_PERCENTAGE or "CLEARED" not in AssessmentResult:
        api_error(400, "PARENT_REPORT_NOT_AVAILABLE", "This level is not eligible for a completed progress report yet.")

    NextLevelCode = str(ReportMeta.get("nextLevelCode") or "").strip()
    if NextLevelCode and NextLevelCode == ReportLevelCode:
        api_error(400, "PARENT_REPORT_NEXT_LEVEL_INVALID", "The next learning level could not be resolved correctly for this report.")

    return ReportMeta, PerformanceMeta


def _admin_clean_delivery_error(ErrorValue: Exception | str | None) -> str:
    TextValue = str(ErrorValue or "").strip()
    if not TextValue:
        return "Email delivery failed. Please check the recipient email or email configuration."
    LowerValue = TextValue.lower()
    if "authentication" in LowerValue or "username and password" in LowerValue or "535" in LowerValue:
        return "SMTP authentication failed. Please check the sender email and app password."
    if "recipient" in LowerValue or "address" in LowerValue or "invalid" in LowerValue:
        return "Email delivery failed. Please check the recipient email address."
    if "timeout" in LowerValue or "timed out" in LowerValue:
        return "Email delivery timed out. Please try again."
    if "network" in LowerValue or "unreachable" in LowerValue or "connection failed" in LowerValue:
        return "Email delivery could not connect to the SMTP service. Please verify the SMTP host, port, and Render outbound email access."
    if "configure" in LowerValue or "smtp" in LowerValue:
        return "Email service is not configured. Please check SMTP settings."
    return TextValue[:500]


def _admin_parent_report_recipients(StudentValue: Student, RequestValue: ParentReportEmailRequest) -> list[dict[str, str]]:
    Mode = str(RequestValue.recipientMode or "").upper().strip()
    FatherEmail = _admin_clean_email(StudentValue.father_email)
    MotherEmail = _admin_clean_email(StudentValue.mother_email)
    CustomEmail = _admin_clean_email(RequestValue.customEmail)

    Recipients: list[dict[str, str]] = []
    if Mode == "FATHER":
        if not FatherEmail:
            api_error(400, "FATHER_EMAIL_MISSING", "Father email is not available for this student.")
        Recipients.append({"email": FatherEmail, "type": "FATHER"})
    elif Mode == "MOTHER":
        if not MotherEmail:
            api_error(400, "MOTHER_EMAIL_MISSING", "Mother email is not available for this student.")
        Recipients.append({"email": MotherEmail, "type": "MOTHER"})
    elif Mode == "BOTH":
        if FatherEmail:
            Recipients.append({"email": FatherEmail, "type": "FATHER"})
        if MotherEmail and MotherEmail.lower() != FatherEmail.lower():
            Recipients.append({"email": MotherEmail, "type": "MOTHER"})
        if not Recipients:
            api_error(400, "PARENT_EMAIL_MISSING", "Parent email is not available for this student.")
    elif Mode == "CUSTOM":
        if not CustomEmail:
            api_error(400, "CUSTOM_EMAIL_MISSING", "Please enter a custom recipient email.")
        Recipients.append({"email": CustomEmail, "type": "CUSTOM"})
    else:
        api_error(400, "INVALID_RECIPIENT_MODE", "Please choose a valid recipient for the parent report.")

    for Recipient in Recipients:
        _admin_validate_parent_report_recipient_email(Recipient["email"])
    return Recipients


def _admin_parent_report_email_body(StudentName: str, ReportLevel: str) -> str:
    return (
        f"Dear Parent,\n\n"
        f"Please find attached {StudentName}'s MathPath Progress Report for {ReportLevel}.\n\n"
        "The report includes the completed level summary, assessment outcome, practice performance, "
        "next learning step, and parent guidance for continued support.\n\n"
        "Warm regards,\n"
        "MathPath Team"
    )


def _admin_create_parent_report_email_logs(db: Session, *, StudentValue: Student, ReportMeta: dict, Recipients: list[dict[str, str]], FileName: str, UserValue: User, Status: str = "PENDING", ErrorMessage: str | None = None) -> list[ParentReportEmailLog]:
    Logs: list[ParentReportEmailLog] = []
    for Recipient in Recipients:
        LogValue = ParentReportEmailLog(
            student_id=StudentValue.id,
            student_code=StudentValue.student_code,
            module_code=ReportMeta.get("reportModuleCode"),
            level_code=ReportMeta.get("reportLevelCode"),
            recipient_email=Recipient["email"],
            recipient_type=Recipient["type"],
            file_name=FileName,
            status=Status,
            sent_by_user_id=UserValue.id,
            error_message=ErrorMessage,
        )
        db.add(LogValue)
        Logs.append(LogValue)
    db.commit()
    return Logs


def _admin_update_parent_report_email_logs(db: Session, Logs: list[ParentReportEmailLog], *, Status: str, ErrorMessage: str | None = None):
    NowValue = datetime.now(datetime_timezone.utc)
    for LogValue in Logs:
        LogValue.status = Status
        LogValue.error_message = ErrorMessage
        LogValue.delivery_status = Status
        LogValue.last_attempt_at = NowValue
        LogValue.attempt_count = int(LogValue.attempt_count or 0) + (1 if Status in {"SENT", "FAILED"} else 0)
        if Status == "SENT":
            LogValue.sent_at = NowValue
            LogValue.error_message = None
        elif Status == "FAILED":
            LogValue.sent_at = None
    db.commit()


def _admin_deliver_parent_report_email_now(
    db: Session,
    *,
    Logs: list[ParentReportEmailLog],
    RecipientEmails: list[str],
    Subject: str,
    Body: str,
    AttachmentBytes: bytes,
    AttachmentFileName: str,
    ActorUserId: str | None,
) -> dict:
    """Deliver a parent report through SMTP and persist a final audit state.

    This intentionally avoids leaving rows stuck in PENDING. The UI receives a
    final SENT/FAILED result and the delivery-history table remains trustworthy.
    """
    try:
        SendResult = SendEmailWithAttachment(
            Recipients=RecipientEmails,
            Subject=Subject,
            Body=Body,
            AttachmentBytes=AttachmentBytes,
            AttachmentFileName=AttachmentFileName,
        )
    except (EmailConfigurationError, EmailSendError, Exception) as ErrorValue:
        MessageValue = _admin_clean_delivery_error(ErrorValue)
        _admin_update_parent_report_email_logs(db, Logs, Status="FAILED", ErrorMessage=MessageValue)
        NotifyParentReportDeliveryLogs(
            db,
            actor_user_id=ActorUserId,
            logs=Logs,
            event="PARENT_REPORT_FAILED",
            status="FAILED",
            file_name=AttachmentFileName,
            error_message=MessageValue,
        )
        db.commit()
        return {
            "sent": False,
            "queued": False,
            "status": "FAILED",
            "message": MessageValue,
            "errorMessage": MessageValue,
            "provider": "SMTP",
        }

    for LogValue in Logs:
        LogValue.delivery_provider = str(SendResult.get("provider") or "SMTP")
        LogValue.provider_message_id = SendResult.get("providerMessageId")
        LogValue.provider_response = str(SendResult.get("providerResponse") or "SMTP accepted message")
    _admin_update_parent_report_email_logs(db, Logs, Status="SENT")
    NotifyParentReportDeliveryLogs(
        db,
        actor_user_id=ActorUserId,
        logs=Logs,
        event="PARENT_REPORT_SENT",
        status="SENT",
        file_name=AttachmentFileName,
    )
    db.commit()
    return {
        "sent": True,
        "queued": False,
        "status": "SENT",
        "message": "Parent progress report email was sent successfully.",
        "provider": str(SendResult.get("provider") or "SMTP"),
        "providerResponse": str(SendResult.get("providerResponse") or "SMTP accepted message"),
    }



def _admin_parent_report_module_label(db: Session, ModuleCode: str | None) -> tuple[str, str]:
    SafeCode = str(ModuleCode or "-").strip() or "-"
    ModuleValue = db.query(Module).filter(Module.module_code == SafeCode).first() if SafeCode != "-" else None
    ModuleName = ModuleValue.module_name if ModuleValue else SafeCode
    ModuleLabel = f"{ModuleName} · {SafeCode}" if ModuleName and ModuleName != SafeCode else SafeCode
    return ModuleName, ModuleLabel


def _admin_parent_report_level_label(db: Session, ModuleCode: str | None, LevelCode: str | None) -> tuple[str, str]:
    SafeCode = str(LevelCode or "-").strip() or "-"
    LevelValue = None
    if SafeCode != "-":
        QueryValue = db.query(Level)
        if ModuleCode:
            ModuleValue = db.query(Module).filter(Module.module_code == ModuleCode).first()
            if ModuleValue:
                QueryValue = QueryValue.filter(Level.module_id == ModuleValue.id)
        LevelValue = QueryValue.filter(Level.level_code == SafeCode).first()
    LevelName = LevelValue.level_name if LevelValue else SafeCode
    LevelLabel = f"{SafeCode} · {LevelName}" if LevelName and LevelName != SafeCode else SafeCode
    return LevelName, LevelLabel


@router.get("/system/smtp-diagnostic")
def admin_smtp_diagnostic(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return DiagnoseSmtpConfiguration()


@router.get("/results/parent-report-deliveries")
def list_parent_report_deliveries(
    moduleId: str | None = None,
    levelId: str | None = None,
    status: str | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    QueryValue = db.query(ParentReportEmailLog)
    if moduleId and moduleId != "ALL":
        QueryValue = QueryValue.filter(ParentReportEmailLog.module_code == moduleId)
    if levelId and levelId != "ALL":
        QueryValue = QueryValue.filter(ParentReportEmailLog.level_code == levelId)
    if status and status != "ALL":
        QueryValue = QueryValue.filter(ParentReportEmailLog.status == status)

    Logs = QueryValue.order_by(ParentReportEmailLog.created_at.desc()).all()
    SearchText = str(search or "").strip().lower()
    Payload = []
    for LogValue in Logs:
        StudentValue = db.get(Student, LogValue.student_id) if LogValue.student_id else None
        StudentUser = db.get(User, StudentValue.user_id) if StudentValue and StudentValue.user_id else None
        SentByUser = db.get(User, LogValue.sent_by_user_id) if LogValue.sent_by_user_id else None
        StudentName = StudentUser.full_name if StudentUser and StudentUser.full_name else (LogValue.student_code or "Student")
        StudentCode = LogValue.student_code or (StudentValue.student_code if StudentValue else "-")
        ModuleName, ModuleLabel = _admin_parent_report_module_label(db, LogValue.module_code)
        LevelName, LevelLabel = _admin_parent_report_level_label(db, LogValue.module_code, LogValue.level_code)
        RowValue = {
            "id": LogValue.id,
            "studentId": LogValue.student_id,
            "studentName": StudentName,
            "studentCode": StudentCode,
            "moduleCode": LogValue.module_code or "-",
            "moduleName": ModuleName,
            "moduleLabel": ModuleLabel,
            "levelCode": LogValue.level_code or "-",
            "levelName": LevelName,
            "levelLabel": LevelLabel,
            "recipientEmail": LogValue.recipient_email,
            "recipientType": LogValue.recipient_type,
            "fileName": LogValue.file_name,
            "status": LogValue.status,
            "sentAt": LogValue.sent_at.isoformat() if LogValue.sent_at else None,
            "createdAt": LogValue.created_at.isoformat() if LogValue.created_at else None,
            "sentBy": SentByUser.full_name if SentByUser and SentByUser.full_name else "MathPath Admin",
            "errorMessage": LogValue.error_message,
        }
        if SearchText:
            SearchSource = " ".join(str(RowValue.get(Key) or "") for Key in [
                "studentName", "studentCode", "moduleCode", "moduleName", "levelCode", "levelName", "recipientEmail", "recipientType", "status", "fileName"
            ]).lower()
            if SearchText not in SearchSource:
                continue
        Payload.append(RowValue)
    return {"logs": Payload}


@router.post("/results/send-parent-summary")
def send_parent_progress_summary_email(
    RequestValue: ParentReportEmailRequest,
    BackgroundTasksValue: BackgroundTasks,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    StudentValue, ModuleValue, LevelValue = _admin_validate_parent_report_scope(
        db,
        StudentId=RequestValue.studentId,
        ModuleId=RequestValue.moduleId,
        LevelId=RequestValue.levelId,
    )
    StudentUser = db.get(User, StudentValue.user_id) if StudentValue.user_id else None
    StudentName = StudentUser.full_name if StudentUser and StudentUser.full_name else StudentValue.student_code

    Recipients = _admin_parent_report_recipients(StudentValue, RequestValue)
    Payload = _admin_build_student_report(db, StudentValue.id, ModuleValue.id, LevelValue.id, RequestValue.lessonId, RequestValue.dpsId)
    ReportData = _admin_build_parent_progress_pdf_data(Payload, RequestValue.timezone, RequestValue.timezoneOffsetMinutes, db)
    ReportMeta, _ = _admin_validate_parent_report_data(ReportData, ExpectedModule=ModuleValue, ExpectedLevel=LevelValue)
    ReportLevel = ReportMeta.get("reportLevelCode") or LevelValue.level_code or "Level"

    FileName = _admin_parent_report_file_name(StudentName, ReportLevel)
    PdfBytes = BuildParentProgressPdfBytes(FileName, ReportData)
    Subject = f"MathPath Progress Report - {StudentName} - {ReportLevel}"
    Body = _admin_parent_report_email_body(StudentName, ReportLevel)
    Logs = _admin_create_parent_report_email_logs(
        db,
        StudentValue=StudentValue,
        ReportMeta=ReportMeta,
        Recipients=Recipients,
        FileName=FileName,
        UserValue=user,
        Status="PENDING",
    )

    DeliveryResult = _admin_deliver_parent_report_email_now(
        db,
        Logs=Logs,
        RecipientEmails=[Item["email"] for Item in Recipients],
        Subject=Subject,
        Body=Body,
        AttachmentBytes=PdfBytes,
        AttachmentFileName=FileName,
        ActorUserId=user.id,
    )

    return {
        **DeliveryResult,
        "recipients": [Item["email"] for Item in Recipients],
        "fileName": FileName,
        "deliveryIds": [LogValue.id for LogValue in Logs],
    }


@router.post("/results/parent-report-deliveries/{delivery_id}/resend")
def resend_parent_report_delivery(delivery_id: str, BackgroundTasksValue: BackgroundTasks, RequestValue: ParentReportResendRequest | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    ExistingLog = db.get(ParentReportEmailLog, delivery_id)
    if not ExistingLog:
        api_error(404, "DELIVERY_LOG_NOT_FOUND", "Parent report delivery record not found.")

    StudentValue = db.get(Student, ExistingLog.student_id) if ExistingLog.student_id else None
    if not StudentValue:
        api_error(404, "STUDENT_NOT_FOUND", "Student linked to this delivery record was not found.")

    StudentUser = db.get(User, StudentValue.user_id) if StudentValue.user_id else None
    StudentName = StudentUser.full_name if StudentUser and StudentUser.full_name else StudentValue.student_code

    ResendMode = str((RequestValue.recipientMode if RequestValue else "SAME") or "SAME").upper().strip()
    if ResendMode in {"", "SAME"}:
        RecipientEmail = _admin_clean_email(ExistingLog.recipient_email)
        if not RecipientEmail:
            api_error(400, "INVALID_PARENT_EMAIL", "The recipient email on this delivery record is missing or invalid.")
        _admin_validate_parent_report_recipient_email(RecipientEmail)
        Recipients = [{"email": RecipientEmail, "type": ExistingLog.recipient_type or "CUSTOM"}]
    else:
        Recipients = _admin_parent_report_recipients(
            StudentValue,
            ParentReportEmailRequest(
                studentId=StudentValue.id,
                moduleId=None,
                levelId=None,
                recipientMode=ResendMode,
                customEmail=RequestValue.customEmail if RequestValue else None,
            ),
        )

    ModuleValue = None
    if ExistingLog.module_code and ExistingLog.module_code != "-":
        ModuleValue = db.query(Module).filter(Module.module_code == ExistingLog.module_code).first()
    LevelQuery = db.query(Level)
    if ModuleValue:
        LevelQuery = LevelQuery.filter(Level.module_id == ModuleValue.id)
    LevelValue = None
    if ExistingLog.level_code and ExistingLog.level_code != "-":
        LevelValue = LevelQuery.filter(Level.level_code == ExistingLog.level_code).first()
    if ExistingLog.level_code and ExistingLog.level_code != "-" and not LevelValue:
        api_error(400, "REPORT_LEVEL_NOT_FOUND", "The report level linked to this delivery record could not be found.")

    if not ModuleValue:
        api_error(400, "REPORT_MODULE_NOT_FOUND", "The report module linked to this delivery record could not be found.")
    if not LevelValue:
        api_error(400, "REPORT_LEVEL_NOT_FOUND", "The report level linked to this delivery record could not be found.")

    _admin_validate_parent_report_scope(db, StudentId=StudentValue.id, ModuleId=ModuleValue.id, LevelId=LevelValue.id)
    Payload = _admin_build_student_report(db, StudentValue.id, ModuleValue.id, LevelValue.id, None, None)
    ReportData = _admin_build_parent_progress_pdf_data(Payload, None, None, db)
    ReportMeta, _ = _admin_validate_parent_report_data(ReportData, ExpectedModule=ModuleValue, ExpectedLevel=LevelValue)
    ReportLevel = ReportMeta.get("reportLevelCode") or ExistingLog.level_code or "Level"

    FileName = _admin_parent_report_file_name(StudentName, ReportLevel)
    PdfBytes = BuildParentProgressPdfBytes(FileName, ReportData)
    Subject = f"MathPath Progress Report - {StudentName} - {ReportLevel}"
    Body = _admin_parent_report_email_body(StudentName, ReportLevel)
    Logs = _admin_create_parent_report_email_logs(
        db,
        StudentValue=StudentValue,
        ReportMeta=ReportMeta,
        Recipients=Recipients,
        FileName=FileName,
        UserValue=user,
    )

    DeliveryResult = _admin_deliver_parent_report_email_now(
        db,
        Logs=Logs,
        RecipientEmails=[Item["email"] for Item in Recipients],
        Subject=Subject,
        Body=Body,
        AttachmentBytes=PdfBytes,
        AttachmentFileName=FileName,
        ActorUserId=user.id,
    )

    return {
        **DeliveryResult,
        "recipients": [Item["email"] for Item in Recipients],
        "fileName": FileName,
        "deliveryIds": [LogValue.id for LogValue in Logs],
    }


@router.delete("/results/parent-report-deliveries/{delivery_id}")
def delete_parent_report_delivery(delivery_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    LogValue = db.get(ParentReportEmailLog, delivery_id)
    if not LogValue:
        api_error(404, "DELIVERY_LOG_NOT_FOUND", "Delivery record was not found or may have already been deleted.")
    NotifyParentReportDeliveryDeleted(db, actor_user_id=user.id, log=LogValue)
    db.delete(LogValue)
    db.commit()
    return {
        "deleted": True,
        "message": "Parent report delivery record deleted successfully.",
        "deliveryId": delivery_id,
    }


@router.get("/attempts/{attempt_id}")
def admin_attempt_detail(attempt_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    attempt = db.get(Attempt, attempt_id)
    if not attempt:
        api_error(404, "NOT_FOUND", "Attempt not found.")

    student = db.get(Student, attempt.student_id)
    student_user = db.get(User, student.user_id) if student else None
    assignment = db.get(Assignment, attempt.assignment_id) if attempt.assignment_id else None
    dps = db.get(DPS, attempt.dps_id)
    lesson = db.get(Lesson, dps.lesson_id) if dps else None
    level = db.get(Level, lesson.level_id) if lesson else None
    module = db.get(Module, level.module_id) if level else None

    payload = result_payload(db, attempt, include_review=True)
    payload.update(attempt_date_payload(attempt))
    payload.update(benchmark_payload_for_attempt(attempt))
    payload["student"] = {
        "studentId": student.id if student else None,
        "studentName": student_user.full_name if student_user else "-",
        "studentCode": student.student_code if student else "-",
        "className": student.class_name if student else None,
        "section": student.section if student else None,
    }
    payload["assignment"] = {
        "assignmentId": assignment.id if assignment else None,
        "title": assignment.title if assignment else None,
    }
    payload["dps"] = {
        "dpsId": dps.id if dps else None,
        "dpsNumber": dps.dps_number if dps else None,
        "dpsTitle": dps.dps_title if dps else None,
        "lessonId": lesson.id if lesson else None,
        "lessonNumber": lesson.lesson_number if lesson else None,
        "lessonTitle": lesson.lesson_title if lesson else None,
        "levelId": level.id if level else None,
        "levelCode": level.level_code if level else None,
        "levelName": level.level_name if level else None,
        "moduleId": module.id if module else None,
        "moduleCode": module.module_code if module else None,
        "moduleName": module.module_name if module else None,
    }
    return payload




@router.get("/assessment-readiness/testing-overrides")
def admin_list_assessment_testing_overrides(
    studentId: str | None = None,
    moduleId: str | None = None,
    levelId: str | None = None,
    activeOnly: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    QueryValue = db.query(AssessmentReadinessTestingOverride)
    if studentId:
        QueryValue = QueryValue.filter(AssessmentReadinessTestingOverride.student_id == studentId)
    if moduleId:
        QueryValue = QueryValue.filter(AssessmentReadinessTestingOverride.module_id == moduleId)
    if levelId:
        QueryValue = QueryValue.filter(AssessmentReadinessTestingOverride.level_id == levelId)
    if activeOnly:
        QueryValue = QueryValue.filter(AssessmentReadinessTestingOverride.status == "ACTIVE")
    Overrides = QueryValue.order_by(AssessmentReadinessTestingOverride.enabled_at.desc()).all()
    return {
        "testingOverrideEnabled": bool(ASSESSMENT_TESTING_OVERRIDE_ENABLED),
        "testingOverrideLabel": ASSESSMENT_TESTING_OVERRIDE_LABEL,
        "count": len(Overrides),
        "overrides": [_admin_assessment_testing_override_payload(db, OverrideValue) for OverrideValue in Overrides],
    }


@router.post("/assessment-readiness/testing-overrides")
def admin_create_assessment_testing_override(
    payload: AssessmentTestingOverrideCreateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    if not ASSESSMENT_TESTING_OVERRIDE_ENABLED:
        api_error(403, "ASSESSMENT_TESTING_OVERRIDE_DISABLED", "Admin testing override is disabled for this environment.")

    StudentValue = db.get(Student, payload.studentId)
    if not StudentValue:
        api_error(404, "STUDENT_NOT_FOUND", "Student not found.")

    LevelValue = db.get(Level, payload.levelId)
    if not LevelValue:
        api_error(404, "LEVEL_NOT_FOUND", "Level not found.")

    ModuleValue = db.get(Module, payload.moduleId) if payload.moduleId else None
    if payload.moduleId and not ModuleValue:
        api_error(404, "MODULE_NOT_FOUND", "Module not found.")

    if ModuleValue and LevelValue.module_id and LevelValue.module_id != ModuleValue.id:
        api_error(400, "LEVEL_MODULE_MISMATCH", "Selected level does not belong to the selected module.")

    if not ModuleValue:
        ModuleValue = db.get(Module, LevelValue.module_id) if LevelValue.module_id else None

    ExistingActive = _admin_active_assessment_testing_override(
        db,
        StudentId=StudentValue.id,
        LevelId=LevelValue.id,
        ModuleId=ModuleValue.id if ModuleValue else None,
    )
    if ExistingActive:
        return {
            "message": "Testing override is already active for this student and level.",
            "override": _admin_assessment_testing_override_payload(db, ExistingActive),
        }

    ReasonValue = (payload.reason or "").strip() or "Admin enabled controlled assessment readiness testing override."
    OverrideValue = AssessmentReadinessTestingOverride(
        student_id=StudentValue.id,
        student_code=StudentValue.student_code,
        module_id=ModuleValue.id if ModuleValue else None,
        module_code=ModuleValue.module_code if ModuleValue else None,
        module_name=ModuleValue.module_name if ModuleValue else None,
        level_id=LevelValue.id,
        level_code=LevelValue.level_code,
        level_name=LevelValue.level_name,
        status="ACTIVE",
        reason=ReasonValue,
        enabled_by_user_id=user.id,
        enabled_at=datetime.now(datetime_timezone.utc),
    )
    db.add(OverrideValue)
    db.commit()
    db.refresh(OverrideValue)
    return {
        "message": "Testing override enabled for this student and level.",
        "override": _admin_assessment_testing_override_payload(db, OverrideValue),
    }


@router.patch("/assessment-readiness/testing-overrides/{override_id}/deactivate")
def admin_deactivate_assessment_testing_override(
    override_id: str,
    payload: AssessmentTestingOverrideDeactivateRequest | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    OverrideValue = db.get(AssessmentReadinessTestingOverride, override_id)
    if not OverrideValue:
        api_error(404, "TESTING_OVERRIDE_NOT_FOUND", "Testing override record was not found.")

    if str(OverrideValue.status or "").upper() != "ACTIVE":
        return {
            "message": "Testing override is already inactive.",
            "override": _admin_assessment_testing_override_payload(db, OverrideValue),
        }

    DisableReason = (payload.reason if payload else None)
    ExistingReason = OverrideValue.reason or ""
    if DisableReason:
        OverrideValue.reason = f"{ExistingReason}\nDisabled Note: {DisableReason}".strip()
    OverrideValue.status = "DISABLED"
    OverrideValue.disabled_by_user_id = user.id
    OverrideValue.disabled_at = datetime.now(datetime_timezone.utc)
    OverrideValue.updated_at = datetime.now(datetime_timezone.utc)
    db.add(OverrideValue)
    db.commit()
    db.refresh(OverrideValue)
    return {
        "message": "Testing override disabled.",
        "override": _admin_assessment_testing_override_payload(db, OverrideValue),
    }

@router.get("/assessment-eligibility")
def admin_assessment_eligibility(levelId: str | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    students = db.query(Student).filter(Student.is_active == True).all()
    rows = eligibility_for_students(db, students, levelId)
    ready = [row for row in rows if row.get("eligible")]
    return {
        "benchmarkPercentage": 70,
        "totalStudents": len(rows),
        "readyCount": len(ready),
        "notReadyCount": len(rows) - len(ready),
        "readinessGate": AssessmentReadinessGateAuditPayload(rows),
        "rows": rows,
    }


@router.get("/assessment-eligibility/levels/{level_id}")
def admin_assessment_eligibility_by_level(level_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    students = db.query(Student).filter(Student.is_active == True, Student.current_level_id == level_id).all()
    rows = eligibility_for_students(db, students, level_id)
    ready = [row for row in rows if row.get("eligible")]
    return {
        "benchmarkPercentage": 70,
        "levelId": level_id,
        "totalStudents": len(rows),
        "readyCount": len(ready),
        "notReadyCount": len(rows) - len(ready),
        "readinessGate": AssessmentReadinessGateAuditPayload(rows),
        "rows": rows,
    }


@router.get("/students/{student_id}/assessment-eligibility")
def admin_student_assessment_eligibility(student_id: str, levelId: str | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    student = db.get(Student, student_id)
    if not student:
        api_error(404, "NOT_FOUND", "Student not found.")
    return assessment_eligibility_payload(db, student, levelId or student.current_level_id)



@router.get("/assessment-blueprints")
def admin_list_assessment_blueprints(
    status: str | None = None,
    moduleId: str | None = None,
    levelId: str | None = None,
    includeArchived: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    blueprints = list_blueprints(
        db,
        status=status,
        module_id=moduleId,
        level_id=levelId,
        include_archived=includeArchived,
    )
    return {
        "total": len(blueprints),
        "items": [blueprint_payload(db, blueprint) for blueprint in blueprints],
    }


@router.get("/assessment-reattempt-approvals")
def admin_list_assessment_reattempt_approvals(
    status: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    return ListAssessmentReattemptApprovals(db, Status=status)


@router.get("/assessment-reattempt-approvals/{approval_id}")
def admin_get_assessment_reattempt_approval(
    approval_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    approval = db.get(AssessmentReattemptApproval, approval_id)
    if not approval:
        api_error(404, "ASSESSMENT_REATTEMPT_APPROVAL_NOT_FOUND", "Assessment re-attempt approval request not found.")
    return AssessmentReattemptApprovalPayload(db, approval)


@router.post("/assessment-reattempt-approvals/{approval_id}/approve")
def admin_approve_assessment_reattempt(
    approval_id: str,
    payload: AssessmentReattemptDecisionRequest | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    approval = ApproveAssessmentReattempt(db, approval_id, user.id, payload.adminNote if payload else None)
    NotifyAssessmentReattemptDecision(db, approval_id=approval.id, actor_user_id=user.id, decision="APPROVED")
    db.commit()
    return {
        "updated": True,
        "message": "Assessment re-attempt approved.",
        "item": AssessmentReattemptApprovalPayload(db, approval),
    }


@router.post("/assessment-reattempt-approvals/{approval_id}/reject")
def admin_reject_assessment_reattempt(
    approval_id: str,
    payload: AssessmentReattemptDecisionRequest | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    approval = RejectAssessmentReattempt(db, approval_id, user.id, payload.adminNote if payload else None)
    NotifyAssessmentReattemptDecision(db, approval_id=approval.id, actor_user_id=user.id, decision="REJECTED")
    db.commit()
    return {
        "updated": True,
        "message": "Assessment re-attempt rejected.",
        "item": AssessmentReattemptApprovalPayload(db, approval),
    }


@router.get("/assessment-engine/foundation")
def admin_assessment_engine_foundation(
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    return AssessmentEngineFoundation(db)


@router.get("/assessment-blueprints/{blueprint_id}/engine-state")
def admin_assessment_blueprint_engine_state(
    blueprint_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    return BlueprintEngineState(db, blueprint_id)


@router.post("/assessment-blueprints/{blueprint_id}/generate-preview")
def admin_generate_assessment_preview(
    blueprint_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    return GenerateAssessmentPreview(db, blueprint_id, GeneratedByUserId=user.id)


@router.get("/assessment-blueprints/{blueprint_id}/generated-assessment")
def admin_get_generated_assessment(
    blueprint_id: str,
    includeAnswerKey: bool = True,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    assessment = LatestGeneratedVersion(db, blueprint_id, IncludeAnswerKey=includeAnswerKey)
    return {
        "available": assessment is not None,
        "assessment": assessment,
    }


@router.post("/assessment-blueprints/{blueprint_id}/versions/{version_id}/make-available")
def admin_make_assessment_version_available(
    blueprint_id: str,
    version_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    return SetAssessmentVersionAvailability(db, blueprint_id, version_id, True)


@router.post("/assessment-blueprints/{blueprint_id}/versions/{version_id}/pause")
def admin_pause_assessment_version(
    blueprint_id: str,
    version_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    return SetAssessmentVersionAvailability(db, blueprint_id, version_id, False)


@router.post("/assessment-blueprints")
def admin_create_assessment_blueprint(
    payload: AssessmentBlueprintCreateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    blueprint = create_blueprint(
        db,
        title=payload.title,
        module_id=payload.moduleId,
        level_id=payload.levelId,
        total_questions=payload.totalQuestions,
        duration_seconds=payload.durationSeconds,
        lesson_distribution=[item.model_dump() for item in payload.lessonDistribution],
        instructions=payload.instructions,
        status=payload.status,
        created_by_user_id=user.id,
    )
    return blueprint_payload(db, blueprint)


@router.get("/assessment-blueprints/{blueprint_id}")
def admin_get_assessment_blueprint(
    blueprint_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    blueprint = db.get(AssessmentBlueprint, blueprint_id)
    if not blueprint:
        api_error(404, "ASSESSMENT_BLUEPRINT_NOT_FOUND", "Assessment blueprint not found.")
    return blueprint_payload(db, blueprint)


@router.patch("/assessment-blueprints/{blueprint_id}")
def admin_update_assessment_blueprint(
    blueprint_id: str,
    payload: AssessmentBlueprintUpdateRequest,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    blueprint = update_blueprint(
        db,
        blueprint_id,
        title=payload.title,
        total_questions=payload.totalQuestions,
        duration_seconds=payload.durationSeconds,
        lesson_distribution=[item.model_dump() for item in payload.lessonDistribution] if payload.lessonDistribution is not None else None,
        instructions=payload.instructions,
    )
    return blueprint_payload(db, blueprint)


@router.post("/assessment-blueprints/{blueprint_id}/publish")
def admin_publish_assessment_blueprint(
    blueprint_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    blueprint = publish_blueprint(db, blueprint_id, published_by_user_id=user.id)
    return blueprint_payload(db, blueprint)


@router.post("/assessment-blueprints/{blueprint_id}/archive")
def admin_archive_assessment_blueprint(
    blueprint_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    blueprint = archive_blueprint(db, blueprint_id)
    return blueprint_payload(db, blueprint)


@router.delete("/assessment-blueprints/{blueprint_id}")
def admin_delete_assessment_blueprint(
    blueprint_id: str,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    blueprint_snapshot = delete_blueprint(db, blueprint_id)
    return {
        "ok": True,
        "message": "Assessment permanently deleted.",
        "item": blueprint_snapshot,
    }


@router.get("/competition/mock-section-plan")
def admin_get_competition_mock_section_plan(levelId: str, totalQuestions: int | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return CompetitionMockSectionPlan(db, LevelId=levelId, TotalQuestions=totalQuestions)


@router.post("/competition/mock-exams/generate-draft")
def admin_generate_competition_mock_draft(payload: CompetitionMockGenerateRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return GenerateCompetitionMockDraft(
        db,
        LevelId=payload.levelId,
        CreatedBy=user,
        Title=payload.title,
        MockCode=payload.mockCode,
        TotalQuestions=payload.totalQuestions,
        DurationSeconds=payload.durationSeconds,
        CompetitionScope=payload.competitionScope or "GENERAL",
        DifficultyBand=payload.difficultyBand or "COMPETITION",
        SectionCounts=payload.sectionCounts,
    )


@router.get("/competition/mock-exams")
def admin_list_competition_mock_exams(levelId: str | None = None, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return {"mockExams": ListCompetitionMockDrafts(db, LevelId=levelId)}


@router.get("/competition/mock-exams/{mock_exam_id}")
def admin_get_competition_mock_exam(mock_exam_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    ExamRecord = db.get(CompetitionMockExam, mock_exam_id)
    if not ExamRecord or not ExamRecord.is_active:
        api_error(404, "COMPETITION_MOCK_NOT_FOUND", "Competition mock exam was not found.")
    return CompetitionMockExamPayload(db, ExamRecord, IncludeQuestions=True)

@router.delete("/competition/mock-exams/{mock_exam_id}")
def admin_delete_competition_mock_exam(mock_exam_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    Deleted = DeleteCompetitionMockExam(db, MockExamId=mock_exam_id)
    return {"ok": True, "message": "Competition mock exam permanently deleted.", "deleted": Deleted}


@router.patch("/competition/mock-exams/{mock_exam_id}/archive")
def admin_archive_competition_mock_exam(mock_exam_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    Archived = ArchiveCompetitionMockExam(db, MockExamId=mock_exam_id)
    return {"ok": True, "message": "Competition mock exam archived.", "mockExam": Archived}


@router.post("/competition/mock-exams/assign")
def admin_assign_competition_mock_exams(payload: CompetitionMockAssignRequest, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return AssignCompetitionMockExams(
        db,
        LevelId=payload.levelId,
        MockExamIds=payload.mockExamIds,
        AssignedBy=user,
        StudentIds=payload.studentIds,
        AssignToAllInLevel=payload.assignToAllInLevel,
        MaxAttempts=payload.maxAttempts or 1,
        DueAt=payload.dueAt,
        Instructions=payload.instructions,
    )


@router.get("/competition/mock-assignments")
def admin_list_competition_mock_assignments(
    levelId: str | None = None,
    mockExamId: str | None = None,
    studentId: str | None = None,
    status: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(admin_dep),
):
    return {
        "assignments": ListCompetitionMockAssignments(
            db,
            LevelId=levelId,
            MockExamId=mockExamId,
            StudentId=studentId,
            Status=status,
        )
    }


from app.api.routes_teacher import _teacher_competition_row_payload, _competition_duration_text
from app.services.competition_mock_attempt_service import GetCompetitionMockResultForAdmin

@router.get("/competition/mock-tracker")
def admin_competition_mock_tracker(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    from app.api.routes_admin import clean_text
    assignments = (
        db.query(CompetitionMockAssignment)
        .join(CompetitionMockExam, CompetitionMockAssignment.mock_exam_id == CompetitionMockExam.id)
        .filter(
            CompetitionMockAssignment.is_active == True,
            CompetitionMockExam.is_active == True,
            CompetitionMockExam.status != "ARCHIVED",
        )
        .order_by(CompetitionMockAssignment.assigned_at.desc())
        .limit(1000)
        .all()
    )
    rows = []
    for assignment in assignments:
        row = _teacher_competition_row_payload(db, assignment)
        if not row:
            continue
        student = db.get(Student, assignment.student_id)
        teacher_name = None
        teacher_code = None
        if student:
            StoredTeacherName = clean_text(student.teacher) if hasattr(student, "teacher") else None
            TargetTeacher = None
            if hasattr(student, "teacher_id") and getattr(student, "teacher_id", None):
                TargetTeacher = db.get(Teacher, getattr(student, "teacher_id"))
            if not TargetTeacher and StoredTeacherName:
                TargetTeacher = (
                    db.query(Teacher)
                    .join(User, Teacher.user_id == User.id)
                    .filter(User.full_name == StoredTeacherName)
                    .first()
                )
            TargetTeacherUser = db.get(User, TargetTeacher.user_id) if TargetTeacher else None
            teacher_name = TargetTeacherUser.full_name if TargetTeacherUser else StoredTeacherName
            teacher_code = TargetTeacher.teacher_code if TargetTeacher else None
        
        row["teacherName"] = teacher_name
        row["teacherCode"] = teacher_code
        rows.append(row)

    completed = [row for row in rows if row.get("status") == "COMPLETED"]
    pending = [row for row in rows if row.get("status") in {"ASSIGNED", "PENDING"}]
    in_progress = [row for row in rows if row.get("status") == "IN_PROGRESS"]
    avg_score = round(sum(float(row.get("percentage") or 0) for row in completed) / len(completed)) if completed else 0
    avg_accuracy = round(sum(float(row.get("accuracyPercentage") or 0) for row in completed) / len(completed)) if completed else 0
    time_values = [int(row.get("timeTakenSeconds") or 0) for row in completed if row.get("timeTakenSeconds") is not None]
    avg_time = round(sum(time_values) / len(time_values)) if time_values else None
    return {
        "summary": {
            "assignedCount": len(rows),
            "completedCount": len(completed),
            "pendingCount": len(pending),
            "inProgressCount": len(in_progress),
            "averageScore": avg_score,
            "averageAccuracy": avg_accuracy,
            "averageTimeTakenSeconds": avg_time,
            "averageTimeTakenText": _competition_duration_text(avg_time),
        },
        "rows": rows,
    }


@router.delete("/competition/mock-tracker/assignment/{assignment_id}")
def admin_delete_competition_mock_assignment(assignment_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    assignment = db.get(CompetitionMockAssignment, assignment_id)
    if not assignment:
        raise HTTPException(status_code=404, detail="Mock assignment not found.")
    assignment.is_active = False
    db.commit()
    return {"status": "success", "message": "Assignment deleted successfully."}

@router.delete("/competition/mock-tracker/student/{student_id}")
def admin_delete_competition_mock_student(student_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    assignments = db.query(CompetitionMockAssignment).filter(CompetitionMockAssignment.student_id == student_id).all()
    for assignment in assignments:
        assignment.is_active = False
    db.commit()
    return {"status": "success", "message": f"Deleted {len(assignments)} assignments for student."}

@router.get("/competition/mock-attempts/{attempt_id}/result")

def admin_get_competition_mock_result(attempt_id: str, db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    return GetCompetitionMockResultForAdmin(db, attempt_id)

@router.post("/gamification/sync-all")
def gamification_sync_all(db: Session = Depends(get_db), user: User = Depends(admin_dep)):
    from app.services.achievements import AchievementEngine
    from app.models.models import CompetitionMockResultSummary, StudentAchievementStat, StudentBadge
    try:
        # Clear stats and student badges
        db.query(StudentBadge).delete()
        db.query(StudentAchievementStat).delete()
        db.commit()

        # Seed badges
        AchievementEngine.seed_badges(db)

        # Iterate all results and evaluate
        summaries = db.query(CompetitionMockResultSummary).order_by(CompetitionMockResultSummary.completed_at.asc()).all()
        for s in summaries:
            try:
                AchievementEngine.evaluate_mock_exam_submission(db, s.student_id, s)
            except Exception as e:
                db.rollback()
                import logging
                logging.error(f"Error evaluating {s.id}: {e}")
        return {"status": "success", "synced_count": len(summaries)}
    except Exception as e:
        db.rollback()
        return {"status": "error", "detail": str(e)}


