from __future__ import annotations

import base64
import math
from datetime import datetime, timezone, timedelta
import asyncio
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Paragraph
from reportlab.pdfgen import canvas as PdfCanvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


HeaderFill = PatternFill("solid", fgColor="0F172A")
HeaderFont = Font(color="FFFFFF", bold=True)
SubtleFill = PatternFill("solid", fgColor="EAF4FF")
BorderLine = Side(style="thin", color="D8E2F0")
CellBorder = Border(left=BorderLine, right=BorderLine, top=BorderLine, bottom=BorderLine)


def SafeCell(Value: Any) -> Any:
    if Value is None:
        return "-"
    if isinstance(Value, (int, float)):
        return Value
    if isinstance(Value, datetime):
        return Value.strftime("%d-%b-%Y, %I:%M %p")
    Text = str(Value)
    if Text and Text[0] in {"=", "+", "-", "@"}:
        return "'" + Text
    return Text


def NormalizeRows(Rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return [dict(Row) for Row in Rows]


def _ColumnWidth(HeaderText: str) -> float:
    Widths = {
        "Student Name": 18,
        "Student Code": 17,
        "Teacher Name": 18,
        "Teacher Code": 17,
        "Module Code": 17,
        "Module": 21,
        "Level Code": 15,
        "Level": 22,
        "Lesson": 46,
        "DPS": 54,
        "Status": 17,
        "Score": 13,
        "Total Marks": 16,
        "Accuracy %": 16,
        "Benchmark Status": 20,
        "Correct Answers": 18,
        "Completion Date": 18,
        "Completion Time": 18,
        "Time Taken": 16,
        "Required DPS": 16,
        "DPS Cleared": 17,
        "Passed DPS": 16,
        "Pending DPS": 16,
        "Needs Re-Attempt": 20,
        "Average Score": 17,
        "Average Accuracy %": 20,
        "Performance Zone": 20,
        "Assessment Readiness": 22,
        "Promotion Status": 20,
        "From Module": 16,
        "From Level": 16,
        "To Module": 16,
        "To Level": 16,
        "Promoted Levels": 18,
        "Promotion History Records": 26,
        "Promotion Assessment": 32,
        "Promotion Score": 18,
        "Promotion Percentage": 22,
        "Promoted Date": 18,
        "Promoted Time": 18,
        "Promoted By": 22,
        "Last Activity Date": 18,
        "Assessment": 30,
        "Result": 18,
    }
    if HeaderText in Widths:
        return float(Widths[HeaderText])
    if HeaderText.endswith("Code"):
        return 17.0
    if HeaderText.endswith("Name"):
        return 24.0
    if "Date" in HeaderText or "Time" in HeaderText:
        return 18.0
    if "%" in HeaderText or "Score" in HeaderText:
        return 17.0
    return 18.0


def AddRowsSheet(WorkbookValue: Workbook, Title: str, Rows: Iterable[dict[str, Any]], EmptyMessage: str = "No report data found for the selected scope."):
    Sheet = WorkbookValue.create_sheet(Title[:31])
    RowList = NormalizeRows(Rows)
    if not RowList:
        Sheet.append([EmptyMessage])
        Sheet["A1"].font = Font(bold=True, color="334155")
        Sheet["A1"].fill = SubtleFill
        Sheet["A1"].border = CellBorder
        Sheet["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        Sheet.row_dimensions[1].height = 24
        Sheet.column_dimensions["A"].width = 58
        return Sheet

    Headers = list(RowList[0].keys())
    Sheet.append(Headers)
    for Cell in Sheet[1]:
        Cell.font = HeaderFont
        Cell.fill = HeaderFill
        Cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        Cell.border = CellBorder
    Sheet.row_dimensions[1].height = 24

    for Row in RowList:
        Sheet.append([SafeCell(Row.get(Header)) for Header in Headers])

    for RowIndex, RowCells in enumerate(Sheet.iter_rows(min_row=2), start=2):
        Sheet.row_dimensions[RowIndex].height = 28
        for Cell in RowCells:
            Cell.border = CellBorder
            Cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    Sheet.freeze_panes = "A2"
    Sheet.auto_filter.ref = Sheet.dimensions
    for Index, Header in enumerate(Headers, start=1):
        Sheet.column_dimensions[get_column_letter(Index)].width = _ColumnWidth(str(Header))
    return Sheet

def AddSummarySheet(WorkbookValue: Workbook, SummaryRows: Iterable[tuple[str, Any]]):
    Sheet = WorkbookValue.active
    Sheet.title = "Report Summary"
    Sheet.append(["Report Field", "Report Value"])
    for Cell in Sheet[1]:
        Cell.font = HeaderFont
        Cell.fill = HeaderFill
        Cell.border = CellBorder
        Cell.alignment = Alignment(horizontal="center", vertical="center")
    Sheet.row_dimensions[1].height = 24
    for Label, Value in SummaryRows:
        Sheet.append([SafeCell(Label), SafeCell(Value)])
    for RowIndex, RowCells in enumerate(Sheet.iter_rows(min_row=2), start=2):
        Sheet.row_dimensions[RowIndex].height = 22
        for Cell in RowCells:
            Cell.border = CellBorder
            Cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        RowCells[0].font = Font(bold=True, color="334155")
        RowCells[0].fill = SubtleFill
    Sheet.column_dimensions["A"].width = 32
    Sheet.column_dimensions["B"].width = 58
    Sheet.freeze_panes = "A2"
    return Sheet

def BuildWorkbookResponse(FileName: str, SummaryRows: Iterable[tuple[str, Any]], Sheets: list[tuple[str, Iterable[dict[str, Any]]]]) -> StreamingResponse:
    WorkbookValue = Workbook()
    AddSummarySheet(WorkbookValue, list(SummaryRows))
    for Title, Rows in Sheets:
        AddRowsSheet(WorkbookValue, Title, Rows)

    Buffer = BytesIO()
    WorkbookValue.save(Buffer)
    Buffer.seek(0)
    SafeFileName = FileName if FileName.lower().endswith(".xlsx") else f"{FileName}.xlsx"
    return StreamingResponse(
        Buffer,
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": f'attachment; filename="{SafeFileName}"'},
    )


def ReportGeneratedOn() -> str:
    IndiaTime = datetime.now(timezone(timedelta(hours=5, minutes=30)))
    return IndiaTime.strftime("%d-%b-%Y, %I:%M %p")

PDF_MIME = "application/pdf"


# =============================================================================
# MathPath Parent Progress Report - premium canvas renderer
#
# One art-directed system: a single palette (Mp*), a single type scale, and
# bespoke canvas-drawn visuals (achievement medallion, journey path, mastery
# bars) instead of stock chart widgets. Page 1 is the milestone story; Page 2
# is the evidence (accuracy profile, level-by-level mastery, progression
# timeline) plus guidance for home. Content is deliberately not repeated
# between the two pages.
# =============================================================================

MpInk = colors.HexColor("#0A1633")
MpText = colors.HexColor("#26334D")
MpMuted = colors.HexColor("#69789A")
MpFaint = colors.HexColor("#93A3C2")
MpLine = colors.HexColor("#DCE6F5")
MpPage = colors.HexColor("#F5F9FF")
MpWhite = colors.white
MpBlue = colors.HexColor("#2563EB")
MpBlueDark = colors.HexColor("#1D4ED8")
MpBlueSoft = colors.HexColor("#EAF1FE")
MpCyan = colors.HexColor("#0E9FC4")
MpCyanLight = colors.HexColor("#67E8F9")
MpTeal = colors.HexColor("#0D9488")
MpTealSoft = colors.HexColor("#E4F8F4")
MpGreen = colors.HexColor("#059669")
MpGreenSoft = colors.HexColor("#E7F8F0")
MpGold = colors.HexColor("#F59E0B")
MpGoldDark = colors.HexColor("#B45309")
MpGoldSoft = colors.HexColor("#FEF3C7")
MpOrange = colors.HexColor("#F97316")
MpOrangeDark = colors.HexColor("#C2410C")
MpOrangeSoft = colors.HexColor("#FFEDD5")
MpPurple = colors.HexColor("#7C3AED")
MpPurpleSoft = colors.HexColor("#F1EBFE")
MpSlateSoft = colors.HexColor("#EDF2FA")

MpFontRegular = "Helvetica"
MpFontBold = "Helvetica-Bold"


def _RegisterMpFonts():
    global MpFontRegular, MpFontBold
    CandidatePairs = [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf", "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
        ("C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/arialbd.ttf"),
        ("C:/Windows/Fonts/calibri.ttf", "C:/Windows/Fonts/calibrib.ttf"),
    ]
    for RegularPath, BoldPath in CandidatePairs:
        try:
            if Path(RegularPath).exists() and Path(BoldPath).exists():
                pdfmetrics.registerFont(TTFont("MathPathReport-Regular", RegularPath))
                pdfmetrics.registerFont(TTFont("MathPathReport-Bold", BoldPath))
                MpFontRegular = "MathPathReport-Regular"
                MpFontBold = "MathPathReport-Bold"
                return
        except Exception:
            continue


_RegisterMpFonts()


def _PdfText(Value: Any) -> str:
    Text = SafeCell(Value)
    Text = str(Text).replace("—", "-").replace("–", "-").replace("•", "-").replace("’", "'")
    return Text


def _FindMathPathLogo() -> str | None:
    CurrentFile = Path(__file__).resolve()
    CandidatePaths = []
    for Parent in CurrentFile.parents:
        CandidatePaths.extend([
            Parent / "frontend" / "public" / "mathpath-logo.png",
            Parent / "public" / "mathpath-logo.png",
            Parent / "app" / "icon.png",
            Parent / "frontend" / "app" / "icon.png",
        ])
    for Candidate in CandidatePaths:
        if Candidate.exists() and Candidate.is_file():
            return str(Candidate)
    return None


def _MpNum(Value: Any) -> float:
    try:
        return float(str(Value).replace("%", "").strip())
    except Exception:
        return 0.0


def _MpTier(PercentageValue: Any) -> dict[str, Any]:
    Numeric = _MpNum(PercentageValue)
    if Numeric >= 90:
        return {"main": MpGold, "dark": MpGoldDark, "soft": MpGoldSoft, "label": "Excellence Zone"}
    if Numeric >= 70:
        return {"main": MpBlue, "dark": MpBlueDark, "soft": MpBlueSoft, "label": "Milestone Cleared"}
    return {"main": MpOrange, "dark": MpOrangeDark, "soft": MpOrangeSoft, "label": "Rising Star"}


def _MpZoneStyle(ZoneText: Any) -> tuple[Any, Any]:
    Zone = str(ZoneText or "").lower()
    if "excellence" in Zone:
        return MpGoldSoft, MpGoldDark
    if "growth" in Zone or "milestone" in Zone:
        return MpBlueSoft, MpBlueDark
    if "improvement" in Zone or "practice" in Zone:
        return MpOrangeSoft, MpOrangeDark
    return MpSlateSoft, MpMuted


def _MpStyle(Name: str, Size: float, Leading: float, Color=MpText, Bold: bool = False, Align=TA_LEFT) -> ParagraphStyle:
    return ParagraphStyle(
        Name,
        fontName=MpFontBold if Bold else MpFontRegular,
        fontSize=Size,
        leading=Leading,
        textColor=Color,
        alignment=Align,
        spaceAfter=0,
        spaceBefore=0,
    )


def _MpEsc(Value: Any) -> str:
    return _PdfText(Value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _MpPara(Canvas, TextValue: Any, X: float, TopY: float, Width: float, Style: ParagraphStyle) -> float:
    Value = Paragraph(_MpEsc(TextValue), Style)
    _, Height = Value.wrap(Width, 160 * mm)
    Value.drawOn(Canvas, X, TopY - Height)
    return Height


def _MpFit(Canvas, TextValue: Any, MaxWidth: float, Start: float, MinSize: float = 6.0, FontName: str | None = None) -> float:
    FontName = FontName or MpFontBold
    Size = Start
    Text = _PdfText(TextValue)
    while Size > MinSize and Canvas.stringWidth(Text, FontName, Size) > MaxWidth:
        Size -= 0.25
    return Size


def _MpDrawFitted(Canvas, TextValue: Any, X: float, Y: float, MaxWidth: float, Start: float, Color=MpInk, Bold: bool = True, Align: str = "left", MinSize: float = 6.0) -> float:
    FontName = MpFontBold if Bold else MpFontRegular
    Size = _MpFit(Canvas, TextValue, MaxWidth, Start, MinSize, FontName)
    Text = _PdfText(TextValue)
    Canvas.saveState()
    Canvas.setFillColor(Color)
    Canvas.setFont(FontName, Size)
    if Align == "center":
        Canvas.drawCentredString(X, Y, Text)
    elif Align == "right":
        Canvas.drawRightString(X, Y, Text)
    else:
        Canvas.drawString(X, Y, Text)
    Canvas.restoreState()
    return Size


def _MpTracked(Canvas, TextValue: Any, X: float, Y: float, Size: float = 6.6, Color=MpMuted, Track: float = 0.9, Align: str = "left") -> float:
    Text = _PdfText(TextValue).upper()
    Width = Canvas.stringWidth(Text, MpFontBold, Size) + Track * max(0, len(Text) - 1)
    if Align == "center":
        DrawX = X - Width / 2
    elif Align == "right":
        DrawX = X - Width
    else:
        DrawX = X
    Canvas.saveState()
    Canvas.setFillColor(Color)
    Canvas.setFont(MpFontBold, Size)
    Canvas.drawString(DrawX, Y, Text, charSpace=Track)
    Canvas.restoreState()
    return Width


def _MpPanel(Canvas, X: float, Y: float, W: float, H: float, Fill=MpWhite, Stroke=MpLine, Radius: float = 10, StrokeWidth: float = 0.8, Shadow: bool = False):
    Canvas.saveState()
    if Shadow:
        Canvas.setFillColor(colors.Color(0.16, 0.24, 0.42, alpha=0.08))
        Canvas.roundRect(X + 0.8, Y - 1.8, W, H, Radius, fill=1, stroke=0)
    Canvas.setFillColor(Fill)
    Canvas.setStrokeColor(Stroke)
    Canvas.setLineWidth(StrokeWidth)
    Canvas.roundRect(X, Y, W, H, Radius, fill=1, stroke=1)
    Canvas.restoreState()


def _MpChip(Canvas, TextValue: Any, CX: float, CY: float, Fill, TextColor, Size: float = 6.4, PadX: float = 6.0, H: float = 11.0) -> float:
    Text = _PdfText(TextValue)
    Width = Canvas.stringWidth(Text, MpFontBold, Size) + PadX * 2
    Canvas.saveState()
    Canvas.setFillColor(Fill)
    Canvas.roundRect(CX - Width / 2, CY - H / 2, Width, H, H / 2, fill=1, stroke=0)
    Canvas.setFillColor(TextColor)
    Canvas.setFont(MpFontBold, Size)
    Canvas.drawCentredString(CX, CY - Size * 0.34, Text)
    Canvas.restoreState()
    return Width


def _MpBar(Canvas, X: float, Y: float, W: float, H: float, Fraction: float, Color, Track=MpSlateSoft):
    Frac = max(0.0, min(1.0, Fraction))
    Canvas.saveState()
    Canvas.setFillColor(Track)
    Canvas.roundRect(X, Y, W, H, H / 2, fill=1, stroke=0)
    if Frac > 0:
        FillW = max(H, W * Frac)
        Canvas.setFillColor(Color)
        Canvas.roundRect(X, Y, FillW, H, H / 2, fill=1, stroke=0)
        Canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.55))
        Canvas.circle(X + FillW - H / 2, Y + H / 2, H * 0.2, fill=1, stroke=0)
    Canvas.restoreState()


def _MpStar(Canvas, CX: float, CY: float, Radius: float, Color=MpGold, InnerRatio: float = 0.45, Points: int = 5, RotationDeg: float = 90.0):
    Canvas.saveState()
    Canvas.setFillColor(Color)
    PathValue = Canvas.beginPath()
    Step = math.pi / Points
    Start = math.radians(RotationDeg)
    for Index in range(Points * 2):
        Rad = Radius if Index % 2 == 0 else Radius * InnerRatio
        Angle = Start + Index * Step
        PX = CX + Rad * math.cos(Angle)
        PY = CY + Rad * math.sin(Angle)
        if Index == 0:
            PathValue.moveTo(PX, PY)
        else:
            PathValue.lineTo(PX, PY)
    PathValue.close()
    Canvas.drawPath(PathValue, stroke=0, fill=1)
    Canvas.restoreState()


def _MpGlyph(Canvas, Kind: str, CX: float, CY: float, S: float, Color=MpBlue, StrokeW: float = 1.1):
    Canvas.saveState()
    Canvas.setStrokeColor(Color)
    Canvas.setFillColor(Color)
    Canvas.setLineWidth(StrokeW)
    Canvas.setLineCap(1)
    Canvas.setLineJoin(1)
    if Kind == "check":
        PathValue = Canvas.beginPath()
        PathValue.moveTo(CX - S * 0.55, CY + S * 0.02)
        PathValue.lineTo(CX - S * 0.12, CY - S * 0.42)
        PathValue.lineTo(CX + S * 0.58, CY + S * 0.45)
        Canvas.drawPath(PathValue, stroke=1, fill=0)
    elif Kind == "flag":
        Canvas.line(CX - S * 0.42, CY - S * 0.72, CX - S * 0.42, CY + S * 0.72)
        PathValue = Canvas.beginPath()
        PathValue.moveTo(CX - S * 0.42, CY + S * 0.72)
        PathValue.lineTo(CX + S * 0.62, CY + S * 0.32)
        PathValue.lineTo(CX - S * 0.42, CY - S * 0.02)
        PathValue.close()
        Canvas.drawPath(PathValue, stroke=0, fill=1)
    elif Kind == "calendar":
        Canvas.roundRect(CX - S * 0.6, CY - S * 0.58, S * 1.2, S * 1.05, S * 0.16, fill=0, stroke=1)
        Canvas.line(CX - S * 0.6, CY + S * 0.14, CX + S * 0.6, CY + S * 0.14)
        Canvas.line(CX - S * 0.28, CY + S * 0.47, CX - S * 0.28, CY + S * 0.68)
        Canvas.line(CX + S * 0.28, CY + S * 0.47, CX + S * 0.28, CY + S * 0.68)
        Canvas.circle(CX - S * 0.2, CY - S * 0.18, S * 0.08, fill=1, stroke=0)
        Canvas.circle(CX + S * 0.2, CY - S * 0.18, S * 0.08, fill=1, stroke=0)
    elif Kind == "doc":
        Canvas.roundRect(CX - S * 0.45, CY - S * 0.62, S * 0.9, S * 1.24, S * 0.12, fill=0, stroke=1)
        Canvas.line(CX - S * 0.22, CY + S * 0.26, CX + S * 0.22, CY + S * 0.26)
        Canvas.line(CX - S * 0.22, CY, CX + S * 0.22, CY)
        Canvas.line(CX - S * 0.22, CY - S * 0.26, CX + S * 0.04, CY - S * 0.26)
    elif Kind == "medal":
        Canvas.line(CX - S * 0.26, CY + S * 0.68, CX - S * 0.1, CY + S * 0.24)
        Canvas.line(CX + S * 0.26, CY + S * 0.68, CX + S * 0.1, CY + S * 0.24)
        Canvas.circle(CX, CY - S * 0.14, S * 0.42, fill=0, stroke=1)
        Canvas.circle(CX, CY - S * 0.14, S * 0.15, fill=1, stroke=0)
    elif Kind == "target":
        Canvas.circle(CX, CY, S * 0.6, fill=0, stroke=1)
        Canvas.circle(CX, CY, S * 0.32, fill=0, stroke=1)
        Canvas.circle(CX, CY, S * 0.09, fill=1, stroke=0)
    elif Kind == "levelup":
        Canvas.line(CX, CY - S * 0.6, CX, CY + S * 0.28)
        PathValue = Canvas.beginPath()
        PathValue.moveTo(CX - S * 0.36, CY + S * 0.14)
        PathValue.lineTo(CX, CY + S * 0.64)
        PathValue.lineTo(CX + S * 0.36, CY + S * 0.14)
        PathValue.close()
        Canvas.drawPath(PathValue, stroke=0, fill=1)
    elif Kind == "spark":
        PathValue = Canvas.beginPath()
        PathValue.moveTo(CX, CY + S)
        PathValue.lineTo(CX + S * 0.24, CY + S * 0.24)
        PathValue.lineTo(CX + S, CY)
        PathValue.lineTo(CX + S * 0.24, CY - S * 0.24)
        PathValue.lineTo(CX, CY - S)
        PathValue.lineTo(CX - S * 0.24, CY - S * 0.24)
        PathValue.lineTo(CX - S, CY)
        PathValue.lineTo(CX - S * 0.24, CY + S * 0.24)
        PathValue.close()
        Canvas.drawPath(PathValue, stroke=0, fill=1)
    elif Kind == "arrow":
        Canvas.line(CX - S * 0.65, CY, CX + S * 0.42, CY)
        PathValue = Canvas.beginPath()
        PathValue.moveTo(CX + S * 0.16, CY + S * 0.3)
        PathValue.lineTo(CX + S * 0.62, CY)
        PathValue.lineTo(CX + S * 0.16, CY - S * 0.3)
        Canvas.drawPath(PathValue, stroke=1, fill=0)
    elif Kind == "home":
        PathValue = Canvas.beginPath()
        PathValue.moveTo(CX - S * 0.62, CY + S * 0.05)
        PathValue.lineTo(CX, CY + S * 0.62)
        PathValue.lineTo(CX + S * 0.62, CY + S * 0.05)
        Canvas.drawPath(PathValue, stroke=1, fill=0)
        Canvas.roundRect(CX - S * 0.42, CY - S * 0.58, S * 0.84, S * 0.6, S * 0.08, fill=0, stroke=1)
    Canvas.restoreState()


def _MpBackground(Canvas, PageNumber: int = 1):
    Width, Height = A4
    Canvas.saveState()
    Canvas.setFillColor(MpPage)
    Canvas.rect(0, 0, Width, Height, fill=1, stroke=0)
    Canvas.setFillColor(colors.Color(0.85, 0.92, 1.0, alpha=0.55))
    Canvas.circle(Width + 4 * mm, 62 * mm, 38 * mm, fill=1, stroke=0)
    Canvas.setFillColor(colors.Color(0.90, 0.88, 1.0, alpha=0.35))
    Canvas.circle(-6 * mm, 128 * mm, 30 * mm, fill=1, stroke=0)
    Canvas.setFillColor(colors.HexColor("#DBE9FB"))
    for X in range(14, int(Width), 16):
        for Y in range(16, int(Height), 16):
            if (X * 7 + Y * 3 + PageNumber) % 89 == 0:
                Canvas.circle(X, Y, 0.5, fill=1, stroke=0)
    Canvas.restoreState()


def _MpFooter(Canvas, PageNumber: int, TotalPages: int, StudentCode: str, ReportLevel: str):
    Width, _ = A4
    Left = 14 * mm
    Right = Width - 14 * mm
    Canvas.saveState()
    Canvas.setStrokeColor(MpLine)
    Canvas.setLineWidth(0.7)
    Canvas.line(Left, 13.2 * mm, Right, 13.2 * mm)
    Canvas.setFillColor(MpMuted)
    Canvas.setFont(MpFontRegular, 6.8)
    Canvas.drawString(Left, 8.8 * mm, f"MathPath Parent Progress Report · {_PdfText(ReportLevel)} · Generated from verified learning records")
    Canvas.drawRightString(Right, 8.8 * mm, f"{_PdfText(StudentCode)} · Page {PageNumber} of {TotalPages}")
    Canvas.restoreState()


def _MpLogoChip(Canvas, LogoPath: str | None, X: float, Y: float, W: float, H: float, Shadow: bool = True):
    _MpPanel(Canvas, X, Y, W, H, MpWhite, MpLine, 8, 0.7, Shadow=Shadow)
    if LogoPath:
        try:
            Canvas.drawImage(LogoPath, X + 4, Y + 3, W - 8, H - 6, preserveAspectRatio=True, mask="auto", anchor="c")
            return
        except Exception:
            pass
    Canvas.saveState()
    Side = H - 8
    Canvas.setFillColor(MpBlue)
    Canvas.roundRect(X + 5, Y + 4, Side, Side, 2.5, fill=1, stroke=0)
    Canvas.setFillColor(MpWhite)
    for I in range(2):
        for J in range(2):
            Canvas.circle(X + 5 + Side * (0.32 + 0.36 * I), Y + 4 + Side * (0.32 + 0.36 * J), Side * 0.1, fill=1, stroke=0)
    Canvas.setFillColor(MpInk)
    Canvas.setFont(MpFontBold, H * 0.4)
    Canvas.drawString(X + 5 + Side + 4, Y + H / 2 - H * 0.15, "MathPath")
    Canvas.restoreState()


def _MpResolvePhoto(PhotoRef: Any) -> ImageReader | None:
    """Resolve a stored student photo reference into an ImageReader.

    Supports base64 data URIs and locally stored /uploads paths. Remote http(s)
    URLs are intentionally not fetched at render time; callers fall back to the
    initials avatar instead of risking a blocked or broken render.
    """
    if not PhotoRef or not isinstance(PhotoRef, str):
        return None
    Ref = PhotoRef.strip()
    if not Ref:
        return None
    try:
        if Ref.startswith("data:image"):
            _, _, Encoded = Ref.partition(",")
            if not Encoded:
                return None
            return ImageReader(BytesIO(base64.b64decode(Encoded)))
        if Ref.startswith("http://") or Ref.startswith("https://"):
            return None
        Candidate = Ref.lstrip("/")
        CandidatePaths = [Path(Candidate)]
        for Parent in Path(__file__).resolve().parents:
            CandidatePaths.append(Parent / Candidate)
        for PathValue in CandidatePaths:
            if PathValue.exists() and PathValue.is_file():
                return ImageReader(str(PathValue))
    except Exception:
        return None
    return None


def _MpAvatar(Canvas, CX: float, CY: float, Radius: float, PhotoReader: ImageReader | None, StudentName: str, Tier: dict[str, Any]):
    Canvas.saveState()
    Canvas.setFillColor(MpWhite)
    Canvas.circle(CX, CY, Radius + 2.8, fill=1, stroke=0)
    Canvas.setStrokeColor(Tier["main"])
    Canvas.setLineWidth(1.7)
    Canvas.circle(CX, CY, Radius + 1.7, fill=0, stroke=1)
    Drawn = False
    if PhotoReader is not None:
        try:
            ImgW, ImgH = PhotoReader.getSize()
            if ImgW > 0 and ImgH > 0:
                Scale = max((Radius * 2) / ImgW, (Radius * 2) / ImgH)
                DrawW, DrawH = ImgW * Scale, ImgH * Scale
                Canvas.saveState()
                ClipPath = Canvas.beginPath()
                ClipPath.circle(CX, CY, Radius)
                Canvas.clipPath(ClipPath, stroke=0, fill=0)
                Canvas.drawImage(PhotoReader, CX - DrawW / 2, CY - DrawH / 2, DrawW, DrawH, mask="auto")
                Canvas.restoreState()
                Drawn = True
        except Exception:
            Drawn = False
    if not Drawn:
        Canvas.setFillColor(Tier["main"])
        Canvas.circle(CX, CY, Radius, fill=1, stroke=0)
        Canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.2))
        Canvas.circle(CX - Radius * 0.32, CY + Radius * 0.36, Radius * 0.55, fill=1, stroke=0)
        Words = [Word for Word in _PdfText(StudentName).split() if Word and Word[0].isalpha()]
        Initials = "".join(Word[0] for Word in Words[:2]).upper() or "MP"
        Size = Radius * (0.9 if len(Initials) > 1 else 1.05)
        Canvas.setFillColor(MpWhite)
        Canvas.setFont(MpFontBold, Size)
        Canvas.drawCentredString(CX, CY - Size * 0.36, Initials)
    Canvas.restoreState()


def _MpAchievementBadge(Canvas, CX: float, CY: float, PercentText: str, ScoreText: str, Tier: dict[str, Any]):
    """Bespoke achievement medallion: segmented XP ring, scalloped rosette,
    star crest, and a ribbon banner carrying the performance band."""
    RingR = 20 * mm
    DiscR = 14.6 * mm
    Numeric = max(0.0, min(100.0, _MpNum(PercentText)))

    Canvas.saveState()
    Canvas.setLineCap(1)

    # Soft halo behind the medallion.
    Canvas.setFillColor(colors.Color(0.55, 0.68, 0.95, alpha=0.10))
    Canvas.circle(CX, CY, RingR + 4.5 * mm, fill=1, stroke=0)

    # Segmented progress ring (gamified XP ring, not a stock donut).
    Segments = 30
    SegmentSpan = 360.0 / Segments
    LitSegments = int(round(Numeric / 100.0 * Segments))
    Canvas.setLineWidth(3.4)
    for Index in range(Segments):
        StartAngle = 90.0 - Index * SegmentSpan
        Canvas.setStrokeColor(Tier["main"] if Index < LitSegments else MpLine)
        Canvas.arc(CX - RingR, CY - RingR, CX + RingR, CY + RingR, StartAngle - 1.2, -(SegmentSpan - 4.2))

    # Scalloped rosette edge.
    Canvas.setFillColor(Tier["dark"])
    for Index in range(12):
        Angle = math.radians(Index * 30)
        Canvas.circle(CX + DiscR * math.cos(Angle), CY + DiscR * math.sin(Angle), 2.5 * mm, fill=1, stroke=0)

    # Medallion discs.
    Canvas.setFillColor(colors.Color(0.06, 0.09, 0.2, alpha=0.16))
    Canvas.circle(CX, CY - 1.4, DiscR + 0.6, fill=1, stroke=0)
    Canvas.setFillColor(Tier["main"])
    Canvas.circle(CX, CY, DiscR, fill=1, stroke=0)
    Canvas.setFillColor(Tier["soft"])
    Canvas.circle(CX, CY, DiscR - 2.4 * mm, fill=1, stroke=0)
    Canvas.setFillColor(MpWhite)
    Canvas.circle(CX, CY, DiscR - 4.0 * mm, fill=1, stroke=0)

    # Star crest and score text.
    _MpStar(Canvas, CX, CY + 6.6 * mm, 3.1 * mm, Tier["main"])
    _MpStar(Canvas, CX, CY + 6.6 * mm, 1.15 * mm, MpWhite)
    Canvas.setFillColor(MpInk)
    PctSize = _MpFit(Canvas, PercentText, (DiscR - 4.0 * mm) * 1.7, 19.5, 11.0)
    Canvas.setFont(MpFontBold, PctSize)
    Canvas.drawCentredString(CX, CY - 2.6 * mm, _PdfText(PercentText))
    _MpTracked(Canvas, f"Score {ScoreText}", CX, CY - 7.4 * mm, 5.4, MpMuted, 0.5, "center")

    # Sparkles.
    _MpStar(Canvas, CX - RingR - 1.5 * mm, CY + RingR * 0.62, 2.0 * mm, MpGold, 0.32, 4)
    _MpStar(Canvas, CX + RingR + 1.0 * mm, CY + RingR * 0.30, 1.5 * mm, MpCyan, 0.32, 4)
    _MpStar(Canvas, CX + RingR * 0.66, CY + RingR + 2.2 * mm, 1.7 * mm, MpPurple, 0.32, 4)

    # Ribbon banner with the performance band.
    Label = _PdfText(Tier["label"]).upper()
    LabelSize = 6.6
    BannerW = min(52 * mm, max(34 * mm, Canvas.stringWidth(Label, MpFontBold, LabelSize) + 0.9 * (len(Label) - 1) + 12 * mm))
    BannerH = 7.6 * mm
    BannerX = CX - BannerW / 2
    BannerY = CY - RingR - 9.6 * mm
    for Direction in (-1, 1):
        EdgeX = BannerX if Direction < 0 else BannerX + BannerW
        TailPath = Canvas.beginPath()
        TailPath.moveTo(EdgeX + Direction * -2, BannerY + 1.2)
        TailPath.lineTo(EdgeX + Direction * 5.2 * mm, BannerY - 0.4)
        TailPath.lineTo(EdgeX + Direction * 3.2 * mm, BannerY + BannerH / 2 - 0.4)
        TailPath.lineTo(EdgeX + Direction * 5.2 * mm, BannerY + BannerH - 2.0)
        TailPath.lineTo(EdgeX + Direction * -2, BannerY + BannerH - 3.4)
        TailPath.close()
        Canvas.setFillColor(Tier["dark"])
        Canvas.drawPath(TailPath, stroke=0, fill=1)
    Canvas.setFillColor(Tier["main"])
    Canvas.roundRect(BannerX, BannerY, BannerW, BannerH, 2.2, fill=1, stroke=0)
    Canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.18))
    Canvas.roundRect(BannerX, BannerY + BannerH * 0.52, BannerW, BannerH * 0.48, 2.2, fill=1, stroke=0)
    _MpTracked(Canvas, Label, CX, BannerY + BannerH / 2 - 2.1, LabelSize, MpWhite, 0.9, "center")
    Canvas.restoreState()


def _MpJourneyPath(Canvas, X: float, Y: float, W: float, Nodes: list[dict[str, Any]], Tier: dict[str, Any]):
    """Level-map path: cleared levels, the freshly completed level, and the
    next step, joined by a progress line."""
    Count = len(Nodes)
    if Count == 0:
        return
    StartX = X + 14 * mm
    EndX = X + W - 14 * mm
    Positions = [StartX + (EndX - StartX) * Index / (Count - 1) for Index in range(Count)] if Count > 1 else [(StartX + EndX) / 2]
    Canvas.saveState()
    Canvas.setLineCap(1)
    for Index in range(Count - 1):
        NextState = Nodes[Index + 1].get("state")
        Canvas.setLineWidth(1.6)
        if NextState == "next":
            Canvas.setStrokeColor(MpFaint)
            Canvas.setDash([2.6, 2.6], 0)
        else:
            Canvas.setStrokeColor(MpGreen)
            Canvas.setDash([])
        Canvas.line(Positions[Index] + 5.2 * mm, Y, Positions[Index + 1] - 5.2 * mm, Y)
    Canvas.setDash([])

    for Index, Node in enumerate(Nodes):
        PosX = Positions[Index]
        State = Node.get("state")
        Code = _PdfText(Node.get("code", "-"))
        if State == "done":
            Canvas.setFillColor(MpGreen)
            Canvas.circle(PosX, Y, 3.4 * mm, fill=1, stroke=0)
            _MpGlyph(Canvas, "check", PosX, Y, 2.4 * mm, MpWhite, 1.15)
            StatusText, StatusColor = "Cleared", MpGreen
        elif State == "current":
            Canvas.setStrokeColor(Tier["soft"])
            Canvas.setLineWidth(2.6)
            Canvas.circle(PosX, Y, 5.2 * mm, fill=0, stroke=1)
            Canvas.setFillColor(Tier["main"])
            Canvas.circle(PosX, Y, 4.2 * mm, fill=1, stroke=0)
            _MpGlyph(Canvas, "check", PosX, Y, 3.0 * mm, MpWhite, 1.5)
            _MpStar(Canvas, PosX + 4.6 * mm, Y + 4.8 * mm, 1.7 * mm, MpGold, 0.32, 4)
            StatusText, StatusColor = "Completed", Tier["dark"]
        else:
            Canvas.setFillColor(MpWhite)
            Canvas.setStrokeColor(MpBlue)
            Canvas.setLineWidth(1.1)
            Canvas.setDash([2.2, 2.0], 0)
            Canvas.circle(PosX, Y, 3.8 * mm, fill=1, stroke=1)
            Canvas.setDash([])
            _MpGlyph(Canvas, "flag", PosX + 0.4 * mm, Y, 2.3 * mm, MpBlue, 1.0)
            StatusText, StatusColor = Node.get("statusOverride") or "Next Up", MpBlue
        _MpTracked(Canvas, StatusText, PosX, Y + 6.8 * mm, 5.4, StatusColor, 0.8, "center")
        _MpDrawFitted(Canvas, Code, PosX, Y - 9.6 * mm, 30 * mm, 8.2, MpInk, True, "center", 6.0)
    Canvas.restoreState()


def _MpParentProgressCopy(PercentageValue: str, FirstName: str, ReportLevel: str, NextLevel: str) -> dict[str, str]:
    Numeric = _MpNum(PercentageValue)
    NextLevelValue = str(NextLevel or "").strip()
    NextLevelAvailable = NextLevelValue not in {"", "-", "Next Level", "Next Level Pending Setup"}

    if Numeric >= 90:
        if not NextLevelAvailable:
            return {
                "takeaway": f"{FirstName} has completed {ReportLevel} with excellent assessment performance. The next structured MathPath level is pending setup.",
                "intro": f"Use this strong milestone to keep {FirstName} motivated while the next learning step is prepared.",
                "celebrate": f"Recognize {FirstName}'s excellent completion of {ReportLevel} and the focus shown during the assessment milestone.",
                "nextFocus": "Keep the learning routine steady while the next MathPath level is prepared.",
                "atHome": "Keep practice short, regular, and encouraging so confidence and consistency continue together.",
                "note": f"{FirstName}'s report records the completed level, strong assessment outcome, practice performance, and next structured learning step status.",
            }
        return {
            "takeaway": f"{FirstName} has completed {ReportLevel} with excellent assessment performance and is ready for the next structured MathPath step.",
            "intro": f"Use this strong milestone to keep {FirstName} motivated while maintaining a steady and balanced learning routine.",
            "celebrate": f"Recognize {FirstName}'s excellent completion of {ReportLevel} and the focus shown during the assessment milestone.",
            "nextFocus": f"Begin {NextLevel} with confidence while continuing the same accuracy-first learning habits.",
            "atHome": "Keep practice short, regular, and encouraging so confidence and consistency continue together.",
            "note": f"{FirstName}'s report records the completed level, strong assessment outcome, practice performance, and next structured learning step.",
        }

    if not NextLevelAvailable:
        return {
            "takeaway": f"{FirstName} has completed {ReportLevel} and cleared the assessment milestone. The next structured MathPath level is pending setup.",
            "intro": f"Use this completed-level milestone to support {FirstName} with calm, regular practice while the next step is prepared.",
            "celebrate": f"Appreciate {FirstName}'s completion of {ReportLevel} and the effort shown in clearing the assessment milestone.",
            "nextFocus": "Maintain steady accuracy habits while the next MathPath level is prepared.",
            "atHome": "Encourage steady practice, avoid pressure for speed, and give positive feedback for careful work.",
            "note": f"{FirstName}'s report records the completed level, assessment clearance, practice performance, and next structured learning step status.",
        }

    return {
        "takeaway": f"{FirstName} has completed {ReportLevel}, cleared the assessment milestone, and is ready to begin the next structured MathPath step.",
        "intro": f"Use this completed-level milestone to support {FirstName} with calm, regular practice and steady confidence-building.",
        "celebrate": f"Appreciate {FirstName}'s completion of {ReportLevel} and the effort shown in clearing the assessment milestone.",
        "nextFocus": f"Start {NextLevel} gradually while continuing to strengthen accuracy and confidence.",
        "atHome": "Encourage steady practice, avoid pressure for speed, and give positive feedback for careful work.",
        "note": f"{FirstName}'s report records the completed level, assessment clearance, practice performance, and next structured learning step.",
    }


def BuildParentProgressPdfResponse(FileName: str, ReportData: dict[str, Any]) -> StreamingResponse:
    Student = ReportData.get("student", {}) or {}
    Report = ReportData.get("report", {}) or {}
    Performance = ReportData.get("performance", {}) or {}
    Summary = ReportData.get("summary", {}) or {}
    Movements = list(ReportData.get("movements", []) or [])
    Levels = list(ReportData.get("levels", []) or [])
    Journey = ReportData.get("journey", {}) or {}

    StudentName = _PdfText(Student.get("name", "-"))
    StudentCode = _PdfText(Student.get("code", "-"))
    ClassSection = _PdfText(Student.get("classSection", "-"))
    GeneratedOn = _PdfText(ReportData.get("generatedOn") or Report.get("generatedOn") or ReportGeneratedOn())
    ReportLevel = _PdfText(Report.get("reportLevelCode", "-"))
    ReportLevelName = _PdfText(Report.get("reportLevelName") or ReportLevel)
    ReportModuleCode = _PdfText(Report.get("reportModuleCode", "-"))
    ReportModuleName = _PdfText(Report.get("reportModuleName") or Report.get("reportModuleCode") or "-")
    ModuleDisplayNames = {"YLM": "Young Learners Module"}
    if ReportModuleName == ReportModuleCode and ReportModuleCode in ModuleDisplayNames:
        ReportModuleName = ModuleDisplayNames[ReportModuleCode]
    if not ReportModuleName or ReportModuleName == "-":
        ReportModuleName = ModuleDisplayNames.get(ReportModuleCode, ReportModuleCode or "Learning Journey")

    NextLevel = _PdfText(Report.get("nextLevelCode", "-"))
    NextLevelAvailable = NextLevel not in {"", "-", "Next Level", "Next Level Pending Setup"}
    AssessmentName = _PdfText(Performance.get("assessmentName", "Level Assessment"))
    AssessmentScore = _PdfText(Performance.get("assessmentScore", "-"))
    AssessmentPercentage = _PdfText(Performance.get("assessmentPercentage", "-"))
    AssessmentResult = _PdfText(Performance.get("assessmentResult", "Assessment Milestone Cleared"))
    AssessmentDate = _PdfText(Performance.get("assessmentDate", "-"))
    PracticeAccuracy = _PdfText(Performance.get("practiceAccuracy", "-"))
    PracticeCompleted = Performance.get("practiceCompleted")
    PracticeTotal = Performance.get("practiceTotal")
    if PracticeCompleted is None or PracticeTotal is None:
        ProgressText = _PdfText(Performance.get("practiceProgress", "-"))
        Portions = ProgressText.replace("Practice Sheets", "").strip().split("/")
        try:
            PracticeCompleted = int(Portions[0].strip())
            PracticeTotal = int(Portions[1].strip()) if len(Portions) > 1 else 0
        except Exception:
            PracticeCompleted, PracticeTotal = 0, 0
    PracticeSheetsText = f"{PracticeCompleted} / {PracticeTotal}" if PracticeTotal else f"{PracticeCompleted}"

    Message = _PdfText(Summary.get("message", "The student's progress summary is ready for review."))
    NextStep = _PdfText(Summary.get("nextStep", f"Begin {NextLevel} Practice"))
    FirstName = StudentName.split()[0] if StudentName and StudentName != "-" else "The student"
    LogoPath = _FindMathPathLogo()
    Tier = _MpTier(AssessmentPercentage)
    Copy = _MpParentProgressCopy(AssessmentPercentage, FirstName, ReportLevel, NextLevel)
    PhotoReader = _MpResolvePhoto(Student.get("photoRef") or Student.get("photoDataUrl") or Student.get("photoUrl") or Student.get("profilePhotoUrl"))

    # Journey aggregates (with graceful fallbacks when the caller predates them).
    PracticeAvg = Journey.get("practiceAverageAccuracy")
    AssessmentAvg = Journey.get("assessmentAverageAccuracy")
    OverallAvg = Journey.get("overallAverageAccuracy")
    if PracticeAvg is None:
        PracticeAvg = _MpNum(PracticeAccuracy)
    if AssessmentAvg is None:
        AssessmentAvg = _MpNum(AssessmentPercentage)
    if OverallAvg is None:
        OverallAvg = round((_MpNum(PracticeAvg) + _MpNum(AssessmentAvg)) / 2)
    AssessmentsCleared = Journey.get("assessmentsCleared")
    if AssessmentsCleared is None:
        AssessmentsCleared = 1 if "Cleared" in AssessmentResult else 0
    PromotedLevels = Journey.get("promotedLevels")
    if PromotedLevels is None:
        PromotedLevels = len(Movements)

    # Level breakdown fallback: synthesize the report level row from performance data.
    if not Levels:
        AccuracyValue = _MpNum(PracticeAccuracy)
        Zone = "Excellence Zone" if AccuracyValue >= 90 else "Growth Zone" if AccuracyValue >= 70 else "Needs Improvement" if PracticeCompleted else "Not Started"
        Levels = [{
            "levelCode": ReportLevel,
            "levelName": ReportLevelName,
            "requiredDps": PracticeTotal or 0,
            "completedDps": PracticeCompleted or 0,
            "averageAccuracy": int(AccuracyValue),
            "performanceZone": Zone,
            "promotionStatus": "Promoted" if Movements else "In Progress",
            "isReportLevel": True,
        }]

    Buffer = BytesIO()
    Pdf = PdfCanvas.Canvas(Buffer, pagesize=A4)
    PageW, PageH = A4
    L = 14 * mm
    R = PageW - 14 * mm
    CW = R - L

    # ------------------------------------------------------------------ Page 1
    _MpBackground(Pdf, 1)

    # Header band.
    BandH = 42 * mm
    Pdf.saveState()
    Pdf.setFillColor(MpInk)
    Pdf.rect(0, PageH - BandH, PageW, BandH, fill=1, stroke=0)
    Pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.05))
    Pdf.circle(PageW - 26 * mm, PageH - 6 * mm, 26 * mm, fill=1, stroke=0)
    Pdf.circle(PageW - 74 * mm, PageH - 40 * mm, 17 * mm, fill=1, stroke=0)
    Pdf.circle(8 * mm, PageH - 40 * mm, 13 * mm, fill=1, stroke=0)
    Pdf.setFillColor(colors.Color(1, 1, 1, alpha=0.09))
    Pdf.setFont(MpFontBold, 11)
    for Glyph, GX, GY in [("+", PageW - 40 * mm, PageH - 10 * mm), ("x", PageW - 96 * mm, PageH - 8 * mm), ("=", PageW - 12 * mm, PageH - 30 * mm), ("+", PageW - 118 * mm, PageH - 36 * mm)]:
        Pdf.drawString(GX, GY, Glyph)
    Pdf.restoreState()
    _MpLogoChip(Pdf, LogoPath, L, PageH - 31 * mm, 42 * mm, 16 * mm, Shadow=False)
    _MpTracked(Pdf, "MathPath · Official Learning Record", R, PageH - 13.5 * mm, 6.6, MpCyanLight, 1.1, "right")
    Pdf.setFillColor(MpWhite)
    Pdf.setFont(MpFontBold, 19)
    Pdf.drawRightString(R, PageH - 21.5 * mm, "Parent Progress Report")
    Pdf.setFillColor(colors.HexColor("#AAB8D9"))
    Pdf.setFont(MpFontRegular, 7.6)
    Pdf.drawRightString(R, PageH - 27.5 * mm, f"Generated {GeneratedOn}")

    # Identity card.
    CardH = 32 * mm
    CardY = PageH - 47 * mm - CardH
    _MpPanel(Pdf, L, CardY, CW, CardH, MpWhite, MpLine, 12, 0.85, Shadow=True)
    _MpAvatar(Pdf, L + 16 * mm, CardY + CardH / 2, 11 * mm, PhotoReader, StudentName, Tier)
    NameX = L + 31 * mm
    NameW = 62 * mm
    _MpDrawFitted(Pdf, StudentName, NameX, CardY + 20.5 * mm, NameW, 14.5, MpInk, True, "left", 9.0)
    Pdf.setFillColor(MpMuted)
    Pdf.setFont(MpFontRegular, 7.8)
    Pdf.drawString(NameX, CardY + 14.6 * mm, f"{StudentCode} · Class {ClassSection}")
    ModuleLabel = f"MODULE · {ReportModuleName}".upper()
    ModuleSize = _MpFit(Pdf, ModuleLabel, NameW - 10 * mm, 6.6, 5.2)
    ModuleW = Pdf.stringWidth(ModuleLabel, MpFontBold, ModuleSize) + 9 * mm
    Pdf.setFillColor(MpBlueSoft)
    Pdf.roundRect(NameX, CardY + 6.2 * mm, ModuleW, 6.4 * mm, 3.2 * mm, fill=1, stroke=0)
    Pdf.setFillColor(MpBlueDark)
    Pdf.setFont(MpFontBold, ModuleSize)
    Pdf.drawString(NameX + 4.5 * mm, CardY + 8.3 * mm, ModuleLabel)

    ChipW = 42 * mm
    ChipH = 21 * mm
    ChipY = CardY + (CardH - ChipH) / 2
    CompletedX = R - ChipW * 2 - 12 * mm
    _MpPanel(Pdf, CompletedX, ChipY, ChipW, ChipH, Tier["soft"], Tier["main"], 9, 0.9)
    _MpTracked(Pdf, "Completed Level", CompletedX + ChipW / 2, ChipY + ChipH - 6.4 * mm, 5.6, Tier["dark"], 0.8, "center")
    _MpDrawFitted(Pdf, ReportLevel, CompletedX + ChipW / 2, ChipY + 6.4 * mm, ChipW - 12 * mm, 13.0, MpInk, True, "center", 7.5)
    _MpGlyph(Pdf, "check", CompletedX + 5.4 * mm, ChipY + ChipH - 5.2 * mm, 2.0 * mm, Tier["dark"], 1.1)
    NextX = R - ChipW - 6 * mm
    Pdf.saveState()
    Pdf.setDash([2.6, 2.2], 0)
    _MpPanel(Pdf, NextX, ChipY, ChipW, ChipH, MpWhite, MpBlue, 9, 0.9)
    Pdf.restoreState()
    _MpTracked(Pdf, "Next Level", NextX + ChipW / 2, ChipY + ChipH - 6.4 * mm, 5.6, MpBlueDark, 0.8, "center")
    _MpDrawFitted(Pdf, NextLevel if NextLevelAvailable else "Pending Setup", NextX + ChipW / 2, ChipY + 6.4 * mm, ChipW - 10 * mm, 13.0, MpBlue, True, "center", 6.6)

    # Achievement section: medallion left, milestone story right.
    SectTop = CardY - 7 * mm
    SectH = 70 * mm
    SectY = SectTop - SectH
    BadgeCX = L + 30 * mm
    BadgeCY = SectY + 40 * mm
    _MpAchievementBadge(Pdf, BadgeCX, BadgeCY, AssessmentPercentage, AssessmentScore, Tier)

    StoryX = L + 64 * mm
    StoryW = R - StoryX
    _MpTracked(Pdf, "Assessment Milestone", StoryX, SectY + 63.5 * mm, 6.8, MpCyan, 1.1)
    _MpDrawFitted(Pdf, f"{FirstName} Completed {ReportLevelName}", StoryX, SectY + 56.5 * mm, StoryW, 15.5, MpInk, True, "left", 10.5)
    _MpPara(Pdf, Message, StoryX, SectY + 53 * mm, StoryW, _MpStyle("MpStory", 9.0, 12.8, MpText))
    FactRows = [
        ("doc", "Assessment", AssessmentName, MpBlue, MpBlueSoft),
        ("calendar", "Completed On", AssessmentDate, MpTeal, MpTealSoft),
        ("medal", "Result", AssessmentResult, MpGoldDark, MpGoldSoft),
    ]
    FactY = SectY + 26 * mm
    for Glyph, Label, Value, GlyphColor, DiscFill in FactRows:
        Pdf.setFillColor(DiscFill)
        Pdf.circle(StoryX + 3.4 * mm, FactY + 2.6 * mm, 3.4 * mm, fill=1, stroke=0)
        _MpGlyph(Pdf, Glyph, StoryX + 3.4 * mm, FactY + 2.6 * mm, 2.4 * mm, GlyphColor, 1.0)
        _MpTracked(Pdf, Label, StoryX + 9.4 * mm, FactY + 4.0 * mm, 5.4, MpMuted, 0.8)
        _MpDrawFitted(Pdf, Value, StoryX + 9.4 * mm, FactY - 0.4 * mm, StoryW - 11 * mm, 9.0, MpInk, True, "left", 6.6)
        FactY -= 9.4 * mm

    # Learning journey path.
    JourneyH = 34 * mm
    JourneyTop = SectY - 7 * mm
    JourneyY = JourneyTop - JourneyH
    _MpPanel(Pdf, L, JourneyY, CW, JourneyH, MpWhite, MpLine, 12, 0.85, Shadow=True)
    _MpTracked(Pdf, "Learning Journey Path", L + 8 * mm, JourneyY + JourneyH - 7 * mm, 6.4, MpCyan, 1.1)
    Pdf.setFillColor(MpFaint)
    Pdf.setFont(MpFontRegular, 6.6)
    Pdf.drawRightString(R - 8 * mm, JourneyY + JourneyH - 7 * mm, "Levels recorded in this report scope")
    PathNodes: list[dict[str, Any]] = []
    SeenCodes: set[str] = set()
    for Movement in reversed(Movements):
        Code = _PdfText(Movement.get("fromLevel", ""))
        if Code and Code not in {"-", ReportLevel} and Code not in SeenCodes:
            PathNodes.append({"code": Code, "state": "done"})
            SeenCodes.add(Code)
    PathNodes = PathNodes[-3:]
    PathNodes.append({"code": ReportLevel, "state": "current"})
    PathNodes.append({
        "code": NextLevel if NextLevelAvailable else "TBD",
        "state": "next",
        "statusOverride": None if NextLevelAvailable else "Pending Setup",
    })
    _MpJourneyPath(Pdf, L, JourneyY + 15.5 * mm, CW, PathNodes, Tier)

    # Stats band: journey counters (no accuracy repeats from page 2).
    StatH = 20 * mm
    StatTop = JourneyY - 7 * mm
    StatY = StatTop - StatH
    StatGap = 5 * mm
    StatW = (CW - StatGap * 3) / 4
    StatItems = [
        ("doc", "Practice Sheets", PracticeSheetsText, MpPurple, MpPurpleSoft),
        ("target", "Practice Accuracy", PracticeAccuracy, MpBlue, MpBlueSoft),
        ("medal", "Assessments Cleared", str(AssessmentsCleared), MpGoldDark, MpGoldSoft),
        ("levelup", "Levels Promoted", str(PromotedLevels), MpGreen, MpGreenSoft),
    ]
    for Index, (Glyph, Label, Value, GlyphColor, DiscFill) in enumerate(StatItems):
        StatX = L + Index * (StatW + StatGap)
        _MpPanel(Pdf, StatX, StatY, StatW, StatH, MpWhite, MpLine, 10, 0.75, Shadow=True)
        Pdf.setFillColor(DiscFill)
        Pdf.circle(StatX + 8.4 * mm, StatY + StatH / 2, 4.6 * mm, fill=1, stroke=0)
        _MpGlyph(Pdf, Glyph, StatX + 8.4 * mm, StatY + StatH / 2, 3.0 * mm, GlyphColor, 1.15)
        _MpTracked(Pdf, Label, StatX + 15.4 * mm, StatY + StatH - 7.6 * mm, 5.2, MpMuted, 0.55)
        _MpDrawFitted(Pdf, Value, StatX + 15.4 * mm, StatY + 4.6 * mm, StatW - 18.4 * mm, 12.5, MpInk, True, "left", 7.0)

    # Parent takeaway with next-step action chip.
    TakeH = 26 * mm
    TakeTop = StatY - 7 * mm
    TakeY = TakeTop - TakeH
    _MpPanel(Pdf, L, TakeY, CW, TakeH, MpTealSoft, colors.HexColor("#BFE8E0"), 12, 0.85)
    Pdf.setFillColor(MpTeal)
    Pdf.roundRect(L, TakeY, 2.6 * mm, TakeH, 1.3 * mm, fill=1, stroke=0)
    _MpGlyph(Pdf, "spark", L + 8.8 * mm, TakeY + TakeH - 6.6 * mm, 2.2 * mm, MpTeal)
    _MpTracked(Pdf, "Parent Takeaway", L + 13 * mm, TakeY + TakeH - 7.6 * mm, 6.4, MpTeal, 1.1)
    ActionW = 48 * mm
    _MpPara(Pdf, Copy["takeaway"], L + 9 * mm, TakeY + TakeH - 11 * mm, CW - ActionW - 22 * mm, _MpStyle("MpTakeaway", 8.8, 12.0, MpText))
    ActionX = R - ActionW - 6 * mm
    ActionY = TakeY + (TakeH - 13 * mm) / 2
    _MpPanel(Pdf, ActionX, ActionY, ActionW, 13 * mm, MpBlue, MpBlueDark, 6.5 * mm, 0.9)
    _MpTracked(Pdf, "Next Step", ActionX + ActionW / 2 + 2.2 * mm, ActionY + 8.0 * mm, 5.0, colors.HexColor("#BFDBFE"), 0.9, "center")
    _MpDrawFitted(Pdf, NextStep, ActionX + ActionW / 2 + 2.2 * mm, ActionY + 3.2 * mm, ActionW - 14 * mm, 8.2, MpWhite, True, "center", 5.8)
    _MpGlyph(Pdf, "arrow", ActionX + 6.0 * mm, ActionY + 6.5 * mm, 2.5 * mm, MpWhite, 1.3)

    _MpTracked(Pdf, "Continued on page 2 · Accuracy profile · Level mastery · Progression timeline · Home guidance", L + CW / 2, TakeY - 9 * mm, 5.6, MpFaint, 0.9, "center")
    _MpFooter(Pdf, 1, 2, StudentCode, ReportLevel)
    Pdf.showPage()

    # ------------------------------------------------------------------ Page 2
    _MpBackground(Pdf, 2)
    _MpLogoChip(Pdf, LogoPath, L, PageH - 27 * mm, 36 * mm, 13 * mm, Shadow=True)
    Pdf.setFillColor(MpInk)
    Pdf.setFont(MpFontBold, 13.5)
    Pdf.drawRightString(R, PageH - 19.5 * mm, "Detailed Performance Review")
    Pdf.setFillColor(MpMuted)
    Pdf.setFont(MpFontRegular, 7.2)
    Pdf.drawRightString(R, PageH - 24.5 * mm, f"{StudentName} · {StudentCode} · Generated {GeneratedOn}")
    Pdf.setStrokeColor(MpLine)
    Pdf.setLineWidth(0.8)
    Pdf.line(L, PageH - 31 * mm, R, PageH - 31 * mm)

    Cursor = PageH - 38 * mm

    def SectionHeading(Eyebrow: str, Title: str, Note: str | None = None) -> float:
        nonlocal Cursor
        _MpTracked(Pdf, Eyebrow, L, Cursor, 6.4, MpCyan, 1.1)
        Pdf.setFillColor(MpInk)
        Pdf.setFont(MpFontBold, 12.5)
        Pdf.drawString(L, Cursor - 6.4 * mm, _PdfText(Title))
        if Note:
            Pdf.setFillColor(MpFaint)
            Pdf.setFont(MpFontRegular, 6.8)
            Pdf.drawRightString(R, Cursor - 6.4 * mm, _PdfText(Note))
        Cursor -= 11.5 * mm
        return Cursor

    # Section A: accuracy profile.
    SectionHeading("Accuracy Profile", "How Accuracy Is Building", "Averages across all recorded attempts")
    ProfileH = 38 * mm
    ProfileY = Cursor - ProfileH
    _MpPanel(Pdf, L, ProfileY, CW, ProfileH, MpWhite, MpLine, 12, 0.85, Shadow=True)
    ProfileRows = [
        ("Practice", _MpNum(PracticeAvg), MpPurple),
        ("Assessments", _MpNum(AssessmentAvg), MpBlue),
        ("Overall", _MpNum(OverallAvg), MpTeal),
    ]
    RowY = ProfileY + ProfileH - 9.2 * mm
    BarX = L + 36 * mm
    BarW = CW - 36 * mm - 24 * mm
    for Label, Value, Color in ProfileRows:
        _MpTracked(Pdf, Label, L + 8 * mm, RowY - 1.2 * mm, 6.0, MpMuted, 0.8)
        _MpBar(Pdf, BarX, RowY - 1.8 * mm, BarW, 3.4 * mm, Value / 100.0, Color)
        _MpDrawFitted(Pdf, f"{int(round(Value))}%", R - 7 * mm, RowY - 1.6 * mm, 14 * mm, 9.5, MpInk, True, "right")
        RowY -= 9.2 * mm
    Pdf.setFillColor(MpFaint)
    Pdf.setFont(MpFontRegular, 6.4)
    Pdf.drawString(L + 8 * mm, ProfileY + 3.8 * mm, "Practice reflects daily practice sheets; Assessments reflect level assessment attempts recorded in this report scope.")
    Cursor = ProfileY - 7 * mm

    # Section B: level-by-level mastery.
    VisibleLevels = Levels[:4]
    HiddenLevels = max(0, len(Levels) - len(VisibleLevels))
    SectionHeading("Level Mastery", "Level-By-Level Breakdown", f"{len(Levels)} level(s) tracked")
    HeaderBandH = 7 * mm
    LevelRowH = 12.5 * mm
    MasteryH = HeaderBandH + LevelRowH * len(VisibleLevels) + (5 * mm if HiddenLevels else 0) + 4 * mm
    MasteryY = Cursor - MasteryH
    _MpPanel(Pdf, L, MasteryY, CW, MasteryH, MpWhite, MpLine, 12, 0.85, Shadow=True)
    ColLevel = L + 8 * mm
    ColPractice = L + 52 * mm
    ColAccuracy = L + 104 * mm
    ColZone = L + 148 * mm
    CaptionY = MasteryY + MasteryH - 5.4 * mm
    for Caption, CapX in [("Level", ColLevel), ("Practice Sheets", ColPractice), ("Accuracy", ColAccuracy), ("Zone", ColZone)]:
        _MpTracked(Pdf, Caption, CapX, CaptionY, 5.4, MpFaint, 0.8)
    RowTop = MasteryY + MasteryH - HeaderBandH - 2 * mm
    for Index, LevelRow in enumerate(VisibleLevels):
        RowBase = RowTop - LevelRowH * (Index + 1) + 2 * mm
        IsReportLevel = bool(LevelRow.get("isReportLevel")) or _PdfText(LevelRow.get("levelCode")) == ReportLevel
        if IsReportLevel:
            Pdf.setFillColor(Tier["soft"])
            Pdf.roundRect(L + 3 * mm, RowBase - 1.4 * mm, CW - 6 * mm, LevelRowH - 1.2 * mm, 5, fill=1, stroke=0)
            Pdf.setFillColor(Tier["main"])
            Pdf.roundRect(L + 3 * mm, RowBase - 1.4 * mm, 2.0 * mm, LevelRowH - 1.2 * mm, 1.0 * mm, fill=1, stroke=0)
        elif Index % 2 == 1:
            Pdf.setFillColor(MpSlateSoft)
            Pdf.roundRect(L + 3 * mm, RowBase - 1.4 * mm, CW - 6 * mm, LevelRowH - 1.2 * mm, 5, fill=1, stroke=0)
        CodeText = _PdfText(LevelRow.get("levelCode", "-"))
        CodeSize = _MpDrawFitted(Pdf, CodeText, ColLevel, RowBase + 4.6 * mm, 30 * mm, 9.2, MpInk, True, "left", 6.6)
        if str(LevelRow.get("promotionStatus") or "") == "Promoted":
            CheckX = ColLevel + Pdf.stringWidth(CodeText, MpFontBold, CodeSize) + 3.2 * mm
            Pdf.setFillColor(MpGreenSoft)
            Pdf.circle(CheckX, RowBase + 5.6 * mm, 2.0 * mm, fill=1, stroke=0)
            _MpGlyph(Pdf, "check", CheckX, RowBase + 5.6 * mm, 1.4 * mm, MpGreen, 0.9)
        LevelNameText = _PdfText(LevelRow.get("levelName") or "")
        if LevelNameText and LevelNameText != CodeText:
            _MpDrawFitted(Pdf, LevelNameText, ColLevel, RowBase + 0.8 * mm, 40 * mm, 6.4, MpMuted, False, "left", 5.2)
        Required = int(_MpNum(LevelRow.get("requiredDps")))
        Completed = int(_MpNum(LevelRow.get("completedDps")))
        PracticeFrac = (Completed / Required) if Required else 0.0
        _MpBar(Pdf, ColPractice, RowBase + 3.2 * mm, 30 * mm, 3.2 * mm, PracticeFrac, MpPurple)
        Pdf.setFillColor(MpMuted)
        Pdf.setFont(MpFontRegular, 6.8)
        Pdf.drawString(ColPractice + 32 * mm, RowBase + 3.4 * mm, f"{Completed} / {Required}" if Required else f"{Completed}")
        AccuracyValue = _MpNum(LevelRow.get("averageAccuracy"))
        _MpBar(Pdf, ColAccuracy, RowBase + 3.2 * mm, 26 * mm, 3.2 * mm, AccuracyValue / 100.0, MpBlue)
        Pdf.setFillColor(MpInk)
        Pdf.setFont(MpFontBold, 7.2)
        Pdf.drawString(ColAccuracy + 28 * mm, RowBase + 3.3 * mm, f"{int(round(AccuracyValue))}%")
        ZoneFill, ZoneText = _MpZoneStyle(LevelRow.get("performanceZone"))
        _MpChip(Pdf, LevelRow.get("performanceZone") or "Not Started", ColZone + 15 * mm, RowBase + 4.6 * mm, ZoneFill, ZoneText, 5.8, 4.6, 9.6)
    if HiddenLevels:
        Pdf.setFillColor(MpFaint)
        Pdf.setFont(MpFontRegular, 6.4)
        Pdf.drawCentredString(L + CW / 2, MasteryY + 3.0 * mm, f"+ {HiddenLevels} more level(s) tracked in the full learning record")
    Cursor = MasteryY - 7 * mm

    # Section C: progression timeline.
    VisibleMovements = Movements[:3]
    HiddenMovements = max(0, len(Movements) - len(VisibleMovements))
    SectionHeading("Progression Timeline", "Completed Level Movements", "Most recent first")
    MoveRowH = 10 * mm
    TimelineH = (MoveRowH * len(VisibleMovements) + (5 * mm if HiddenMovements else 0) + 6 * mm) if VisibleMovements else 16 * mm
    TimelineY = Cursor - TimelineH
    _MpPanel(Pdf, L, TimelineY, CW, TimelineH, MpWhite, MpLine, 12, 0.85, Shadow=True)
    if VisibleMovements:
        DotX = L + 40 * mm
        if len(VisibleMovements) > 1:
            Pdf.setStrokeColor(MpLine)
            Pdf.setLineWidth(1.1)
            Pdf.line(DotX, TimelineY + TimelineH - 3 * mm - MoveRowH / 2, DotX, TimelineY + TimelineH - 3 * mm - MoveRowH * (len(VisibleMovements) - 1) - MoveRowH / 2)
        for Index, Movement in enumerate(VisibleMovements):
            RowMid = TimelineY + TimelineH - 3 * mm - MoveRowH * Index - MoveRowH / 2
            Pdf.setFillColor(MpMuted)
            Pdf.setFont(MpFontRegular, 6.6)
            DateText = _PdfText(Movement.get("date", "-"))
            Pdf.drawString(L + 7 * mm, RowMid - 1.0 * mm, DateText[:24])
            Pdf.setFillColor(MpGreen if Index else Tier["main"])
            Pdf.circle(DotX, RowMid, 1.9 * mm, fill=1, stroke=0)
            Pdf.setFillColor(MpWhite)
            Pdf.circle(DotX, RowMid, 0.75 * mm, fill=1, stroke=0)
            FromText = _PdfText(Movement.get("fromLevel", "-"))
            ToText = _PdfText(Movement.get("toLevel", "-"))
            TextX = DotX + 5.5 * mm
            Pdf.setFillColor(MpInk)
            Pdf.setFont(MpFontBold, 8.6)
            Pdf.drawString(TextX, RowMid + 0.6 * mm, FromText)
            FromW = Pdf.stringWidth(FromText, MpFontBold, 8.6)
            _MpGlyph(Pdf, "arrow", TextX + FromW + 4.6 * mm, RowMid + 1.6 * mm, 2.2 * mm, MpMuted, 1.0)
            Pdf.setFillColor(MpInk)
            Pdf.setFont(MpFontBold, 8.6)
            Pdf.drawString(TextX + FromW + 9.6 * mm, RowMid + 0.6 * mm, ToText)
            _MpDrawFitted(Pdf, Movement.get("assessment", "-"), TextX, RowMid - 3.6 * mm, 62 * mm, 6.4, MpMuted, False, "left", 5.2)
            ScoreLabel = f"{_PdfText(Movement.get('score', '-'))} · {_PdfText(Movement.get('percentage', '-'))}"
            _MpChip(Pdf, ScoreLabel, R - 24 * mm, RowMid, Tier["soft"] if Index == 0 else MpSlateSoft, Tier["dark"] if Index == 0 else MpMuted, 6.2, 5.0, 10.0)
        if HiddenMovements:
            Pdf.setFillColor(MpFaint)
            Pdf.setFont(MpFontRegular, 6.4)
            Pdf.drawCentredString(L + CW / 2, TimelineY + 2.6 * mm, f"+ {HiddenMovements} earlier movement(s) in the full promotion history")
    else:
        _MpGlyph(Pdf, "flag", L + CW / 2 - 46 * mm, TimelineY + TimelineH / 2, 2.4 * mm, MpFaint, 1.0)
        Pdf.setFillColor(MpMuted)
        Pdf.setFont(MpFontRegular, 7.6)
        Pdf.drawCentredString(L + CW / 2 + 2 * mm, TimelineY + TimelineH / 2 - 1.2 * mm, "The first level movement will be recorded here after the next promotion.")
    Cursor = TimelineY - 7 * mm

    # Section D: guidance for home.
    GuidanceNeed = 11.5 * mm + 27 * mm + 3.5 * mm + 14 * mm
    if Cursor - GuidanceNeed < 16 * mm:
        Cursor = GuidanceNeed + 16 * mm
    SectionHeading("Guidance For Home", f"How To Support {FirstName}")
    CardH2 = 27 * mm
    CardGap = 5 * mm
    CardW2 = (CW - CardGap * 2) / 3
    CardsY = Cursor - CardH2
    GuidanceCards = [
        ("spark", "Celebrate", Copy["celebrate"], MpGoldDark, MpGoldSoft, MpGold),
        ("target", "Next Focus", Copy["nextFocus"], MpBlueDark, MpBlueSoft, MpBlue),
        ("home", "At Home", Copy["atHome"], MpTeal, MpTealSoft, MpTeal),
    ]
    for Index, (Glyph, Title, Body, GlyphColor, DiscFill, AccentColor) in enumerate(GuidanceCards):
        GX = L + Index * (CardW2 + CardGap)
        _MpPanel(Pdf, GX, CardsY, CardW2, CardH2, MpWhite, MpLine, 10, 0.75, Shadow=True)
        Pdf.setFillColor(AccentColor)
        Pdf.roundRect(GX, CardsY + CardH2 - 2.2 * mm, CardW2, 2.2 * mm, 1.1 * mm, fill=1, stroke=0)
        Pdf.setFillColor(DiscFill)
        Pdf.circle(GX + 7.2 * mm, CardsY + CardH2 - 8.4 * mm, 3.2 * mm, fill=1, stroke=0)
        _MpGlyph(Pdf, Glyph, GX + 7.2 * mm, CardsY + CardH2 - 8.4 * mm, 2.1 * mm, GlyphColor, 1.0)
        _MpTracked(Pdf, Title, GX + 12.4 * mm, CardsY + CardH2 - 9.4 * mm, 6.2, MpInk, 0.9)
        _MpPara(Pdf, Body, GX + 5.5 * mm, CardsY + CardH2 - 12.6 * mm, CardW2 - 11 * mm, _MpStyle(f"MpGuide{Index}", 7.4, 9.4, MpText))
    NoteY = CardsY - 3.5 * mm - 14 * mm
    _MpPanel(Pdf, L, NoteY, CW, 14 * mm, MpPurpleSoft, colors.HexColor("#DED6FE"), 9, 0.7)
    _MpTracked(Pdf, "MathPath Note", L + 7 * mm, NoteY + 8.6 * mm, 5.8, MpPurple, 1.0)
    _MpPara(Pdf, Copy["note"], L + 7 * mm, NoteY + 7.2 * mm, CW - 14 * mm, _MpStyle("MpNote", 7.2, 9.2, MpText))

    _MpFooter(Pdf, 2, 2, StudentCode, ReportLevel)
    Pdf.save()
    Buffer.seek(0)

    SafeFileName = FileName if FileName.lower().endswith(".pdf") else f"{FileName}.pdf"
    return StreamingResponse(
        Buffer,
        media_type=PDF_MIME,
        headers={"Content-Disposition": f'attachment; filename="{SafeFileName}"'},
    )


def BuildParentProgressPdfBytes(FileName: str, ReportData: dict[str, Any]) -> bytes:
    """Build the finalized parent progress PDF and return raw bytes for email attachments."""
    Response = BuildParentProgressPdfResponse(FileName, ReportData)

    async def _Collect() -> bytes:
        Chunks: list[bytes] = []
        async for Chunk in Response.body_iterator:
            if isinstance(Chunk, bytes):
                Chunks.append(Chunk)
            elif isinstance(Chunk, str):
                Chunks.append(Chunk.encode("utf-8"))
            else:
                Chunks.append(bytes(Chunk))
        return b"".join(Chunks)

    try:
        return asyncio.run(_Collect())
    except RuntimeError:
        Loop = asyncio.new_event_loop()
        try:
            return Loop.run_until_complete(_Collect())
        finally:
            Loop.close()
